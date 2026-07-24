import json
from datetime import timedelta
from decimal import ROUND_HALF_UP, Decimal, InvalidOperation

from django.conf import settings
from django.contrib import messages
from django.contrib.auth import update_session_auth_hash
from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator
from django.db import transaction
from django.db.models import (
    Case,
    Count,
    OuterRef,
    DecimalField,
    ExpressionWrapper,
    F,
    ProtectedError,
    Q,
    Subquery,
    Sum,
    Value,
    When,
)
from django.db.models.functions import Coalesce, Greatest
from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.template.loader import render_to_string
from django.urls import reverse
from django.utils import timezone
from django.utils.dateparse import parse_date
from django.utils.formats import date_format
from django.utils.text import slugify
from django.views.decorators.http import require_GET, require_POST

from .decorators import super_admin_required
from .notifications import dismiss as dismiss_notification
from .utils import get_month_filter
from .forms import (
    BillEditReasonForm,
    BillingSettingsForm,
    BillPaymentForm,
    CashDrawerEditForm,
    CashDrawerInForm,
    CashDrawerOutForm,
    CategoryForm,
    ChequeForm,
    CustomerBalanceAdjustmentForm,
    CustomerForm,
    CustomerPriceForm,
    CustomerSettlementForm,
    DailyOtherWorkForm,
    MachineForm,
    MaterialForm,
    MaterialPurchaseHeaderForm,
    MaterialSupplierForm,
    MaterialWeighEntryForm,
    OrderHeaderForm,
    PettyCashExpenseForm,
    PettyCashReimbursementForm,
    ProductForm,
    ProductionEntryForm,
    ProductQuickForm,
    ProfileDetailsForm,
    ProfilePasswordForm,
    RiderForm,
    SetUserPasswordForm,
    StockAdjustmentForm,
    SupplierQuickForm,
    UserCreateForm,
    UserEditForm,
    VehicleForm,
    VehicleTripForm,
)
from .models import (
    AuditLog,
    Bill,
    BillEditAudit,
    BillingSettings,
    BillItem,
    CashDrawer,
    CashTransfer,
    Category,
    Cheque,
    Customer,
    CustomerBalanceAdjustment,
    CustomerPrice,
    DailyMachineRun,
    DailyOtherWork,
    HeldBill,
    Machine,
    Material,
    MaterialPurchase,
    MaterialPurchaseItem,
    MaterialSupplier,
    MaterialWeighEntry,
    Order,
    OrderItem,
    Payment,
    PettyCashEntry,
    PettyCashFund,
    PettyCashReimbursement,
    Product,
    ProductionEntry,
    Rider,
    StockAdjustment,
    SupplierBill,
    SupplierBillItem,
    User,
    Vehicle,
    VehicleTrip,
    generate_password,
)

#: A cheque is "maturing soon" this many days out.
CHEQUE_WARNING_DAYS = 3

MONEY = DecimalField(max_digits=12, decimal_places=2)
ZERO = Decimal("0.00")

#: Prefix on ProductionEntry.reason for entries the bill-save path auto-created
#: to cover an oversell. _reverse_bill looks these up by prefix to undo them
#: when the bill is edited or deleted, and the stock ledger uses it to render
#: those rows distinctly. Do not change without a data migration.
OVERSALE_REASON_PREFIX = "Oversale —"


def _paginate(request, object_list, per_page=None):
    """One page of object_list, read off ?page=.

    get_page rather than page: ?page= arrives from bookmarks and hand-edited
    URLs as well as from the pager, so a missing, unparsable or out-of-range
    number lands on the nearest real page instead of raising.

    Takes a queryset or a list. A queryset is sliced in SQL and only the page
    is fetched; a list has already been built, so pass one only where the rows
    are computed in Python — a running balance has to see every earlier row,
    and cannot be worked out a page at a time.
    """
    paginator = Paginator(object_list, per_page or settings.PAGINATE_BY)
    return paginator.get_page(request.GET.get("page"))


def _is_super_admin(user):
    """Mirrors the check in `super_admin_required`, for views that stay open to
    managers but hand them a reduced form."""
    return getattr(user, "role", None) == User.Role.SUPER_ADMIN


def _warning_signature(today, cheques):
    """A key for one day's set of cheque warnings.

    The dashboard card can be dismissed, but a warning about money that hasn't
    arrived should not stay dismissed for ever. Keying the dismissal to the day
    and the exact cheques means it comes back tomorrow, and immediately if a
    different cheque starts maturing.
    """
    # Sorted, so the key identifies the set rather than the order it was listed
    # in — re-sorting the same warnings must not resurrect a dismissal.
    ids = ",".join(str(pk) for pk in sorted(cheque.pk for cheque in cheques))
    return f"{today.isoformat()}:{ids}"


def landing(request):
    """Public marketing page at /.

    Signed-in users bypass it — they've already seen the pitch and would rather
    land on the dashboard. Anonymous visitors get the marketing page with a
    Sign in call-to-action; the logout redirect also comes here so the
    just-signed-out user sees the front door rather than the login form.
    """
    if request.user.is_authenticated:
        return redirect("core:dashboard")
    modules = [
        "Products", "Categories", "Stock Ledger", "Customers",
        "Suppliers", "Bills", "Held Bills", "Cheques",
        "Cash Drawer", "Supplier Bills", "Customer Ledger", "Sales Report",
        "Production", "Daily Machine Run", "Petty Cash", "Material Purchasing",
        "Order Book", "Vehicle Tracker", "User Management", "Outstanding Report",
    ]
    return render(request, "landing.html", {"module_list": modules})


@require_POST
@login_required
def notification_dismiss(request):
    """Mark a single bell notification dismissed for the configured window.

    Expects a `key` in POST — the same stable key the notification carried
    when rendered. Fire-and-forget: the bell removes the card client-side
    before the request even completes, and this write only exists so the
    same alert doesn't come back on the next page load.
    """
    key = (request.POST.get("key") or "").strip()
    if key:
        dismiss_notification(request.session, key)
    if request.headers.get("X-Requested-With") == "XMLHttpRequest":
        return JsonResponse({"ok": True})
    return redirect(request.META.get("HTTP_REFERER") or "core:dashboard")


def _cash_drawer_balance(queryset=None):
    """Net cash on hand: 'in' adds, 'out' and 'transfer' both remove.

    Takes a queryset so the same rule can price a slice of the log — the
    opening balance of a date range is this over everything before it.
    """
    if queryset is None:
        queryset = CashDrawer.objects.all()
    return queryset.aggregate(
        total=Coalesce(
            Sum(
                Case(
                    When(txn_type=CashDrawer.TxnType.IN, then=F("amount")),
                    default=-F("amount"),
                    output_field=MONEY,
                )
            ),
            ZERO,
            output_field=MONEY,
        )
    )["total"]


@login_required
def dashboard(request):
    today = timezone.localdate()
    horizon = today + timedelta(days=CHEQUE_WARNING_DAYS)

    # Debtors carry a negative balance, so the sum is negative; flip it so the
    # card reads as a positive amount owed to the company.
    owed = Customer.objects.filter(balance__lt=0).aggregate(
        total=Coalesce(Sum("balance"), ZERO, output_field=MONEY)
    )["total"]
    total_outstanding = -owed

    todays_sales = (
        Bill.objects.filter(bill_date=today)
        .exclude(status=Bill.Status.CANCELLED)
        .aggregate(total=Coalesce(Sum("total_amount"), ZERO, output_field=MONEY))["total"]
    )

    # Pending only: a held cheque is one we have chosen not to bank yet, and a
    # deposited one is finished with. No lower bound either — a cheque that
    # matured last week and still has not been banked is the most urgent row
    # on the page, not one to hide.
    maturing_cheques = (
        Cheque.objects.filter(maturity_date__lte=horizon, status=Cheque.Status.PENDING)
        .select_related("customer")
        .order_by("maturity_date")
    )

    recent_bills = Bill.objects.select_related("customer")[:5]

    # `owed` is the positive amount the customer owes, for display.
    top_customers = (
        Customer.objects.filter(balance__lt=0)
        .annotate(
            owed=ExpressionWrapper(Value(0) - F("balance"), output_field=MONEY)
        )
        .order_by("balance")[:5]
    )

    # Chart Data: Sales Trend (last 7 days)
    sales_labels = []
    sales_values = []
    bill_counts = []
    for i in range(6, -1, -1):
        d = today - timedelta(days=i)
        sales_labels.append(d.strftime("%a"))
        
        day_bills = Bill.objects.filter(bill_date=d).exclude(status=Bill.Status.CANCELLED)
        day_sales = day_bills.aggregate(total=Coalesce(Sum("total_amount"), ZERO, output_field=MONEY))["total"]
        
        sales_values.append(float(day_sales))
        bill_counts.append(day_bills.count())

    # Chart Data: Payment Mix (last 30 days)
    thirty_days_ago = today - timedelta(days=30)
    payment_mix = (
        Bill.objects.filter(bill_date__gte=thirty_days_ago)
        .exclude(status=Bill.Status.CANCELLED)
        .values("payment_type")
        .annotate(total=Coalesce(Sum("total_amount"), ZERO, output_field=MONEY))
        .order_by("-total")
    )
    payment_labels = []
    payment_values = []
    payment_type_map = dict(Bill.PaymentType.choices)

    for pm in payment_mix:
        if pm["total"] > 0:
            payment_labels.append(payment_type_map.get(pm["payment_type"], pm["payment_type"]))
            payment_values.append(float(pm["total"]))

    # ── Stock health ──────────────────────────────────────────────────────
    # Split rather than a single "needs attention" count so the operator can
    # see at a glance whether the shelf is genuinely bare or just running low.
    low_stock_qs = Product.objects.filter(
        is_active=True, qty__gt=0, qty__lte=settings.LOW_STOCK_THRESHOLD
    ).order_by("qty", "name")
    out_of_stock_qs = Product.objects.filter(is_active=True, qty__lte=0).order_by(
        "qty", "name"
    )
    low_stock_items = list(low_stock_qs[:6])
    low_stock_count = low_stock_qs.count()
    out_of_stock_count = out_of_stock_qs.count()

    # ── This month vs previous month ──────────────────────────────────────
    # Two figures the operator asks for weekly. Percent lives in the template
    # to keep the view thin, but the delta stays here because "did we grow"
    # is a business question not a display one.
    month_start = today.replace(day=1)
    if month_start.month == 1:
        prev_month_start = month_start.replace(year=month_start.year - 1, month=12)
    else:
        prev_month_start = month_start.replace(month=month_start.month - 1)

    def _sum_bills(start, end_exclusive):
        return (
            Bill.objects.filter(bill_date__gte=start, bill_date__lt=end_exclusive)
            .exclude(status=Bill.Status.CANCELLED)
            .aggregate(
                total=Coalesce(Sum("total_amount"), ZERO, output_field=MONEY),
                count=Count("id"),
            )
        )

    this_month = _sum_bills(month_start, month_start + timedelta(days=32))
    prev_month = _sum_bills(prev_month_start, month_start)
    this_month_total = this_month["total"]
    prev_month_total = prev_month["total"]

    if prev_month_total > 0:
        month_delta_pct = float(
            (this_month_total - prev_month_total) / prev_month_total * 100
        )
    elif this_month_total > 0:
        month_delta_pct = 100.0
    else:
        month_delta_pct = 0.0

    # ── Top-selling products this month ──────────────────────────────────
    top_products = (
        BillItem.objects.filter(
            bill__bill_date__gte=month_start,
        )
        .exclude(bill__status=Bill.Status.CANCELLED)
        .values("product__name", "product__size")
        .annotate(
            qty_sold=Coalesce(Sum("qty"), ZERO, output_field=MONEY),
            revenue=Coalesce(Sum("line_total"), ZERO, output_field=MONEY),
        )
        .order_by("-revenue")[:5]
    )

    return render(
        request,
        "core/dashboard.html",
        {
            "total_outstanding": total_outstanding,
            "todays_sales": todays_sales,
            "cash_balance": _cash_drawer_balance(),
            "maturing_cheques": maturing_cheques,
            "maturing_count": maturing_cheques.count(),
            # Identifies exactly this set of warnings on this day, so a dismissal
            # lasts until the warnings actually change rather than for ever.
            "cheque_signature": _warning_signature(today, maturing_cheques),
            "recent_bills": recent_bills,
            "top_customers": top_customers,
            "warning_days": CHEQUE_WARNING_DAYS,
            "today": today,
            "sales_labels_json": json.dumps(sales_labels),
            "sales_values_json": json.dumps(sales_values),
            "bill_counts_json": json.dumps(bill_counts),
            "payment_labels_json": json.dumps(payment_labels),
            "payment_values_json": json.dumps(payment_values),
            "low_stock_items": low_stock_items,
            "low_stock_count": low_stock_count,
            "out_of_stock_count": out_of_stock_count,
            "low_stock_threshold": settings.LOW_STOCK_THRESHOLD,
            "this_month_total": this_month_total,
            "this_month_count": this_month["count"],
            "prev_month_total": prev_month_total,
            "month_delta_pct": month_delta_pct,
            "month_label": month_start.strftime("%B %Y"),
            "top_products": top_products,
        },
    )


# --------------------------------------------------------------------- users
# Super-admin only. There is no self-registration: every account is created
# here, and its first password is generated rather than chosen.

#: Where a generated password waits between the POST that made it and the page
#: that shows it. The session, not a message, because it is popped exactly once
#: and must not survive into a second render.
CREDENTIALS_KEY = "new_credentials"


def _stash_credentials(request, username, password):
    request.session[CREDENTIALS_KEY] = {"username": username, "password": password}


def _pop_credentials(request):
    """The generated password, once.

    pop rather than read: refreshing the list, or coming back to it later, must
    not put the password back on screen. Once it has been rendered it is gone —
    the stored hash is all that is left, and a lost password is reset, not
    recovered.
    """
    return request.session.pop(CREDENTIALS_KEY, None)


@super_admin_required
def user_list(request):
    users = User.objects.order_by("username")
    page_obj = _paginate(request, users)
    return render(
        request,
        "core/user_list.html",
        {
            "page_obj": page_obj,
            "users": page_obj.object_list,
            "credentials": _pop_credentials(request),
        },
    )


@super_admin_required
def user_create(request):
    form = UserCreateForm(request.POST or None)
    if request.method == "POST" and form.is_valid():
        password = generate_password()
        user = form.save(commit=False)
        # set_password hashes it. The plain text exists only in this request,
        # long enough to be shown once.
        user.set_password(password)
        user.save()

        _stash_credentials(request, user.username, password)
        messages.success(request, f"{user.username} was created.")
        return redirect("core:user_list")

    return render(request, "core/user_form.html", {"form": form, "target": None})


@super_admin_required
def user_edit(request, pk):
    target = get_object_or_404(User, pk=pk)
    # A super admin editing themselves gets a reduced form: no role, no active
    # switch. See UserEditForm — the fields are dropped, not just hidden.
    is_self = target.pk == request.user.pk

    form = UserEditForm(request.POST or None, instance=target, is_self=is_self)
    if request.method == "POST" and form.is_valid():
        form.save()
        messages.success(request, f"{target.username} was updated.")
        return redirect("core:user_list")

    return render(
        request,
        "core/user_form.html",
        {"form": form, "target": target, "is_self": is_self},
    )


@super_admin_required
def user_set_password(request, pk):
    """A super admin types a new password for any user.

    Replaces the older reset-to-random flow: the operator picks the
    password themselves and hands it over directly, no on-screen dance
    with a generated string to write down. Own account still allowed —
    the session auth hash is re-stamped so the admin isn't kicked out
    of their own session mid-change.
    """
    target = get_object_or_404(User, pk=pk)
    form = SetUserPasswordForm(request.POST or None)

    if request.method == "POST" and form.is_valid():
        password = form.cleaned_data["new_password1"]
        target.set_password(password)
        target.save(update_fields=["password"])

        if target.pk == request.user.pk:
            update_session_auth_hash(request, target)

        messages.success(
            request,
            f"{target.username}'s password was updated.",
        )
        return redirect("core:user_list")

    return render(
        request,
        "core/user_set_password.html",
        {"form": form, "target": target},
    )


@login_required
def profile(request):
    """A user's own profile page — view details and change their password.

    Reachable by any signed-in user from the sidebar footer chip. Two forms
    on one page: basic details (name + email, username stays put) and
    password (current + new + confirm). Each has its own submit so the
    operator can change one without touching the other.
    """
    user = request.user
    details_form = ProfileDetailsForm(instance=user)
    password_form = ProfilePasswordForm(user=user)

    if request.method == "POST":
        action = request.POST.get("action") or ""
        if action == "details":
            details_form = ProfileDetailsForm(request.POST, instance=user)
            if details_form.is_valid():
                details_form.save()
                messages.success(request, "Profile updated.")
                return redirect("core:profile")
            messages.error(request, f"Profile not saved: {details_form.first_error()}")
        elif action == "password":
            password_form = ProfilePasswordForm(request.POST, user=user)
            if password_form.is_valid():
                user.set_password(password_form.cleaned_data["new_password1"])
                user.save(update_fields=["password"])
                # Re-stamp the session auth hash so this request doesn't
                # sign the user straight out.
                update_session_auth_hash(request, user)
                messages.success(request, "Password changed.")
                return redirect("core:profile")
            messages.error(request, f"Password not saved: {password_form.first_error()}")

    return render(
        request,
        "core/profile.html",
        {
            "details_form": details_form,
            "password_form": password_form,
        },
    )


@require_POST
@super_admin_required
def user_deactivate(request, pk):
    target = get_object_or_404(User, pk=pk)

    # Also the reason no "last super admin" check is needed: the only person
    # who can deactivate accounts is a super admin, and they cannot be the one
    # going inactive — so an active super admin always remains.
    if target.pk == request.user.pk:
        messages.error(request, "You can't deactivate your own account.")
        return redirect("core:user_list")

    target.is_active = False
    target.save(update_fields=["is_active"])
    messages.success(
        request,
        f"{target.username} was deactivated and can no longer sign in. "
        f"The records they created are unchanged.",
    )
    return redirect("core:user_list")


@require_POST
@super_admin_required
def user_activate(request, pk):
    target = get_object_or_404(User, pk=pk)
    target.is_active = True
    target.save(update_fields=["is_active"])
    messages.success(request, f"{target.username} can sign in again.")
    return redirect("core:user_list")


@require_POST
@super_admin_required
def user_delete(request, pk):
    """Hard-delete a user account.

    Refuses self-delete for the same reason `user_deactivate` does — the only
    people who can run this are super admins, and losing the last one would
    lock the app.

    User is `on_delete=PROTECT`-referenced from most audit trails (bill edits,
    settlements, adjustments, material purchases, weigh entries, etc.), so any
    account that has ever done anything cannot be hard-deleted. That's the
    point: the audit trail must stay whole. Catch the ProtectedError and steer
    the operator to Deactivate instead, matching the delete pattern used for
    Product/Category/Customer.
    """
    target = get_object_or_404(User, pk=pk)

    if target.pk == request.user.pk:
        messages.error(request, "You can't delete your own account.")
        return redirect("core:user_list")

    username = target.username
    try:
        target.delete()
    except ProtectedError:
        messages.error(
            request,
            f"{username} has records attached (bills, payments, audits) and "
            f"can't be deleted. Deactivate the account instead — it stops "
            f"them signing in without losing the history.",
        )
        return redirect("core:user_list")

    messages.success(request, f"{username} was deleted.")
    return redirect("core:user_list")


# --------------------------------------------------------------- audit log
# Super-admin only. Reads every business write across the system, filterable
# by month/user/action/target-type, with a "delete this whole month" button.


#: Rows per page on the log — the log is scanned in bulk (unlike the ledger),
#: so a shorter page pays off in scroll speed.
AUDIT_PAGE_SIZE = 40


def _audit_month_choices():
    """Distinct months present in the log, newest first.

    Powers both the month dropdown and the "delete this month" button list.
    Read as a list here rather than a queryset because the template iterates
    twice — once for the filter, once for the delete panel — and repeating
    the query would waste a round-trip for no gain.
    """
    months = (
        AuditLog.objects.values_list("month", flat=True)
        .distinct()
        .order_by("-month")
    )
    return list(months)


@super_admin_required
def audit_log_list(request):
    """The activity feed for the whole system.

    Filters read off the query string so a bookmark round-trips:
      ?month=YYYY-MM  (or ?month=all)
      ?user=<id>
      ?action=create|update|delete
      ?target=<ModelName>
      ?q=<free text against target_label/summary>
    """
    logs = AuditLog.objects.select_related("user").all()

    month_filter = get_month_filter(request)
    logs = month_filter.apply(logs, field="month")

    user_id = (request.GET.get("user") or "").strip()
    if user_id.isdigit():
        logs = logs.filter(user_id=int(user_id))

    action = (request.GET.get("action") or "").strip()
    if action in {choice for choice, _ in AuditLog.Action.choices}:
        logs = logs.filter(action=action)

    target = (request.GET.get("target") or "").strip()
    if target:
        logs = logs.filter(target_type=target)

    query = (request.GET.get("q") or "").strip()
    if query:
        logs = logs.filter(
            Q(target_label__icontains=query)
            | Q(summary__icontains=query)
            | Q(username_snapshot__icontains=query)
        )

    page_obj = _paginate(request, logs, per_page=AUDIT_PAGE_SIZE)

    # Cardinality summaries. Cheap enough — a month's log is tens of thousands
    # of rows at most in this project's scale.
    scoped = month_filter.apply(AuditLog.objects.all(), field="month")
    action_counts = (
        scoped.values("action")
        .annotate(n=Count("id"))
        .order_by("-n")
    )
    action_totals = {row["action"]: row["n"] for row in action_counts}

    return render(
        request,
        "core/audit_log.html",
        {
            "page_obj": page_obj,
            "logs": page_obj.object_list,
            "month_filter": month_filter,
            "month_choices": _audit_month_choices(),
            "user_choices": User.objects.order_by("username"),
            "target_choices": list(
                AuditLog.objects.values_list("target_type", flat=True)
                .distinct()
                .order_by("target_type")
            ),
            "action_choices": AuditLog.Action.choices,
            "action_totals": action_totals,
            "total_all_time": AuditLog.objects.count(),
            "total_in_scope": scoped.count(),
            "filters": {
                "user": user_id,
                "action": action,
                "target": target,
                "q": query,
            },
        },
    )


@require_POST
@super_admin_required
def audit_log_delete_month(request):
    """Delete every log row for the given month.

    Confirmed via a modal in the template. Reads `month=YYYY-MM` from POST;
    anything else — including the "all months" sentinel from the filter — is
    refused. Deleting the whole log unconditionally is a separate button
    below, deliberately kept apart from the routine month-purge to force a
    second confirmation.
    """
    from datetime import date as _date
    raw = (request.POST.get("month") or "").strip()
    try:
        year, month = raw.split("-")
        month_start = _date(int(year), int(month), 1)
    except (ValueError, TypeError):
        messages.error(request, "Pick a specific month to delete.")
        return redirect("core:audit_log_list")

    deleted, _ = AuditLog.objects.filter(month=month_start).delete()
    messages.success(
        request,
        f"Deleted {deleted} log entr{'y' if deleted == 1 else 'ies'} for "
        f"{month_start.strftime('%B %Y')}.",
    )
    return redirect("core:audit_log_list")


# ----------------------------------------------------------- billing settings
# Super-admin only. Singleton row edited in place — no create/list/delete.


@super_admin_required
def billing_settings(request):
    """Edit the company header/footer printed on every bill.

    Singleton — no pk in the URL. The model's `load()` returns the same row
    every time, so a fresh POST just re-saves it; nothing here has to worry
    about "which record" the operator is editing.
    """
    instance = BillingSettings.load()

    if request.method == "POST":
        form = BillingSettingsForm(request.POST, instance=instance)
        if form.is_valid():
            form.save()
            messages.success(request, "Billing settings updated.")
            return redirect("core:billing_settings")
        messages.error(request, "Please fix the highlighted fields.")
    else:
        form = BillingSettingsForm(instance=instance)

    # A sample bill id lets the "Preview" button open a real bill styled with
    # the current settings — no seed data, no pk in the URL is fine but the
    # template needs one.
    sample_bill_id = Bill.objects.order_by("-id").values_list("id", flat=True).first()

    return render(
        request,
        "core/billing_settings.html",
        {"form": form, "settings": instance, "sample_bill_id": sample_bill_id},
    )


# ---------------------------------------------------------------- categories
# Super-admin only. Managers are redirected to the dashboard with an error.


@super_admin_required
def category_list(request):
    query = request.GET.get("q", "").strip()

    # order_by repeats Category.Meta.ordering, which the annotate() below would
    # otherwise drop — see _bills_with_counts.
    categories = Category.objects.annotate(product_count=Count("products")).order_by(
        "name"
    )
    if query:
        categories = categories.filter(
            Q(name__icontains=query) | Q(description__icontains=query)
        )

    page_obj = _paginate(request, categories)

    return render(
        request,
        "core/category_list.html",
        {
            "page_obj": page_obj,
            "categories": page_obj.object_list,
            "query": query,
        },
    )


@super_admin_required
def category_create(request):
    form = CategoryForm(request.POST or None)

    if request.method == "POST" and form.is_valid():
        category = form.save()
        messages.success(request, f"Category '{category.name}' was created.")
        return redirect("core:category_list")

    return render(
        request,
        "core/category_form.html",
        {"form": form, "is_edit": False},
    )


@super_admin_required
def category_update(request, pk):
    category = get_object_or_404(Category, pk=pk)
    form = CategoryForm(request.POST or None, instance=category)

    if request.method == "POST" and form.is_valid():
        category = form.save()
        messages.success(request, f"Category '{category.name}' was updated.")
        return redirect("core:category_list")

    return render(
        request,
        "core/category_form.html",
        {"form": form, "category": category, "is_edit": True},
    )


@require_POST
@super_admin_required
def category_delete(request, pk):
    category = get_object_or_404(Category, pk=pk)
    name = category.name

    try:
        category.delete()
    except ProtectedError:
        # Product.category is on_delete=PROTECT, so a category still holding
        # products cannot be removed. Explain rather than 500.
        count = category.products.count()
        messages.error(
            request,
            f"Cannot delete '{name}' — {count} product{'' if count == 1 else 's'} "
            f"still belong to it. Reassign or remove them first.",
        )
    else:
        messages.success(request, f"Category '{name}' was deleted.")

    return redirect("core:category_list")


# ------------------------------------------------------------------ products
# Managers may list, create and edit. Deleting is super-admin only.


@login_required
def product_list(request):
    query = request.GET.get("q", "").strip()
    category_id = request.GET.get("category", "").strip()
    # ?view=stock — the sidebar's Stock Ledgers link — narrows the list to
    # products the operator would want to open a ledger for: negative
    # (oversold), zero, or low.
    stock_view = request.GET.get("view", "").strip() == "stock"

    # order_by repeats Product.Meta.ordering, which the annotate() below would
    # otherwise drop — see _bills_with_counts.
    products = (
        Product.objects.select_related("category")
        .annotate(custom_price_count=Count("customer_prices"))
        .order_by("name", "size")
    )
    if query:
        products = products.filter(name__icontains=query)

    # An unparsable ?category= is ignored rather than 500ing on a bad filter.
    selected_category = None
    if category_id.isdigit():
        selected_category = int(category_id)
        products = products.filter(category_id=selected_category)

    if stock_view:
        products = products.filter(qty__lte=settings.LOW_STOCK_THRESHOLD)

    page_obj = _paginate(request, products)

    # For the native <datalist> autocomplete on the search box. Flat list
    # of names; the browser filters as the operator types and Enter submits
    # the form as normal.
    suggest_names = list(
        Product.objects.order_by("name")
        .values_list("name", flat=True)
        .distinct()
    )

    return render(
        request,
        "core/product_list.html",
        {
            "page_obj": page_obj,
            "products": page_obj.object_list,
            "categories": Category.objects.all(),
            "query": query,
            "selected_category": selected_category,
            "stock_view": stock_view,
            "is_filtered": bool(query or selected_category or stock_view),
            "low_stock_threshold": settings.LOW_STOCK_THRESHOLD,
            "suggest_names": suggest_names,
        },
    )


@login_required
def product_create(request):
    form = ProductForm(request.POST or None)

    if request.method == "POST" and form.is_valid():
        product = form.save()
        messages.success(request, f"Product '{product}' was created.")
        return redirect("core:product_list")

    return render(
        request,
        "core/product_form.html",
        {"form": form, "is_edit": False},
    )


@login_required
def product_update(request, pk):
    product = get_object_or_404(Product, pk=pk)
    form = ProductForm(request.POST or None, instance=product)

    if request.method == "POST" and form.is_valid():
        product = form.save()
        messages.success(request, f"Product '{product}' was updated.")
        return redirect("core:product_list")

    return render(
        request,
        "core/product_form.html",
        {"form": form, "product": product, "is_edit": True},
    )


@require_POST
@super_admin_required
def product_delete(request, pk):
    product = get_object_or_404(Product, pk=pk)
    label = str(product)

    try:
        product.delete()
    except ProtectedError:
        # Bill items, supplier bill items and production entries all PROTECT
        # their product, so anything with history stays put.
        messages.error(
            request,
            f"Cannot delete '{label}' — it is used by bills, supplier bills or "
            f"production entries. Deactivate it instead to hide it from new bills.",
        )
    else:
        messages.success(request, f"Product '{label}' was deleted.")

    return redirect("core:product_list")


@require_POST
@login_required
def product_toggle_active(request, pk):
    """Flip is_active from the list page. Answers JSON for the inline toggle."""
    product = get_object_or_404(Product, pk=pk)
    product.is_active = not product.is_active
    product.save(update_fields=["is_active"])

    return JsonResponse(
        {
            "ok": True,
            "is_active": product.is_active,
            "label": "Active" if product.is_active else "Inactive",
        }
    )


@login_required
def product_export_excel(request):
    """Download every product with its current stock as an .xlsx.

    Honours the same filters as product_list — ?q=, ?category=, ?view=stock
    and ?status=active|inactive — so the user can narrow the sheet first and
    then export exactly what they see.
    """
    from io import BytesIO

    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    query = request.GET.get("q", "").strip()
    category_id = request.GET.get("category", "").strip()
    stock_view = request.GET.get("view", "").strip() == "stock"
    status = request.GET.get("status", "").strip().lower()

    products = (
        Product.objects.select_related("category")
        .annotate(custom_price_count=Count("customer_prices"))
        .order_by("name", "size")
    )
    if query:
        products = products.filter(name__icontains=query)
    if category_id.isdigit():
        products = products.filter(category_id=int(category_id))
    if stock_view:
        products = products.filter(qty__lte=settings.LOW_STOCK_THRESHOLD)
    if status == "active":
        products = products.filter(is_active=True)
    elif status == "inactive":
        products = products.filter(is_active=False)

    today = timezone.localdate()
    threshold = settings.LOW_STOCK_THRESHOLD

    wb = Workbook()
    ws = wb.active
    ws.title = "Stock"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1F2937")
    right = Alignment(horizontal="right")
    center = Alignment(horizontal="center")

    # Masthead
    ws["A1"] = "Senovka Plastics — Product Stock Report"
    ws["A1"].font = Font(bold=True, size=14)
    ws.merge_cells("A1:H1")

    ws["A2"] = "As of"
    ws["A2"].font = Font(bold=True)
    ws["B2"] = today.strftime("%d %b %Y")

    # A short "filters used" line, so the exported file is self-documenting.
    filter_bits = []
    if query:
        filter_bits.append(f'Search: "{query}"')
    if category_id.isdigit():
        cat = Category.objects.filter(pk=int(category_id)).first()
        if cat:
            filter_bits.append(f"Category: {cat.name}")
    if stock_view:
        filter_bits.append(f"Low stock only (≤ {threshold})")
    if status in ("active", "inactive"):
        filter_bits.append(f"Status: {status.title()}")
    if filter_bits:
        ws["C2"] = "Filters"
        ws["C2"].font = Font(bold=True)
        ws["D2"] = " · ".join(filter_bits)
        ws.merge_cells("D2:H2")

    # Items header
    HEADERS = [
        "No", "Name", "Size", "Category",
        "Qty", "Stock Status", "Default Price", "Active",
    ]
    header_row = 4
    for idx, name in enumerate(HEADERS, start=1):
        cell = ws.cell(row=header_row, column=idx, value=name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center

    def stock_status(qty):
        if qty < 0:
            return "Oversold"
        if qty == 0:
            return "Out of Stock"
        if qty <= threshold:
            return "Low Stock"
        return "In Stock"

    row_num = header_row + 1
    for i, product in enumerate(products, start=1):
        ws.cell(row=row_num, column=1, value=i).alignment = center
        ws.cell(row=row_num, column=2, value=product.name)
        ws.cell(row=row_num, column=3, value=product.size or "")
        ws.cell(row=row_num, column=4, value=product.category.name)

        c_qty = ws.cell(row=row_num, column=5, value=float(product.qty))
        c_qty.alignment = right
        c_qty.number_format = "#,##0.000"

        ws.cell(row=row_num, column=6, value=stock_status(product.qty)).alignment = center

        c_price = ws.cell(row=row_num, column=7, value=float(product.default_price))
        c_price.alignment = right
        c_price.number_format = "#,##0.00"

        ws.cell(
            row=row_num, column=8,
            value="Yes" if product.is_active else "No",
        ).alignment = center
        row_num += 1

    # Freeze the masthead + header row so the header stays visible on scroll.
    ws.freeze_panes = f"A{header_row + 1}"

    # Column widths — a rough auto-size based on typical content width.
    widths = {"A": 5, "B": 32, "C": 12, "D": 18, "E": 12, "F": 14, "G": 14, "H": 9}
    for letter, width in widths.items():
        ws.column_dimensions[letter].width = width

    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    response = HttpResponse(
        stream.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = (
        f'attachment; filename="products_stock_{today.strftime("%Y-%m-%d")}.xlsx"'
    )
    return response


def _stock_ledger_rows(product):
    """Every stock movement on `product`, oldest first, with a running balance
    and running production total.

    Three sources land in one column layout:
      Production       — a ProductionEntry row (own manufacture, corrections,
                         or an auto Oversale row that covered a shortfall)
      Supplier receipt — a SupplierBillItem (goods arriving from a supplier)
      Sale             — a BillItem (goods leaving on a customer bill)

    An opening "as stock" row is always first, holding the balance the ledger
    has to start from so it ends at Product.qty. Computed rather than stored:
    opening = current stock − sum(inputs) + sum(outputs). If everything has
    been recorded through the app since day one, opening comes out as 0.

    All work happens in Python because the balance and total both depend on
    every earlier row — SQL cannot express that a page at a time.
    """
    events = []

    for entry in ProductionEntry.objects.filter(product=product):
        is_oversale = entry.reason.startswith(OVERSALE_REASON_PREFIX)
        events.append(
            {
                "date": entry.production_date,
                # (date, kind, tiebreaker) — production comes before same-day
                # sales so the sale reads as drawing on that morning's batch,
                # not on stock that arrived later in the day.
                "_sort": (entry.production_date, 0, entry.created_at, entry.pk),
                "kind": "oversale" if is_oversale else "production",
                "production": entry.qty_produced,
                "sales": None,
                "customer": entry.reason or "",
                "bill_number": "",
            }
        )

    supplier_items = (
        SupplierBillItem.objects.filter(product=product)
        .exclude(supplier_bill__status=SupplierBill.Status.CANCELLED)
        .select_related("supplier_bill__supplier")
    )
    for item in supplier_items:
        sb = item.supplier_bill
        events.append(
            {
                "date": sb.bill_date,
                # No created_at on SupplierBillItem — order by pk within the
                # day, which is monotonic and stable across page loads.
                "_sort": (sb.bill_date, 0, sb.bill_date, item.pk),
                "kind": "supplier",
                "production": item.qty,
                "sales": None,
                "customer": f"Supplier: {sb.supplier.name}",
                "bill_number": f"SUP-{sb.pk}",
            }
        )

    bill_items = (
        BillItem.objects.filter(product=product)
        .exclude(bill__status=Bill.Status.CANCELLED)
        .select_related("bill__customer")
    )
    for item in bill_items:
        bill = item.bill
        if bill.is_walk_in:
            who = bill.walk_in_name or "Walk-in"
        elif bill.customer_id:
            who = bill.customer.name
        else:
            who = "—"
        events.append(
            {
                "date": bill.bill_date,
                # Sales sort *after* productions on the same day (kind=1 vs 0).
                "_sort": (bill.bill_date, 1, bill.bill_date, item.pk),
                "kind": "sale",
                "production": None,
                "sales": item.qty,
                "customer": who,
                "bill_number": f"#{bill.pk:04d}",
                # Explicit pk so the template can link straight to the bill
                # without parsing the formatted "#0001" back apart.
                "bill_pk": bill.pk,
            }
        )

    # Manual stock corrections. Signed qty: positive rides in the PRODUCTION
    # column (also bumps the running total, since the shelf gained it),
    # negative rides in the SALES column (the shelf lost it, though not to
    # a sale). The kind='adjust_up' / 'adjust_down' tint keeps them
    # visually distinct from real production or a customer sale.
    for adj in StockAdjustment.objects.filter(product=product).select_related("adjusted_by"):
        is_up = adj.qty >= 0
        events.append(
            {
                "date": adj.adjustment_date,
                # Same day, after production and supplier receipts (kind=2)
                # but before same-day sales (they'd read as adjusting to a
                # position before the day's selling started).
                "_sort": (adj.adjustment_date, 0, adj.created_at, adj.pk),
                "kind": "adjust_up" if is_up else "adjust_down",
                "production": adj.qty if is_up else None,
                "sales": (-adj.qty) if not is_up else None,
                "customer": f"Adjustment — {adj.reason} · by {adj.adjusted_by.username}",
                "bill_number": f"ADJ-{adj.pk}",
                # For the delete button in the template.
                "adjustment_pk": adj.pk,
            }
        )

    events.sort(key=lambda e: e["_sort"])

    total_in = sum(
        (e["production"] or Decimal("0.000") for e in events), Decimal("0.000")
    )
    total_out = sum(
        (e["sales"] or Decimal("0.000") for e in events), Decimal("0.000")
    )
    # Opening = current shelf, minus everything the ledger says came in, plus
    # everything it says went out. If every movement has been recorded, this
    # comes to zero (or whatever the shelf held before the app was in use).
    opening = product.qty - total_in + total_out

    rows = []
    opening_date = events[0]["date"] if events else timezone.localdate()
    rows.append(
        {
            "date": opening_date,
            "production": None,
            "total": None,
            "sales": None,
            "balance": opening,
            "customer": "as stock",
            "bill_number": "",
            "kind": "opening",
        }
    )

    balance = opening
    running_total = Decimal("0.000")
    for e in events:
        if e["production"] is not None:
            balance += e["production"]
            running_total += e["production"]
            rows.append(
                {
                    "date": e["date"],
                    "production": e["production"],
                    "total": running_total,
                    "sales": None,
                    "balance": balance,
                    "customer": e["customer"],
                    "bill_number": e["bill_number"],
                    "kind": e["kind"],
                    "bill_pk": e.get("bill_pk"),
                    "adjustment_pk": e.get("adjustment_pk"),
                }
            )
        else:
            balance -= e["sales"]
            rows.append(
                {
                    "date": e["date"],
                    "production": None,
                    # TOTAL only advances on production/supplier rows — it is
                    # a cumulative production counter, not a running balance.
                    "total": None,
                    "sales": e["sales"],
                    "balance": balance,
                    "customer": e["customer"],
                    "bill_number": e["bill_number"],
                    "kind": e["kind"],
                    "bill_pk": e.get("bill_pk"),
                    "adjustment_pk": e.get("adjustment_pk"),
                }
            )

    return {
        "rows": rows,
        "opening": opening,
        "total_produced": total_in,
        "total_sold": total_out,
        "closing_balance": balance,
    }


@require_POST
@login_required
def stock_adjust_create(request, pk):
    """Manually set a product's shelf to an exact quantity.

    The form's `qty` field holds the *target* stock value the operator
    wants on the shelf.  The actual delta applied to Product.qty is
    (target − current), computed here after a fresh DB read so that a
    concurrent sale or production is priced in before the delta lands.
    Wrapping the shelf update and audit row in one transaction means a
    mid-flight failure cannot leave Product.qty half-moved.
    """
    product = get_object_or_404(Product, pk=pk)
    form = StockAdjustmentForm(request.POST)
    if not form.is_valid():
        messages.error(request, f"Adjustment not saved: {form.first_error()}")
        return redirect("core:stock_ledger", pk=product.pk)

    with transaction.atomic():
        # Fresh read so a concurrent sale/production is priced in.
        current = Product.objects.values_list("qty", flat=True).get(pk=product.pk)
        target = form.cleaned_data["qty"]  # desired final stock value
        delta = target - current           # signed change to apply
        entry = form.save(commit=False)
        entry.product = product
        entry.adjusted_by = request.user
        entry.stock_before = current
        entry.stock_after = target
        entry.qty = delta                  # store delta so ledger & delete work
        entry.save()
        Product.objects.filter(pk=product.pk).update(qty=target)

    messages.success(
        request,
        f"Stock set to {target} on {product.name}. "
        f"(Changed by {'+' if delta >= 0 else ''}{delta}.)",
    )
    return redirect("core:stock_ledger", pk=product.pk)


@require_POST
@super_admin_required
def stock_adjust_delete(request, pk):
    """Reverse an adjustment: undo its stock movement and drop the row.

    Super-admin only — an adjustment is what accountability rests on for
    counted-shelf and scrap corrections, so managers can create them but
    only a super admin can rewind one.
    """
    entry = get_object_or_404(
        StockAdjustment.objects.select_related("product"), pk=pk
    )
    product = entry.product
    with transaction.atomic():
        Product.objects.filter(pk=product.pk).update(qty=F("qty") - entry.qty)
        entry.delete()

    messages.success(request, f"Adjustment on {product.name} reversed.")
    return redirect("core:stock_ledger", pk=product.pk)


@login_required
def stock_ledger(request, pk):
    """One product's full movement history with a running balance.

    The month filter narrows the *display* — the running balance is still
    computed over the whole ledger, so the first row shown carries the
    balance as it stood at the end of the prior month, not zero.
    """
    product = get_object_or_404(Product.objects.select_related("category"), pk=pk)

    ledger = _stock_ledger_rows(product)
    rows = ledger["rows"]

    month_filter = get_month_filter(request)
    if not month_filter.is_all_time:
        rows = [
            r for r in rows
            if month_filter.start <= r["date"] <= month_filter.end
        ]

    page_obj = _paginate(request, rows, settings.PAGINATE_BY_REPORTS)

    # If the ledger's closing figure disagrees with the shelf, say so — a
    # mismatch means a stock move happened outside the app.
    ledger_mismatch = ledger["closing_balance"] != product.qty

    return render(
        request,
        "core/stock_ledger.html",
        {
            "product": product,
            "page_obj": page_obj,
            "rows": page_obj.object_list,
            "month_filter": month_filter,
            "opening_balance": ledger["opening"],
            "total_produced": ledger["total_produced"],
            "total_sold": ledger["total_sold"],
            "closing_balance": ledger["closing_balance"],
            "current_stock": product.qty,
            "ledger_mismatch": ledger_mismatch,
            "low_stock_threshold": settings.LOW_STOCK_THRESHOLD,
        },
    )


def _write_stock_ledger_sheet(ws, product, month_filter):
    """Write one product's whole ledger onto worksheet `ws`.

    Extracted so the single-product export and the multi-product bulk export
    can share every column width, colour and border rule — the two would drift
    the moment either one grew a column, and a stock ledger that prints
    differently in the bulk file from the direct download is exactly the sort
    of thing that quietly loses a reader's trust.
    """
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    ledger = _stock_ledger_rows(product)
    rows = ledger["rows"]

    if not month_filter.is_all_time:
        rows = [
            r for r in rows
            if month_filter.start <= r["date"] <= month_filter.end
        ]

    thin = Side(style="thin")
    border_all = Border(top=thin, bottom=thin, left=thin, right=thin)
    bold = Font(bold=True)
    bold_large = Font(bold=True, size=14)
    center = Alignment(horizontal="center", vertical="center")
    right = Alignment(horizontal="right", vertical="center")
    left = Alignment(horizontal="left", vertical="center")

    header_fill = PatternFill("solid", fgColor="1F2937")
    header_font = Font(bold=True, color="FFFFFF")

    fill_opening = PatternFill("solid", fgColor="F3F4F6")
    fill_production = PatternFill("solid", fgColor="ECFDF5")
    fill_supplier = PatternFill("solid", fgColor="F0F9FF")
    fill_sale = PatternFill("solid", fgColor="FFF1F2")
    fill_adjust = PatternFill("solid", fgColor="FEF3C7")

    ws.column_dimensions["A"].width = 15
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 16
    ws.column_dimensions["F"].width = 18
    ws.column_dimensions["G"].width = 30
    ws.column_dimensions["H"].width = 15

    ws["A1"] = "Senovka Plastics — Stock Ledger"
    ws["A1"].font = bold_large
    ws.merge_cells("A1:H1")

    ws["A2"] = "Product:"; ws["B2"] = product.name; ws["B2"].font = bold
    ws["C2"] = "Size:"; ws["D2"] = product.size or "—"; ws["D2"].font = bold
    ws["E2"] = "Category:"; ws["F2"] = product.category.name; ws["F2"].font = bold
    ws["G2"] = "Period:"; ws["H2"] = month_filter.label; ws["H2"].font = bold

    for cell_addr in ["A2", "C2", "E2", "G2"]:
        ws[cell_addr].font = Font(bold=True)

    ws["A4"] = "Opening Balance"; ws["B4"] = float(ledger["opening"])
    ws["C4"] = "Total Produced"; ws["D4"] = float(ledger["total_produced"])
    ws["E4"] = "Total Sold"; ws["F4"] = float(ledger["total_sold"])
    ws["G4"] = "Closing Balance"; ws["H4"] = float(ledger["closing_balance"])

    ws["B4"].font = bold; ws["D4"].font = bold; ws["F4"].font = bold; ws["H4"].font = bold
    for cell_addr in ["A4", "C4", "E4", "G4"]:
        ws[cell_addr].font = Font(bold=True)

    headers = [
        "Date", "Type", "Production (+)", "Cumulative Prod.",
        "Sales (-)", "Running Balance", "Customer / Detail", "Ref / Bill No",
    ]
    header_row = 6
    for idx, name in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=idx, value=name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border_all

    row_num = header_row + 1
    for r in rows:
        kind = r["kind"]
        kind_label = kind.title().replace("_", " ")
        row_fill = None
        if kind == "opening":
            row_fill = fill_opening
        elif kind == "production" or kind == "oversale":
            row_fill = fill_production
        elif kind == "supplier":
            row_fill = fill_supplier
        elif kind == "sale":
            row_fill = fill_sale
        elif kind in ("adjust_up", "adjust_down"):
            row_fill = fill_adjust

        c_date = ws.cell(row=row_num, column=1, value=r["date"].strftime("%d %b %Y"))
        c_type = ws.cell(row=row_num, column=2, value=kind_label)
        c_prod = ws.cell(row=row_num, column=3, value=float(r["production"]) if r["production"] is not None else "")
        c_cum = ws.cell(row=row_num, column=4, value=float(r["total"]) if r["total"] is not None else "")
        c_sales = ws.cell(row=row_num, column=5, value=float(r["sales"]) if r["sales"] is not None else "")
        c_bal = ws.cell(row=row_num, column=6, value=float(r["balance"]))
        c_cust = ws.cell(row=row_num, column=7, value=r["customer"])
        c_ref = ws.cell(row=row_num, column=8, value=r["bill_number"])

        c_date.alignment = center
        c_type.alignment = center
        c_prod.alignment = right
        c_cum.alignment = right
        c_sales.alignment = right
        c_bal.alignment = right
        c_cust.alignment = left
        c_ref.alignment = center

        for cell in (c_prod, c_cum, c_sales, c_bal):
            cell.number_format = "#,##0.000"

        for c_idx in range(1, 9):
            cell = ws.cell(row=row_num, column=c_idx)
            cell.border = border_all
            if row_fill:
                cell.fill = row_fill
            if kind == "opening":
                cell.font = Font(italic=True)

        row_num += 1

    for c_idx in range(1, 9):
        ws.cell(row=row_num, column=c_idx).border = Border(top=thin)


def _sheet_title_for(base, used_titles):
    """A safe, unique worksheet title.

    Excel caps sheet names at 31 characters and rejects `[ ] : * ? / \\`. Two
    selected products with the same first 31 chars would also collide, so we
    disambiguate with a numeric suffix. The `used_titles` set is mutated in
    place — passing it back to the caller means the next sheet reuses the same
    memory of what's already claimed.
    """
    cleaned = "".join(c for c in base if c not in "[]:*?/\\")[:31].strip() or "Sheet"
    candidate = cleaned
    suffix = 2
    while candidate in used_titles:
        # Reserve room for the suffix so the trimmed version still fits.
        room = 31 - len(f" ({suffix})")
        candidate = f"{cleaned[:room]} ({suffix})"
        suffix += 1
    used_titles.add(candidate)
    return candidate


def _xlsx_response(wb, filename):
    """Serialise a workbook to a Content-Disposition HTTP response."""
    from io import BytesIO

    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    response = HttpResponse(
        stream.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


@login_required
def stock_ledger_excel(request, pk):
    """One product's full movement history as an Excel sheet.

    Respects the active month filter. Delegates the sheet body to
    `_write_stock_ledger_sheet` — the bulk export uses the same helper so the
    two exports never render differently.
    """
    from openpyxl import Workbook

    product = get_object_or_404(Product.objects.select_related("category"), pk=pk)
    month_filter = get_month_filter(request)

    wb = Workbook()
    ws = wb.active
    ws.title = _sheet_title_for("Stock Ledger", set())
    _write_stock_ledger_sheet(ws, product, month_filter)

    clean_name = product.name.replace(" ", "_").lower()
    filename_param = month_filter.param if not month_filter.is_all_time else "all"
    return _xlsx_response(
        wb, f"stock_ledger_{clean_name}_{filename_param}.xlsx"
    )


def _parse_id_list(request, param="ids"):
    """Product/customer id list from either GET (`?ids=1,2,3` or `?ids=1&ids=2`)
    or POST (`ids` as either a single comma-string or a repeated field).

    A single sanitiser both places lets the checkboxes in the list templates
    submit either way — a form POST for a CSRF-clean cross-user link, or a
    GET so a bookmarked selection round-trips.
    """
    raw = []
    if request.method == "POST":
        raw = request.POST.getlist(param) or [request.POST.get(param, "")]
    else:
        raw = request.GET.getlist(param) or [request.GET.get(param, "")]

    ids = []
    for chunk in raw:
        for part in (chunk or "").split(","):
            part = part.strip()
            if part.isdigit():
                ids.append(int(part))
    # Preserve first-seen order but drop duplicates — a workbook with two
    # identical sheets is worthless and the sheet-title logic would rename
    # the second one anyway.
    seen = set()
    unique = []
    for pid in ids:
        if pid not in seen:
            seen.add(pid)
            unique.append(pid)
    return unique


@login_required
def stock_ledger_bulk_excel(request):
    """Multiple products' stock ledgers in one workbook — one sheet each.

    Fired by the "Download stock ledger (Excel)" button on the product list
    once at least one row is ticked. Products are read in `Product.Meta.ordering`
    order so the workbook's tabs match the order the operator sees on screen.
    """
    from openpyxl import Workbook

    ids = _parse_id_list(request)
    if not ids:
        messages.error(request, "Pick at least one product first.")
        return redirect("core:product_list")

    products = list(
        Product.objects.filter(pk__in=ids).select_related("category")
    )
    if not products:
        messages.error(request, "None of the picked products exist.")
        return redirect("core:product_list")

    month_filter = get_month_filter(request)

    wb = Workbook()
    # Workbook() ships with a blank default sheet; remove it so the first
    # product owns its own tab rather than starting at the anonymous "Sheet".
    default_ws = wb.active
    wb.remove(default_ws)

    used_titles = set()
    for product in products:
        title_base = f"{product.name} {product.size}".strip() or f"Product {product.pk}"
        ws = wb.create_sheet(title=_sheet_title_for(title_base, used_titles))
        _write_stock_ledger_sheet(ws, product, month_filter)

    stamp = timezone.localdate().isoformat()
    filename_param = month_filter.param if not month_filter.is_all_time else "all"
    return _xlsx_response(
        wb, f"stock_ledgers_{len(products)}products_{filename_param}_{stamp}.xlsx"
    )


# ----------------------------------------------------------- customer prices
# The same CustomerPrice grid seen from either side: all customers for one
# product, or all products for one customer. Both save through the endpoint
# below, so the two pages can never disagree about what a save does.

#: Timestamp format shared by the first paint and the AJAX reply — a saved row
#: must not drift out of step with the untouched rows around it.
UPDATED_FORMAT = "M j, Y g:i a"

#: Ceiling on one Save All. Comfortably above a full product list, but stops a
#: hand-rolled payload from turning into an unbounded write.
MAX_BULK_ROWS = 500

MALFORMED = "Malformed request — reload the page and try again."


def _format_updated(dt):
    return date_format(timezone.localtime(dt), UPDATED_FORMAT)


@login_required
def product_prices(request, pk):
    """Every customer's negotiated price for one product."""
    product = get_object_or_404(Product.objects.select_related("category"), pk=pk)

    existing = {
        cp.customer_id: cp for cp in CustomerPrice.objects.filter(product=product)
    }

    # Every active party may be priced: a supplier may buy from us too, so
    # negotiated prices for them are a real thing. Only inactive accounts are
    # off the list — unless one is already priced, since hiding the row
    # would hide the data.
    customers = Customer.objects.filter(
        Q(is_active=True) | Q(pk__in=list(existing))
    )

    rows = []
    for customer in customers:
        price = existing.get(customer.pk)
        rows.append(
            {
                "customer": customer,
                "has_custom": price is not None,
                "unit_price": price.unit_price if price else None,
                "updated_at": _format_updated(price.updated_at) if price else "",
            }
        )

    return render(
        request,
        "core/product_prices.html",
        {
            "product": product,
            "rows": rows,
            "custom_count": len(existing),
        },
    )


@login_required
def customer_prices(request, pk):
    """Every product's negotiated price for one customer."""
    customer = get_object_or_404(Customer, pk=pk)

    existing = {
        cp.product_id: cp for cp in CustomerPrice.objects.filter(customer=customer)
    }

    # Same reasoning as above: a retired product that still carries a custom
    # price for this customer stays visible.
    products = Product.objects.select_related("category").filter(
        Q(is_active=True) | Q(pk__in=list(existing))
    )

    rows = []
    for product in products:
        price = existing.get(product.pk)
        rows.append(
            {
                "product": product,
                "has_custom": price is not None,
                "unit_price": price.unit_price if price else None,
                "updated_at": _format_updated(price.updated_at) if price else "",
            }
        )

    return render(
        request,
        "core/customer_prices.html",
        {
            "customer": customer,
            "rows": rows,
            "custom_count": len(existing),
        },
    )


@require_POST
@login_required
def customer_price_save_all(request):
    """Save every edited row of a price table in one request.

    Takes JSON: {"rows": [{"customer_id", "product_id", "unit_price"}, ...]}.
    The whole batch is validated before anything is written, and the write is
    wrapped in one transaction: a single bad row saves nothing. Half-applying
    a batch would leave the operator guessing which half landed.
    """
    try:
        payload = json.loads(request.body or b"{}")
    except json.JSONDecodeError:
        return JsonResponse({"success": False, "error": MALFORMED}, status=400)

    rows = payload.get("rows") if isinstance(payload, dict) else None
    if not isinstance(rows, list):
        return JsonResponse({"success": False, "error": MALFORMED}, status=400)
    if not rows:
        return JsonResponse({"success": False, "error": "No changes to save."}, status=400)
    if len(rows) > MAX_BULK_ROWS:
        return JsonResponse(
            {
                "success": False,
                "error": f"Too many rows in one save (limit {MAX_BULK_ROWS}).",
            },
            status=400,
        )

    valid = []
    errors = []
    for row in rows:
        if not isinstance(row, dict):
            return JsonResponse({"success": False, "error": MALFORMED}, status=400)

        form = CustomerPriceForm(row)
        if form.is_valid():
            valid.append(form.cleaned_data)
        else:
            # Echo the ids back so the page can pin the message to its row.
            errors.append(
                {
                    "customer_id": str(row.get("customer_id", "")),
                    "product_id": str(row.get("product_id", "")),
                    "error": form.first_error(),
                }
            )

    if errors:
        count = len(errors)
        return JsonResponse(
            {
                "success": False,
                "error": f"Nothing saved — fix {count} row{'' if count == 1 else 's'} and try again.",
                "errors": errors,
            },
            status=400,
        )

    results = []
    with transaction.atomic():
        for data in valid:
            # update_or_create leans on the (customer, product) unique
            # constraint, so two operators saving the same row race to an
            # update rather than to a duplicate row or an IntegrityError.
            price, created = CustomerPrice.objects.update_or_create(
                customer=data["customer_id"],
                product=data["product_id"],
                defaults={"unit_price": data["unit_price"]},
            )
            results.append(
                {
                    "customer_id": str(price.customer_id),
                    "product_id": str(price.product_id),
                    "created": created,
                    "price": f"{price.unit_price:.2f}",
                    "updated_at": _format_updated(price.updated_at),
                }
            )

    return JsonResponse({"success": True, "saved": len(results), "results": results})


# ---------------------------------------------------------------- customers
# Managers may list, view, create and edit. Deleting is super-admin only, and
# so is the credit limit — see CustomerForm.


def _customers():
    """Customers annotated with everything the list and detail pages report.

    `owed` is the positive amount a debtor owes us: balances run negative when
    the customer owes, positive when we owe them.

    `pending_cheques` is the sum of Cheque.amount for cheques we're holding
    against this customer that haven't yet been banked. They were already
    added to Customer.balance when received — so the ledger sees them as
    paid — but for credit-limit purposes they are still promises, not money.
    We add them back to the effective debt so a customer can't stack the
    limit up with pending cheques and buy again on top of them. Once one
    of those cheques deposits, it drops out of this sum and the available
    credit rises automatically.

    Available credit = credit_limit − (owed + pending_cheques), floored at
    zero — someone past their limit has none left, never a negative amount.
    """
    owed = Case(
        When(balance__lt=0, then=Value(0) - F("balance")),
        default=Value(ZERO),
        output_field=MONEY,
    )
    return (
        Customer.objects.annotate(owed=owed)
        .annotate(
            pending_cheques=Coalesce(
                Sum(
                    "cheques__amount",
                    filter=Q(cheques__status=Cheque.Status.PENDING),
                    output_field=MONEY,
                ),
                ZERO,
                output_field=MONEY,
            )
        )
        .annotate(
            available_credit=Greatest(
                F("credit_limit") - F("owed") - F("pending_cheques"),
                Value(ZERO),
                output_field=MONEY,
            )
        )
        # distinct=True: without it these four joins multiply each other's rows
        # and every count comes out inflated.
        .annotate(
            bill_count=Count("bills", distinct=True),
            supplier_bill_count=Count("supplier_bills", distinct=True),
            cheque_count=Count("cheques", distinct=True),
            custom_price_count=Count("custom_prices", distinct=True),
        )
        .annotate(
            history_count=F("bill_count") + F("supplier_bill_count") + F("cheque_count")
        )
        # Repeats Customer.Meta.ordering because the annotate()s above group,
        # and a grouped query loses Meta.ordering — see _bills_with_counts.
        .order_by("name")
    )


def _delete_blockers(customer):
    """History that PROTECTs this customer, as printable phrases.

    Takes a customer from `_customers()` — the counts are already annotated.
    """
    counts = [
        (customer.bill_count, "bill"),
        (customer.supplier_bill_count, "supplier bill"),
        (customer.cheque_count, "cheque"),
    ]
    phrases = [f"{n} {label}{'' if n == 1 else 's'}" for n, label in counts if n]

    if len(phrases) <= 1:
        return "".join(phrases)
    return f"{', '.join(phrases[:-1])} and {phrases[-1]}"


@login_required
def customer_list(request):
    query = request.GET.get("q", "").strip()

    # Unrecognised filter values are dropped rather than 500ing or silently
    # showing an unfiltered list that claims to be filtered.
    kind = request.GET.get("kind", "").strip()
    if kind not in {"customers", "suppliers"}:
        kind = ""
    status = request.GET.get("status", "").strip()
    if status not in {"active", "inactive"}:
        status = ""

    customers = _customers()
    if query:
        customers = customers.filter(name__icontains=query)
    if kind:
        customers = customers.filter(is_supplier=kind == "suppliers")
    if status:
        customers = customers.filter(is_active=status == "active")

    page_obj = _paginate(request, customers)

    # Summary cards — aggregated over the *full* filtered set, not just this page,
    # so the totals are always consistent regardless of pagination.
    # balance < 0  → customer owes us   (outstanding)
    # balance > 0  → we owe the customer
    balance_totals = customers.aggregate(
        total_outstanding=Coalesce(
            Sum(
                Case(
                    When(balance__lt=0, then=Value(0) - F("balance")),
                    default=Value(ZERO),
                    output_field=MONEY,
                )
            ),
            ZERO,
            output_field=MONEY,
        ),
        we_owe=Coalesce(
            Sum(
                Case(
                    When(balance__gt=0, then=F("balance")),
                    default=Value(ZERO),
                    output_field=MONEY,
                )
            ),
            ZERO,
            output_field=MONEY,
        ),
    )
    customer_stats = {
        "total_outstanding": balance_totals["total_outstanding"],
        "we_owe": balance_totals["we_owe"],
    }

    # Supplier management stats — cheap aggregates over the full supplier set
    # (not just this page) so the header reads the same on every page.
    supplier_stats = None
    if kind == "suppliers":
        supplier_qs = Customer.objects.filter(is_supplier=True, is_walk_in_account=False)
        totals = supplier_qs.aggregate(
            total_count=Count("pk"),
            total_owed_to=Coalesce(
                Sum(Case(
                    When(balance__gt=0, then=F("balance")),
                    default=Value(ZERO),
                    output_field=MONEY,
                )),
                ZERO,
                output_field=MONEY,
            ),
            total_owed_by=Coalesce(
                Sum(Case(
                    When(balance__lt=0, then=Value(0) - F("balance")),
                    default=Value(ZERO),
                    output_field=MONEY,
                )),
                ZERO,
                output_field=MONEY,
            ),
        )
        supplier_stats = {
            "count": totals["total_count"],
            "we_owe": totals["total_owed_to"],
            "they_owe": totals["total_owed_by"],
        }

    # For the browser's native <datalist> autocomplete on the search box.
    # A flat list of names — the datalist shows matches as the operator
    # types and Enter submits the form as it already would. Suppliers /
    # inactive customers are still in the list because searching for one
    # you can't see anywhere else is still useful.
    suggest_names = list(
        Customer.objects.filter(is_walk_in_account=False)
        .order_by("name")
        .values_list("name", flat=True)
    )

    return render(
        request,
        "core/customer_list.html",
        {
            "page_obj": page_obj,
            "customers": page_obj.object_list,
            "query": query,
            "kind": kind,
            "status": status,
            "is_filtered": bool(query or kind or status),
            "customer_stats": customer_stats,
            "supplier_stats": supplier_stats,
            "suggest_names": suggest_names,
        },
    )



@login_required
def customer_contacts(request):
    """Customer Details page — search, view and quick-edit contact info
    (name, phone, email, address) for any existing customer.
    This is a dedicated contacts directory; it does NOT create new customers.
    """
    query = request.GET.get("q", "").strip()
    pk = request.GET.get("pk")
    selected = None
    save_success = False

    # Build the list (exclude walk-in holding account)
    customers = Customer.objects.filter(is_walk_in_account=False).order_by("name")
    if query:
        customers = customers.filter(
            Q(name__icontains=query)
            | Q(phone__icontains=query)
            | Q(email__icontains=query)
        )

    # Load the selected customer for the detail/edit panel
    if pk:
        try:
            selected = Customer.objects.get(pk=pk, is_walk_in_account=False)
        except Customer.DoesNotExist:
            selected = None

    # Handle inline edit POST
    if request.method == "POST" and selected:
        # Only update contact fields — never touch balance/credit_limit/flags
        selected.name = request.POST.get("name", selected.name).strip() or selected.name
        selected.phone = request.POST.get("phone", "").strip()
        selected.email = request.POST.get("email", "").strip()
        selected.address = request.POST.get("address", "").strip()
        selected.save(update_fields=["name", "phone", "email", "address"])
        messages.success(request, f"Contact details for '{selected.name}' saved.")
        save_success = True

    return render(
        request,
        "core/customer_contacts.html",
        {
            "customers": customers,
            "query": query,
            "selected": selected,
            "save_success": save_success,
        },
    )


@login_required
def customer_list_excel(request):
    """Export the list of customers/suppliers to an Excel sheet.
    Respects any active filters (query q, kind, status).
    """
    from io import BytesIO
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from django.utils import timezone

    query = request.GET.get("q", "").strip()
    kind = request.GET.get("kind", "").strip()
    if kind not in {"customers", "suppliers"}:
        kind = ""
    status = request.GET.get("status", "").strip()
    if status not in {"active", "inactive"}:
        status = ""

    customers = _customers()
    if query:
        customers = customers.filter(name__icontains=query)
    if kind:
        customers = customers.filter(is_supplier=kind == "suppliers")
    if status:
        customers = customers.filter(is_active=status == "active")

    # Aggregate stats over the exported set
    balance_totals = customers.aggregate(
        total_outstanding=Coalesce(
            Sum(
                Case(
                    When(balance__lt=0, then=Value(0) - F("balance")),
                    default=Value(ZERO),
                    output_field=MONEY,
                )
            ),
            ZERO,
            output_field=MONEY,
        ),
        we_owe=Coalesce(
            Sum(
                Case(
                    When(balance__gt=0, then=F("balance")),
                    default=Value(ZERO),
                    output_field=MONEY,
                )
            ),
            ZERO,
            output_field=MONEY,
        ),
        total_limit=Coalesce(Sum("credit_limit"), ZERO, output_field=MONEY),
        total_avail=Coalesce(Sum("available_credit"), ZERO, output_field=MONEY),
        total_count=Count("pk"),
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Customer List"

    # Style definitions
    thin = Side(style="thin")
    border_all = Border(top=thin, bottom=thin, left=thin, right=thin)
    bold = Font(bold=True)
    bold_large = Font(bold=True, size=14)
    center = Alignment(horizontal="center", vertical="center")
    right = Alignment(horizontal="right", vertical="center")
    left = Alignment(horizontal="left", vertical="center")

    header_fill = PatternFill("solid", fgColor="1F2937") # Dark gray
    header_font = Font(bold=True, color="FFFFFF")
    
    # Highlight lines where balance indicates they owe us (outstanding) or we owe them
    fill_outstanding = PatternFill("solid", fgColor="FFF1F2") # Rose tint
    fill_we_owe = PatternFill("solid", fgColor="ECFDF5") # Emerald tint
    fill_normal = PatternFill("solid", fgColor="FFFFFF")

    # Set Column Widths (8 columns A-H)
    ws.column_dimensions["A"].width = 30 # Name
    ws.column_dimensions["B"].width = 16 # Phone
    ws.column_dimensions["C"].width = 32 # Address
    ws.column_dimensions["D"].width = 16 # Balance
    ws.column_dimensions["E"].width = 16 # Credit Limit
    ws.column_dimensions["F"].width = 18 # Available Credit
    ws.column_dimensions["G"].width = 15 # Type
    ws.column_dimensions["H"].width = 12 # Status

    # --- Title Block ---
    report_title = "Senovka Plastics — Customer List"
    if kind == "suppliers":
        report_title = "Senovka Plastics — Supplier List"
    ws["A1"] = report_title
    ws["A1"].font = bold_large
    ws.merge_cells("A1:H1")

    now = timezone.localtime(timezone.now())
    ws["A2"] = f"Exported: {now.strftime('%d %b %Y, %I:%M %p')}"
    ws["A2"].font = Font(italic=True)
    ws.merge_cells("A2:H2")

    # --- Summary Metrics Block ---
    ws["A4"] = "Total Count"; ws["B4"] = balance_totals["total_count"]
    ws["C4"] = "Total Outstanding (They Owe)"; ws["D4"] = float(balance_totals["total_outstanding"])
    ws["E4"] = "Total We Owe"; ws["F4"] = float(balance_totals["we_owe"])
    ws["G4"] = "Total Credit Limit"; ws["H4"] = float(balance_totals["total_limit"])

    ws["B4"].font = bold; ws["D4"].font = bold; ws["F4"].font = bold; ws["H4"].font = bold
    for cell_addr in ["A4", "C4", "E4", "G4"]:
        ws[cell_addr].font = Font(bold=True)

    # --- Table Header ---
    headers = [
        "Name", "Phone", "Address", "Balance", "Credit Limit", 
        "Available Credit", "Type", "Status"
    ]
    header_row = 6
    for idx, name in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=idx, value=name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border_all

    # --- Item Rows ---
    row_num = header_row + 1
    for customer in customers:
        # Determine styling & coloring based on balance
        row_fill = fill_normal
        if customer.balance < 0:
            row_fill = fill_outstanding
        elif customer.balance > 0:
            row_fill = fill_we_owe

        # Type text
        if customer.is_supplier:
            cust_type = "Supplier"
        else:
            cust_type = "Customer"

        # Status text
        cust_status = "Active" if customer.is_active else "Inactive"

        # Write cells
        c_name = ws.cell(row=row_num, column=1, value=customer.name)
        c_phone = ws.cell(row=row_num, column=2, value=customer.phone or "—")
        c_addr = ws.cell(row=row_num, column=3, value=customer.address or "—")
        c_bal = ws.cell(row=row_num, column=4, value=float(customer.balance))
        c_limit = ws.cell(row=row_num, column=5, value=float(customer.credit_limit))
        c_avail = ws.cell(row=row_num, column=6, value=float(customer.available_credit))
        c_type = ws.cell(row=row_num, column=7, value=cust_type)
        c_status = ws.cell(row=row_num, column=8, value=cust_status)

        # Formatting
        c_name.alignment = left
        c_phone.alignment = center
        c_addr.alignment = left
        c_bal.alignment = right
        c_limit.alignment = right
        c_avail.alignment = right
        c_type.alignment = center
        c_status.alignment = center

        for cell in (c_bal, c_limit, c_avail):
            cell.number_format = "#,##0.00"

        # Apply borders and fills to all cells in the row
        for c_idx in range(1, 9):
            cell = ws.cell(row=row_num, column=c_idx)
            cell.border = border_all
            cell.fill = row_fill

        row_num += 1

    # Add double line for totals at the bottom
    totals_row = row_num
    ws.cell(row=totals_row, column=1, value="Totals").font = bold
    ws.cell(row=totals_row, column=4, value=float(sum((c.balance for c in customers), ZERO))).font = bold
    ws.cell(row=totals_row, column=5, value=float(balance_totals["total_limit"])).font = bold
    ws.cell(row=totals_row, column=6, value=float(balance_totals["total_avail"])).font = bold

    ws.cell(row=totals_row, column=4).number_format = "#,##0.00"
    ws.cell(row=totals_row, column=5).number_format = "#,##0.00"
    ws.cell(row=totals_row, column=6).number_format = "#,##0.00"

    ws.cell(row=totals_row, column=4).alignment = right
    ws.cell(row=totals_row, column=5).alignment = right
    ws.cell(row=totals_row, column=6).alignment = right

    double_border = Border(top=thin, bottom=Side(style="double"))
    for c_idx in range(1, 9):
        cell = ws.cell(row=totals_row, column=c_idx)
        cell.border = double_border

    # --- Write stream ---
    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    response = HttpResponse(
        stream.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    
    file_label = "customers" if kind != "suppliers" else "suppliers"
    filename_date = now.strftime("%Y-%m-%d")
    response["Content-Disposition"] = (
        f'attachment; filename="{file_label}_list_{filename_date}.xlsx"'
    )
    return response


@login_required
def customer_detail(request, pk):
    customer = get_object_or_404(_customers(), pk=pk)

    # The adjustments table and its modal are super-admin only, but the query
    # runs regardless so a manager viewing the page does not see a suddenly
    # missing section on reload after a role change.
    adjustments = customer.balance_adjustments.select_related("adjusted_by")
    page_obj = _paginate(request, adjustments)

    return render(
        request,
        "core/customer_detail.html",
        {
            "customer": customer,
            "blockers": _delete_blockers(customer),
            "adjustment_form": CustomerBalanceAdjustmentForm(),
            "adjustments_page": page_obj,
            "adjustments": page_obj.object_list,
        },
    )


def _apply_adjustment(customer, adjustment_type, amount):
    """Move Customer.balance by one adjustment, in the direction of its type.

    Not called anywhere Customer.balance is already touched — this is the only
    thing that should ever move it from an adjustment, so any change to that
    rule lives here.
    """
    if adjustment_type == CustomerBalanceAdjustment.Type.CREDIT:
        Customer.objects.filter(pk=customer.pk).update(balance=F("balance") + amount)
    else:
        Customer.objects.filter(pk=customer.pk).update(balance=F("balance") - amount)


def _reverse_adjustment(customer, adjustment_type, amount):
    """Undo one adjustment's effect on Customer.balance.

    An edit reverses the old row before applying the new one, and a delete
    reverses the row on its way out. Kept separate from `_apply_adjustment`
    even though it is the mirror: reading `reverse` at the call site is
    clearer than reading `apply(opposite_type)`.
    """
    if adjustment_type == CustomerBalanceAdjustment.Type.CREDIT:
        Customer.objects.filter(pk=customer.pk).update(balance=F("balance") - amount)
    else:
        Customer.objects.filter(pk=customer.pk).update(balance=F("balance") + amount)


@require_POST
@super_admin_required
def customer_adjustment_create(request, pk):
    customer = get_object_or_404(Customer, pk=pk)
    form = CustomerBalanceAdjustmentForm(request.POST)

    if not form.is_valid():
        messages.error(request, f"Adjustment not saved: {form.first_error()}")
        return redirect("core:customer_detail", pk=customer.pk)

    with transaction.atomic():
        adjustment = form.save(commit=False)
        adjustment.customer = customer
        adjustment.adjusted_by = request.user
        adjustment.save()
        _apply_adjustment(customer, adjustment.adjustment_type, adjustment.amount)

    sign = "+" if adjustment.adjustment_type == CustomerBalanceAdjustment.Type.CREDIT else "-"
    messages.success(
        request,
        f"Balance adjustment {sign}{adjustment.amount:,.2f} recorded for "
        f"{customer.name}.",
    )
    return redirect("core:customer_detail", pk=customer.pk)


@require_POST
@super_admin_required
def customer_adjustment_edit(request, pk, adjustment_pk):
    customer = get_object_or_404(Customer, pk=pk)
    adjustment = get_object_or_404(
        CustomerBalanceAdjustment, pk=adjustment_pk, customer=customer
    )
    form = CustomerBalanceAdjustmentForm(request.POST, instance=adjustment)

    if not form.is_valid():
        messages.error(request, f"Adjustment not saved: {form.first_error()}")
        return redirect("core:customer_detail", pk=customer.pk)

    with transaction.atomic():
        # Reverse the row as it stood before the edit, so the balance walks
        # from the pre-edit total to the new one and never counts either the
        # old amount or the new amount twice — the same shape as _reverse_bill.
        _reverse_adjustment(customer, adjustment.adjustment_type, adjustment.amount)

        edited = form.save(commit=False)
        # adjusted_by tracks the person of record for the current values, not
        # the person who wrote the row first. That matches how BillEditAudit
        # attributes edits.
        edited.adjusted_by = request.user
        edited.save()
        _apply_adjustment(customer, edited.adjustment_type, edited.amount)

    messages.success(request, f"Adjustment updated for {customer.name}.")
    return redirect("core:customer_detail", pk=customer.pk)


@require_POST
@super_admin_required
def customer_adjustment_delete(request, pk, adjustment_pk):
    customer = get_object_or_404(Customer, pk=pk)
    adjustment = get_object_or_404(
        CustomerBalanceAdjustment, pk=adjustment_pk, customer=customer
    )

    with transaction.atomic():
        _reverse_adjustment(customer, adjustment.adjustment_type, adjustment.amount)
        adjustment.delete()

    messages.success(request, f"Adjustment removed from {customer.name}.")
    return redirect("core:customer_detail", pk=customer.pk)


@login_required
def customer_create(request):
    form = CustomerForm(
        request.POST or None, is_super_admin=_is_super_admin(request.user)
    )

    if request.method == "POST" and form.is_valid():
        customer = form.save()
        messages.success(request, f"Customer '{customer.name}' was created.")
        return redirect("core:customer_list")

    return render(
        request,
        "core/customer_form.html",
        {"form": form, "is_edit": False},
    )


@login_required
def customer_update(request, pk):
    customer = get_object_or_404(Customer, pk=pk)
    form = CustomerForm(
        request.POST or None,
        instance=customer,
        is_super_admin=_is_super_admin(request.user),
    )

    if request.method == "POST" and form.is_valid():
        customer = form.save()
        messages.success(request, f"Customer '{customer.name}' was updated.")
        return redirect("core:customer_list")

    return render(
        request,
        "core/customer_form.html",
        {"form": form, "customer": customer, "is_edit": True},
    )


def _parse_date(raw):
    """A GET date, or None. An unparsable one is ignored rather than 500ing."""
    raw = (raw or "").strip()
    if not raw:
        return None
    try:
        return parse_date(raw)
    except ValueError:
        # Well-formed but impossible, e.g. 2026-02-31.
        return None


def _ledger_rows(customer, from_date=None, to_date=None):
    """Every ledger line for one customer, oldest first, with a running balance.

    Four sources land in one column layout:
      Sale      — a bill, what they now owe us            -> SALE
      Payment   — cash/cheque taken against a bill        -> CHE/CASH
      Purchase  — a supplier bill, what we owe them back  -> CHE/CASH
      Edit note — a bill was rewritten, and why           -> neither

    The edit note is the odd one out: it carries no amount, because an edit's
    figures are already in the sale row it rewrote. It is here so the ledger
    can account for a figure changing under a reader who saw the old one.

    The running balance is accumulated here rather than in the template: each
    row depends on every row before it, which a template cannot express
    without carrying state.

    Cancelled bills and their payments are left out — a cancelled sale is not
    owed, so including it would overstate the balance.
    """
    entries = []

    for bill in customer.bills.exclude(status=Bill.Status.CANCELLED):
        # Delivery and discount are folded into total_amount already; spelled
        # out here so the ledger reads as more than a bare "Sale" whenever
        # either moved the figure.
        extra = []
        if bill.delivery_charge:
            extra.append(f"+{bill.delivery_charge:.2f} delivery")
        if bill.discount_amount:
            extra.append(f"-{bill.discount_amount:.2f} discount")
        description = f"Sale ({', '.join(extra)})" if extra else "Sale"

        entries.append(
            {
                "date": bill.bill_date,
                "kind": 0,  # a sale precedes same-day money against it
                "pk": bill.pk,
                "description": description,
                "sale": bill.total_amount,
                "credit": None,
                "is_note": False,
                # Carried through so the ledger row can offer a "Pay" link on a
                # bill that still owes money. Everything else lands as None so
                # the template can guard on it with a single check.
                "bill_pk": bill.pk,
                "remaining": bill.remaining_balance,
            }
        )

    # Cancelled bills are excluded above and their notes go with them: a
    # cancelled sale isn't in the ledger, so the story of how it was corrected
    # has nothing left to annotate.
    audits = BillEditAudit.objects.filter(bill__customer=customer).exclude(
        bill__status=Bill.Status.CANCELLED
    )
    for audit in audits:
        entries.append(
            {
                "date": audit.edit_date,
                # Last on its day: the note explains rows already read, and
                # sorting it into the middle of them would imply it split the
                # day's money in two.
                "kind": 3,
                "pk": audit.pk,
                "description": f"Bill #{audit.bill_id} edited: {audit.reason}",
                # Both None is what makes this a note. The running balance
                # below adds 0 and carries the previous row's figure forward,
                # which is exactly what a note should do to an account.
                "sale": None,
                "credit": None,
                "is_note": True,
            }
        )

    for supplier_bill in customer.supplier_bills.exclude(
        status=SupplierBill.Status.CANCELLED
    ):
        note = supplier_bill.notes.strip()
        entries.append(
            {
                "date": supplier_bill.bill_date,
                "kind": 1,
                "pk": supplier_bill.pk,
                "description": f"Purchase - {note}" if note else "Purchase",
                "sale": None,
                "credit": supplier_bill.total_amount,
                "is_note": False,
            }
        )

    payments = (
        Payment.objects.filter(bill__customer=customer)
        .exclude(bill__status=Bill.Status.CANCELLED)
        # A cheque that bounced or is being held never became money, and the
        # customer's balance has had it taken back off. Leaving the payment
        # here would walk the running balance away from the account itself.
        .exclude(
            cheques__status__in=[Cheque.Status.BOUNCED, Cheque.Status.HELD]
        )
        .select_related("bill")
    )
    for payment in payments:
        entries.append(
            {
                # paid_at is a moment; the ledger reports days.
                "date": timezone.localdate(payment.paid_at),
                "kind": 2,
                "pk": payment.pk,
                "description": f"{payment.get_method_display()} received",
                "sale": None,
                "credit": payment.amount,
                "is_note": False,
            }
        )

    # Detached payments — money booked against the customer without a bill
    # behind it (opening-balance settlement, top-up sitting as credit). See
    # _allocate_settlement's spillover.
    direct_payments = (
        Payment.objects.filter(bill__isnull=True, customer=customer)
        .exclude(
            cheques__status__in=[Cheque.Status.BOUNCED, Cheque.Status.HELD]
        )
    )
    for payment in direct_payments:
        entries.append(
            {
                "date": timezone.localdate(payment.paid_at),
                "kind": 2,
                "pk": payment.pk,
                "description": f"{payment.get_method_display()} received (against balance)",
                "sale": None,
                "credit": payment.amount,
                "is_note": False,
            }
        )

    # Manual adjustments book like a payment or a charge — a credit reduces
    # what the customer owes us, a debit adds to it — so they share the sale
    # and credit columns. The description carries the reason so a reader who
    # only has the printed ledger can see why the balance moved.
    for adjustment in customer.balance_adjustments.all():
        is_credit = adjustment.adjustment_type == CustomerBalanceAdjustment.Type.CREDIT
        sign = "+" if is_credit else "-"
        entries.append(
            {
                "date": adjustment.adjustment_date,
                "kind": 2,
                "pk": adjustment.pk,
                "description": f"Balance Adjustment ({sign}) — {adjustment.reason}",
                "sale": None if is_credit else adjustment.amount,
                "credit": adjustment.amount if is_credit else None,
                "is_note": False,
            }
        )

    if from_date:
        entries = [e for e in entries if e["date"] >= from_date]
    if to_date:
        entries = [e for e in entries if e["date"] <= to_date]

    # pk breaks the last tie, so two rows on one day never swap between loads.
    entries.sort(key=lambda e: (e["date"], e["kind"], e["pk"]))

    balance = ZERO
    for entry in entries:
        balance += (entry["sale"] or ZERO) - (entry["credit"] or ZERO)
        entry["balance"] = balance

    return entries


@login_required
def customer_ledger(request, pk):
    customer = get_object_or_404(_customers(), pk=pk)

    from_date = _parse_date(request.GET.get("from_date"))
    to_date = _parse_date(request.GET.get("to_date"))
    rows = _ledger_rows(customer, from_date, to_date)

    # _ledger_rows stays whole and the totals below are taken over all of it:
    # the running balance in each row is the sum of every row before it, and
    # the closing balance is the last row's. Only the display is paged.
    #
    # The PDF calls _ledger_rows itself and never sees this — a printed ledger
    # is the whole account, not page 1 of it.
    page_obj = _paginate(request, rows, settings.PAGINATE_BY_REPORTS)

    return render(
        request,
        "core/customer_ledger.html",
        {
            "customer": customer,
            "page_obj": page_obj,
            "rows": page_obj.object_list,
            "from_date": from_date,
            "to_date": to_date,
            "is_filtered": bool(from_date or to_date),
            "total_sale": sum((r["sale"] or ZERO for r in rows), ZERO),
            "total_credit": sum((r["credit"] or ZERO for r in rows), ZERO),
            "closing_balance": rows[-1]["balance"] if rows else ZERO,
        },
    )


@require_POST
@super_admin_required
def customer_delete(request, pk):
    customer = get_object_or_404(_customers(), pk=pk)
    name = customer.name

    blockers = _delete_blockers(customer)
    if blockers:
        # Bills, supplier bills and cheques all PROTECT their customer, so
        # anything with trading history stays put. Say what is holding it
        # rather than letting the delete fail at the database.
        messages.error(
            request,
            f"Cannot delete '{name}' — it still has {blockers}. "
            f"Deactivate it instead to hide it from new bills.",
        )
        return redirect("core:customer_list")

    try:
        customer.delete()
    except ProtectedError:
        # Belt and braces: a PROTECTed relation added later lands here rather
        # than as a 500.
        messages.error(
            request,
            f"Cannot delete '{name}' — other records still reference it. "
            f"Deactivate it instead.",
        )
    else:
        messages.success(request, f"Customer '{name}' was deleted.")

    return redirect("core:customer_list")


# -------------------------------------------------------------------- bills
# Bill creation is one page: picking a customer pulls their own prices over
# AJAX, so nothing reloads between steps.


def _qty_text(value):
    """Stock as the product list prints it: up to 3 dp, no trailing zeros."""
    text = f"{value:.3f}".rstrip("0").rstrip(".")
    return text or "0"


#: What the walk-in holding account is called when it is first created. Only a
#: starting name — the account is found by its flag, so renaming it is safe.
WALK_IN_ACCOUNT_NAME = "Walk-in Customer"


def _walk_in_customer():
    """The one account every walk-in bill hangs off, made on first use.

    Created with no credit limit: a walk-in is a counter sale, and nobody is
    extending credit to a name on a slip.
    """
    customer = Customer.objects.filter(is_walk_in_account=True).order_by("pk").first()
    if customer is not None:
        return customer
    return Customer.objects.create(
        name=WALK_IN_ACCOUNT_NAME,
        is_walk_in_account=True,
        is_supplier=False,
        is_active=True,
        credit_limit=ZERO,
    )


def _billable_customers():
    """Who a bill may be made out to, priced against their current balance.

    Suppliers *are* offered: a party we buy from may also buy from us, and
    the ledger already tracks the net through Customer.balance — a supplier
    bill moves balance one way (they gave us goods, we owe them), a sales
    bill moves it the other (we gave them goods, they owe us).

    Inactive accounts and the walk-in holding account are still excluded —
    walk-in is reached by the Walk-in toggle, and offering it here would let
    a bill be booked to the holding account without a name.
    """
    return list(
        Customer.objects.filter(is_active=True, is_walk_in_account=False)
    )


def _bill_form_context(request, customers):
    """What both bill pages need to draw the form."""
    return {
        "customers": customers,
        "today": timezone.localdate(),
        # The page prices a walk-in against the holding account, which has no
        # CustomerPrice rows — so the endpoint quotes it default prices. Created
        # on first sight of a bill form rather than on first walk-in sale, so
        # the page always has an id to ask about.
        "walk_in_customer_id": _walk_in_customer().pk,
        # Tabs over the product grid. Driven off the table rather than a fixed
        # list, because categories are operator-managed — a hardcoded tab would
        # be wrong the first time one is added or renamed.
        "categories": Category.objects.all(),
        # Straight off the models, so the radio values and account codes the
        # page posts are the ones the save step will store. The legacy
        # "partial" (cash + cheque, full amount) type is left off — it has
        # been split into PARTIAL_CASH and PARTIAL_CHEQUE. The save step
        # still accepts it for editing bills written before the split.
        "payment_types": [
            (value, label)
            for value, label in Bill.PaymentType.choices
            if value != Bill.PaymentType.PARTIAL
        ],
        "account_choices": Payment.Account.choices,
    }


@login_required
def bill_create(request):
    customers = _billable_customers()
    for customer in customers:
        # A new bill prices against the balance as it stands. The edit page
        # sets this differently, which is the only difference between them.
        customer.balance_for_bill = customer.balance

    context = _bill_form_context(request, customers)
    context.update(
        {
            "save_url": reverse("core:bill_save"),
            "is_edit": False,
        }
    )
    return render(request, "core/bill_create.html", context)


@require_GET
@login_required
def bill_products(request, customer_id):
    """What this customer can be billed for, at their own price.

    Feeds the product grid. Inactive products are left out entirely — they
    cannot go on a bill. Out-of-stock ones are reported with qty 0 and drawn as
    dimmed, unsellable cards: the biller looking for a product is better told it
    exists and is finished than left to wonder whether it was ever set up. The
    save step refuses them regardless of what the page allows.

    ?bill=<id> asks the same question for an edit, where the bill's own lines
    have already taken their stock. Those quantities are added back and the
    products kept in the list however low they have run, otherwise a bill that
    cleared the shelf could never be edited.
    """
    customer = Customer.objects.filter(pk=customer_id).first()
    if customer is None:
        return JsonResponse({"error": "That customer no longer exists."}, status=404)

    editing = request.GET.get("bill", "").strip()
    held = {}
    if editing.isdigit():
        held = {
            item.product_id: item.qty
            for item in BillItem.objects.filter(bill_id=int(editing))
        }

    # One query for this customer's overrides, then one for the products. The
    # alternative is a CustomerPrice lookup per row.
    #
    # order_by() clears CustomerPrice.Meta.ordering, which sorts by customer
    # and product name and so drags both tables into a join this lookup has no
    # use for — the rows land in a dict either way.
    overrides = dict(
        CustomerPrice.objects.filter(customer=customer)
        .order_by()
        .values_list("product_id", "unit_price")
    )

    sellable = Product.objects.filter(
        Q(is_active=True) | Q(pk__in=list(held))
    ).select_related("category")

    threshold = Decimal(str(settings.LOW_STOCK_THRESHOLD))

    products = []
    for product in sellable:
        override = overrides.get(product.pk)
        unit_price = product.default_price if override is None else override
        # Stock this bill is already holding is stock this bill may still use.
        available = product.qty + held.get(product.pk, ZERO)
        products.append(
            {
                "id": product.pk,
                "name": product.name,
                "size": product.size,
                "qty": _qty_text(available),
                # Numeric copy so the JS does not have to re-parse for the
                # live "Available: N" reflection — the string is kept for the
                # existing "N in stock" label.
                "qty_number": float(available),
                "unit_price": f"{unit_price:.2f}",
                "has_custom_price": override is not None,
                # Server-computed flags so every UI reads off the same
                # threshold. Zero and negative are both "out of stock"; the
                # card refuses selection either way.
                "is_out_of_stock": available <= ZERO,
                "is_low_stock": ZERO < available <= threshold,
                # Drives the category tabs over the grid. The id is what the
                # tabs match on — names are operator-entered and change.
                "category_id": product.category_id,
                "category": product.category.name,
            }
        )

    # Wrapped now: the array is joined by the low-stock threshold, so both the
    # client-side card grid and the item-panel oversell warning read from one
    # response instead of the JS having to know the setting a second way.
    return JsonResponse(
        {
            "products": products,
            "low_stock_threshold": float(threshold),
        }
    )


class BillError(Exception):
    """A reason to roll the save back, worded for the biller."""


def _decimal(raw, label, places):
    """Parse one money or quantity figure off the payload."""
    text = str(raw if raw is not None else "").strip()
    if text == "":
        raise BillError(f"{label} is required.")
    try:
        value = Decimal(text)
    except InvalidOperation:
        raise BillError(f"{label} must be a number.")
    if not value.is_finite():
        raise BillError(f"{label} must be a number.")
    if value < 0:
        raise BillError(f"{label} cannot be negative.")

    step = Decimal("0.01") if places == 2 else Decimal("0.001")
    return value.quantize(step, rounding=ROUND_HALF_UP)


def _read_bill_date(raw):
    """The date a new bill is billed on, off the payload.

    Required — the picker no longer pre-fills today, so a blank here means the
    biller never chose one rather than that today was intended. A future date
    is refused: the goods have not left the yard yet, and a bill dated forward
    would sit ahead of the running balance in every ledger it appears in.
    """
    text = str(raw if raw is not None else "").strip()
    if text == "":
        raise BillError("Bill date is required.")

    parsed = _parse_date(text)
    if parsed is None:
        raise BillError("That bill date isn't a real date.")
    if parsed > timezone.localdate():
        raise BillError("A bill can't be dated in the future.")
    return parsed


def _optional_decimal(raw, label):
    """Blank means nothing was tendered on this leg, not zero-as-an-error."""
    if raw is None or str(raw).strip() == "":
        return ZERO
    return _decimal(raw, label, 2)


def _read_cheque(raw):
    """Validate one cheque row and return the tidy dict for it. Raises
    BillError with a row-scoped message if any required field is missing —
    the caller adds "Cheque N: " so the biller knows which row went wrong.
    """
    raw = raw or {}

    amount = _decimal(raw.get("amount"), "amount", 2)
    if amount <= ZERO:
        raise BillError("amount must be above 0.")

    cheque_no = str(raw.get("cheque_no") or "").strip()
    bank_name = str(raw.get("bank_name") or "").strip()
    if not cheque_no:
        raise BillError("number is required.")
    if not bank_name:
        raise BillError("bank name is required.")

    received = _parse_date(raw.get("received_date"))
    maturity = _parse_date(raw.get("maturity_date"))
    if received is None:
        raise BillError("received date is required.")
    if maturity is None:
        raise BillError("maturity date is required.")
    if maturity < received:
        raise BillError("maturity date cannot be before the received date.")

    return {
        "cheque_no": cheque_no,
        "bank_name": bank_name,
        "branch": str(raw.get("branch") or "").strip(),
        "acc_no": str(raw.get("acc_no") or "").strip(),
        "amount": amount,
        "received_date": received,
        "maturity_date": maturity,
    }


def _read_cheques(raw_list):
    """Validate an array of cheque rows and return (cheques, total).

    Every cheque type takes at least one row, so a missing or empty list is
    refused. Row errors are prefixed with "Cheque N: " so a biller with five
    rows on screen knows which one to fix.
    """
    if not isinstance(raw_list, list) or not raw_list:
        raise BillError("Add at least one cheque.")

    cheques = []
    total = ZERO
    for index, raw in enumerate(raw_list, start=1):
        try:
            cheque = _read_cheque(raw)
        except BillError as exc:
            raise BillError(f"Cheque {index}: {exc}")
        cheques.append(cheque)
        total += cheque["amount"]
    return cheques, total


def _read_cash_account(raw):
    """Where cash on this leg should end up: physical drawer (blank), or one
    of the named accounts. Invalid strings are refused rather than silently
    kept as physical.
    """
    accounts = {value for value, _ in Payment.Account.choices}
    account = str(raw or "").strip()
    if account and account not in accounts:
        raise BillError("Choose a valid account.")
    return account


def _read_payment(raw, total, customer):
    """Re-derive every payment leg from the payload.

    The page validates all of this already; none of that is evidence, so it
    is all recomputed here from the customer's stored balance.

    `total` is what the bill comes to — goods plus delivery, less any
    discount — not the subtotal of the lines. Collecting against the subtotal
    would ask the customer for the delivery they are not paying and refuse
    them the discount they were given.

    Six payment types are recognised. Two — PARTIAL_CASH and PARTIAL_CHEQUE —
    intentionally leave money outstanding without going through Pay Later's
    credit-limit gate: they are how a bill collects some money now and puts
    the rest on account.
    """
    raw = raw or {}
    kind = str(raw.get("type") or "").strip()
    valid = {value for value, _ in Bill.PaymentType.choices}
    if kind not in valid:
        raise BillError("Choose a payment type.")

    parts = {
        "type": kind,
        "cash": ZERO,
        "cash_account": "",
        # Multi-cheque: an array of dicts as returned by _read_cheque, or [].
        "cheques": [],
    }

    # target is what the customer needs to hand over to settle everything:
    # this bill plus any debt, less any credit. Full-payment types collect it
    # or more; partial types collect less than it; Pay Later collects none.
    target = (total - customer.balance).quantize(Decimal("0.01"))

    # Anything handed over above target is not refused: the customer is paying
    # ahead, and the excess lands on their balance as credit the next bill
    # prices against. balance_change = paid - total carries it there on its
    # own, so there is nothing to do here but let the figure through — see
    # _write_bill. Only the full types take it. A partial that reached target
    # isn't partial, and a mixed one is a split of an exact figure.
    if kind == Bill.PaymentType.FULL_CASH:
        parts["cash"] = _decimal(raw.get("cash"), "Amount received", 2)
        parts["cash_account"] = _read_cash_account(raw.get("account"))
        if target <= ZERO:
            raise BillError("Nothing to collect — the credit covers this bill. Use Pay Later.")
        if parts["cash"] < target:
            raise BillError(
                f"Payment must be at least {target:.2f} — got {parts['cash']:.2f}. "
                f"Use Partial Cash to collect less."
            )

    elif kind == Bill.PaymentType.FULL_CHEQUE:
        cheques, cheque_total = _read_cheques(raw.get("cheques"))
        parts["cheques"] = cheques
        if target <= ZERO:
            raise BillError("Nothing to collect — the credit covers this bill. Use Pay Later.")
        if cheque_total < target:
            raise BillError(
                f"Cheques must total at least {target:.2f} — got {cheque_total:.2f}. "
                f"Use Partial Cheque to collect less."
            )

    elif kind == Bill.PaymentType.PARTIAL_CASH:
        parts["cash"] = _decimal(raw.get("cash"), "Cash amount", 2)
        parts["cash_account"] = _read_cash_account(raw.get("account"))
        if parts["cash"] <= ZERO:
            raise BillError("Cash amount must be above 0.")
        if target <= ZERO:
            raise BillError("Nothing to collect — the credit covers this bill. Use Pay Later.")
        if parts["cash"] >= target:
            raise BillError(
                f"Partial Cash must be less than {target:.2f}. Use Full Cash to pay the whole bill."
            )

    elif kind == Bill.PaymentType.PARTIAL_CHEQUE:
        cheques, cheque_total = _read_cheques(raw.get("cheques"))
        parts["cheques"] = cheques
        if target <= ZERO:
            raise BillError("Nothing to collect — the credit covers this bill. Use Pay Later.")
        if cheque_total >= target:
            raise BillError(
                f"Partial Cheque total must be less than {target:.2f}. Use Full Cheque to pay the whole bill."
            )

    elif kind == Bill.PaymentType.MIXED:
        parts["cash"] = _optional_decimal(raw.get("cash"), "Cash amount")
        parts["cash_account"] = _read_cash_account(raw.get("account"))
        cheques, cheque_total = _read_cheques(raw.get("cheques"))
        parts["cheques"] = cheques
        if parts["cash"] <= ZERO:
            raise BillError("Mixed payment needs a cash amount above 0.")
        if target <= ZERO:
            raise BillError("Nothing to collect — the credit covers this bill. Use Pay Later.")
        combined = parts["cash"] + cheque_total
        if combined != target:
            raise BillError(
                f"Cash + cheques must total {target:.2f} — got {combined:.2f}."
            )

    elif kind == Bill.PaymentType.PAY_LATER:
        pass  # No inputs; credit limit is judged separately.

    elif kind == Bill.PaymentType.PARTIAL:
        # Legacy shape: cash + one cheque, together settling the full amount.
        # Kept only because a bill saved before the split may be re-saved by
        # editing — the page doesn't offer this type any more.
        parts["cash"] = _decimal(raw.get("cash"), "Cash amount", 2)
        cheque = _read_cheque(raw.get("cheque"))
        parts["cheques"] = [cheque]
        if target <= ZERO:
            raise BillError("Nothing to collect — the credit covers this bill. Use Pay Later.")
        combined = parts["cash"] + cheque["amount"]
        if combined != target:
            raise BillError(f"Payment must total {target:.2f} — got {combined:.2f}.")

    cheque_total = sum((c["amount"] for c in parts["cheques"]), ZERO)
    parts["paid"] = ZERO if kind == Bill.PaymentType.PAY_LATER else (parts["cash"] + cheque_total)
    return parts


def _read_walkin_payment(raw, total):
    """The only payment a walk-in sale can take: cash, for the exact total.

    There is no account behind a walk-in bill, so nothing can be left
    outstanding and nothing can be settled by cheque or transfer against a
    balance that does not exist — full cash, in full, is the one shape that
    works. Enforced here regardless of what the page already restricts it to,
    since the page's restriction is a courtesy and not a control.
    """
    raw = raw or {}
    kind = str(raw.get("type") or "").strip()
    if kind != Bill.PaymentType.FULL_CASH:
        raise BillError("A walk-in sale can only be paid by full cash.")

    cash = _decimal(raw.get("cash"), "Amount received", 2)
    account = _read_cash_account(raw.get("account"))

    target = total.quantize(Decimal("0.01"))
    if cash != target:
        raise BillError(f"Payment must total {target:.2f} — got {cash:.2f}.")

    return {
        "type": kind,
        "cash": cash,
        "cash_account": account,
        "cheques": [],
        "paid": cash,
        "credit_override": False,
    }


def _check_credit_limit(customer, total, parts, user):
    """Only Pay Later can leave money outstanding; every other type is held to
    the full amount, so it lands the balance on zero.

    Measured against the bill's total rather than its subtotal: what the
    customer ends up owing is what the bill came to.
    """
    if parts["type"] != Bill.PaymentType.PAY_LATER:
        return

    after = customer.balance - total
    owed = -after if after < ZERO else ZERO
    if owed <= customer.credit_limit:
        return

    # The page hides the override from managers, which is a courtesy, not a
    # control: the flag arrives from the browser and is only worth what this
    # check makes it worth.
    if not _is_super_admin(user):
        raise BillError(
            f"This bill leaves {owed:.2f} outstanding, past the "
            f"{customer.credit_limit:.2f} credit limit. A super admin has to approve it."
        )
    if not parts.get("credit_override"):
        raise BillError(
            f"This bill leaves {owed:.2f} outstanding, past the "
            f"{customer.credit_limit:.2f} credit limit, and needs an override."
        )


def _record_payments(bill, customer, parts, when=None):
    """Payment rows, and the paper trail each one leaves behind.

    One Payment for the cash leg (if any). One Payment plus one Cheque per
    cheque row on the bill.

    If `bill` is None, this is a *detached* settlement — a payment against a
    customer's account with no specific bill behind it (typically an opening
    balance). In that case Payment.bill is left NULL and Payment.customer is
    set instead; CashDrawer and Cheque rows are still written, dated by
    `when` (defaulting to today) and labelled by the customer's name.

    `when` is the date the operator is booking the money on — the bill's
    date for a bill-attached call, or the settlement date for a detached
    one. Kept explicit so a corrective settlement on last month's date
    lands in last month's cash drawer, not today's.
    """
    now = timezone.now()
    if when is None:
        when = bill.bill_date if bill is not None else timezone.localdate()

    # A short label used on CashDrawer.reason so a drawer entry can be read
    # back without joining anything.
    if bill is not None:
        source_label = f"Bill #{bill.pk}"
    else:
        who = customer.name if customer is not None else "settlement"
        source_label = f"Settlement · {who}"

    if parts["cash"] > ZERO:
        payment = Payment.objects.create(
            bill=bill,
            customer=customer if bill is None else None,
            method=Payment.Method.CASH,
            amount=parts["cash"],
            account=parts["cash_account"],
            paid_at=now,
        )
        if parts["cash_account"]:
            CashTransfer.objects.create(
                payment=payment,
                to_account=parts["cash_account"],
                amount=parts["cash"],
                transferred_at=now,
            )
            # Both legs, deliberately. The cash reached the drawer and then
            # left it for the bank; writing only the transfer would take the
            # drawer down by money it never held.
            CashDrawer.objects.create(
                txn_date=when,
                txn_type=CashDrawer.TxnType.IN,
                amount=parts["cash"],
                reason=f"{source_label} cash",
                bill=bill,
            )
            CashDrawer.objects.create(
                txn_date=when,
                txn_type=CashDrawer.TxnType.TRANSFER,
                amount=parts["cash"],
                reason=f"{source_label} cash to {payment.get_account_display()}",
                bill=bill,
            )
        else:
            CashDrawer.objects.create(
                txn_date=when,
                txn_type=CashDrawer.TxnType.IN,
                amount=parts["cash"],
                reason=f"{source_label} cash",
                bill=bill,
            )

    for cheque in parts["cheques"]:
        payment = Payment.objects.create(
            bill=bill,
            customer=customer if bill is None else None,
            method=Payment.Method.CHEQUE,
            amount=cheque["amount"],
            paid_at=now,
        )
        Cheque.objects.create(
            payment=payment,
            bill=bill,
            customer=customer,
            cheque_no=cheque["cheque_no"],
            bank_name=cheque["bank_name"],
            branch=cheque["branch"],
            acc_no=cheque["acc_no"],
            amount=cheque["amount"],
            received_date=cheque["received_date"],
            maturity_date=cheque["maturity_date"],
        )


def _reverse_bill(bill):
    """Undo everything a bill did, leaving it an empty header.

    Shared by edit and delete: an edit is this followed by a fresh write, and a
    delete is this followed by dropping the header. Must run inside a
    transaction — half a reversal is worse than none.
    """
    # Auto-created Oversale production rows: reverse the stock they added and
    # delete them, so a bill that oversold leaves no phantom production
    # behind on its way out. Done before the normal stock restore so both
    # movements are undone in the same order they were applied.
    oversale = ProductionEntry.objects.filter(
        reason__startswith=OVERSALE_REASON_PREFIX + f" Bill #{bill.pk}"
    )
    for entry in oversale:
        Product.objects.filter(pk=entry.product_id).update(
            qty=F("qty") - entry.qty_produced
        )
    oversale.delete()

    # Stock next, off the rows about to be deleted.
    for item in bill.items.all():
        Product.objects.filter(pk=item.product_id).update(qty=F("qty") + item.qty)

    # new_balance = old_balance + balance_change on the way in, so taking the
    # same figure back out is the exact inverse.
    Customer.objects.filter(pk=bill.customer_id).update(
        balance=F("balance") - bill.balance_change
    )

    # CashDrawer.bill is SET_NULL, so these survive the bill as orphans that
    # still count toward the drawer balance. They have to go by hand.
    CashDrawer.objects.filter(bill=bill).delete()

    # Cheque and CashTransfer hang off Payment with CASCADE, so they go too.
    bill.payments.all().delete()
    bill.items.all().delete()


@transaction.atomic
def _save_bill(user, payload):
    """Write one bill and everything it touches, or nothing at all."""
    return _write_bill(Bill(), user, payload)


@transaction.atomic
def _update_bill(bill, user, payload, edit_date, edit_reason):
    """Rewrite a bill as if it had always said this, and record that it was.

    The reversal has to come first: the new lines are validated against stock
    and a balance that no longer carry this bill's own effects, so re-saving an
    unchanged bill is a no-op rather than a double charge.

    The audit note is written in the same transaction as the rewrite, so a bill
    that failed to save has no note claiming it did, and a bill that saved can
    never be missing the reason it changed.
    """
    _reverse_bill(bill)

    # Set before the write rather than saved after it: _write_bill saves the
    # header itself, so a second save() here would only be another round trip.
    bill.edit_date = edit_date
    bill.edit_reason = edit_reason
    bill = _write_bill(bill, user, payload)

    BillEditAudit.objects.create(
        bill=bill,
        edit_date=edit_date,
        reason=edit_reason,
        created_by=user,
    )
    return bill


def _write_bill(bill, user, payload):
    is_walk_in = bool(payload.get("is_walk_in"))
    walk_in_name = str(payload.get("walk_in_name") or "").strip()

    if is_walk_in:
        # No account behind a walk-in sale: Bill.customer is null exactly for
        # this, and walk_in_name is the only record of who bought the goods —
        # so it is required rather than decorative. Nothing here touches any
        # Customer row; a walk-in cannot move a balance that doesn't exist.
        if not walk_in_name:
            raise BillError("Enter the walk-in customer's name.")
        if len(walk_in_name) > 255:
            raise BillError("That walk-in name is too long.")
        customer = None
    else:
        walk_in_name = ""
        # Read fresh: on an edit the reversal above moved the balance with an
        # F() expression, which leaves any object already in memory stale.
        #
        # Suppliers *may* be billed too: a party we buy from may also buy from
        # us. The two flows meet on Customer.balance — a supplier bill moves
        # it up (we owe them), a sales bill moves it down (they owe us) — and
        # the ledger reads the net. See _billable_customers.
        customer = Customer.objects.filter(
            pk=payload.get("customer_id"), is_active=True
        ).first()
        if customer is None:
            raise BillError("That customer can't be billed.")

    raw_lines = payload.get("lines")
    if not isinstance(raw_lines, list) or not raw_lines:
        raise BillError("Add at least one product to the bill.")

    products = {
        product.pk: product
        for product in Product.objects.filter(
            pk__in=[raw.get("product_id") for raw in raw_lines if isinstance(raw, dict)],
            is_active=True,
        )
    }
    # A walk-in has no customer to have quoted it a price at all — every line
    # prices against the product's default, same as a regular customer with no
    # override of their own.
    quoted = (
        dict(
            CustomerPrice.objects.filter(customer=customer)
            .order_by()
            .values_list("product_id", "unit_price")
        )
        if customer is not None
        else {}
    )

    items = []
    subtotal = ZERO
    seen = set()
    for raw in raw_lines:
        if not isinstance(raw, dict):
            raise BillError(MALFORMED)

        product = products.get(raw.get("product_id"))
        if product is None:
            raise BillError("A product on this bill is no longer available.")
        if product.pk in seen:
            raise BillError(f"{product} is on the bill twice.")
        seen.add(product.pk)

        qty = _decimal(raw.get("qty"), "Quantity", 3)
        if qty <= ZERO:
            raise BillError(f"Quantity for {product} must be above 0.")
        unit_price = _decimal(raw.get("unit_price"), "Unit price", 2)

        # Recomputed here: a line total off the browser is a number the biller
        # could have typed.
        line_total = (qty * unit_price).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
        subtotal += line_total
        items.append(
            {
                "product": product,
                "qty": qty,
                "unit_price": unit_price,
                "line_total": line_total,
                "quoted": quoted.get(product.pk, product.default_price),
            }
        )

    # Delivery is charged on top of the goods; the discount comes off the lot.
    # Both are optional, so a payload without them prices exactly as before.
    delivery_charge = _optional_decimal(payload.get("delivery_charge"), "Delivery charge")
    discount_amount = _optional_decimal(payload.get("discount_amount"), "Discount")
    discount_reason = str(payload.get("discount_reason") or "").strip()[:255]

    if discount_amount > ZERO and not discount_reason:
        # Money off a bill is the one figure on it that nothing else explains.
        raise BillError("Give a reason for the discount.")
    if discount_amount == ZERO:
        discount_reason = ""

    total = subtotal + delivery_charge - discount_amount
    if total < ZERO:
        raise BillError(
            f"The discount is more than the bill — {subtotal + delivery_charge:.2f} "
            f"including delivery, discounted by {discount_amount:.2f}."
        )

    # Everything downstream prices against `total`, not the subtotal: the
    # payment collects what the bill actually comes to, and the credit limit
    # measures the debt it actually leaves.
    #
    # A walk-in takes a completely different path: there is no balance to
    # collect against or put debt on, so it is full cash for the exact total
    # rather than anything _read_payment or the credit limit would judge.
    if is_walk_in:
        parts = _read_walkin_payment(payload.get("payment"), total)
    else:
        parts = _read_payment(payload.get("payment"), total, customer)
        parts["credit_override"] = bool((payload.get("payment") or {}).get("credit_override"))
        _check_credit_limit(customer, total, parts, user)

    paid = parts["paid"]
    if paid >= total:
        status = Bill.Status.PAID
    elif paid > ZERO:
        status = Bill.Status.PARTIAL
    else:
        status = Bill.Status.UNPAID

    # How this bill moves the balance. The sale is debt (balance down), the
    # payment settles debt (balance up), so the two net to paid - total. Adding
    # this to the balance is the whole update, which keeps
    # new_balance = old_balance + balance_change true — what the field is for.
    #
    # A walk-in never moves one: there is no customer behind it, and it is
    # always paid in full anyway, so the net would be zero even if there were.
    balance_change = ZERO if is_walk_in else paid - total

    # Snapshot of how much of the customer's positive balance (credit we owed
    # them) this bill absorbed. `customer` was fetched fresh above after any
    # reversal, so its `balance` here is the pre-bill figure — the same one
    # the payment target was measured against — which is what makes this
    # correct even on an edit.
    if not is_walk_in and customer.balance > ZERO:
        credit_applied = min(customer.balance, total)
    else:
        credit_applied = ZERO

    # 1. header. An edit keeps the date it was billed on — the goods left the
    # yard that day whatever gets corrected afterwards.
    bill.customer = customer
    if bill.pk is None:
        bill.bill_date = _read_bill_date(payload.get("bill_date"))
    bill.subtotal = subtotal
    bill.delivery_charge = delivery_charge
    bill.discount_amount = discount_amount
    bill.discount_reason = discount_reason
    bill.total_amount = total
    bill.paid_amount = paid
    bill.balance_change = balance_change
    bill.credit_applied = credit_applied
    bill.payment_type = parts["type"]
    bill.status = status
    bill.is_walk_in = is_walk_in
    bill.walk_in_name = walk_in_name
    bill.notes = str(payload.get("notes") or "").strip()
    bill.save()

    # 2. lines, and the stock they take with them.
    #
    # Overselling is allowed: a bill may take more than the shelf holds. When
    # that happens we auto-create a matching ProductionEntry with reason
    # "Oversale — Bill #N" for the shortfall, then decrement stock normally.
    # The net effect is that Product.qty never goes negative and the stock
    # ledger has a matching production row explaining where the extra units
    # came from. _reverse_bill deletes these auto rows on edit/delete so the
    # phantom stock does not persist past its bill.
    for item in items:
        BillItem.objects.create(
            bill=bill,
            product=item["product"],
            qty=item["qty"],
            unit_price=item["unit_price"],
            line_total=item["line_total"],
        )

        # Fresh read: two tills billing at once may already have moved the
        # shelf since this bill was validated.
        current_qty = Product.objects.values_list("qty", flat=True).get(
            pk=item["product"].pk
        )
        if item["qty"] > current_qty:
            shortage = item["qty"] - current_qty
            ProductionEntry.objects.create(
                product=item["product"],
                production_date=bill.bill_date,
                qty_produced=shortage,
                reason=OVERSALE_REASON_PREFIX + f" Bill #{bill.pk}",
                stock_before=current_qty,
                stock_after=current_qty + shortage,
            )
            Product.objects.filter(pk=item["product"].pk).update(
                qty=F("qty") + shortage
            )

        Product.objects.filter(pk=item["product"].pk).update(
            qty=F("qty") - item["qty"]
        )

    # 3. money. Payment/Cheque/CashDrawer rows don't need a customer — a
    # walk-in still hits the till and the drawer exactly like any other cash
    # sale, it just isn't collected against anyone's account.
    _record_payments(bill, customer, parts)

    # 4. balance. F() so a balance moved by another till in the meantime is
    # adjusted rather than overwritten. Skipped entirely for a walk-in: there
    # is no customer row to move, and balance_change is 0 regardless.
    if not is_walk_in:
        Customer.objects.filter(pk=customer.pk).update(balance=F("balance") + balance_change)

    # 5. prices the biller changed become this customer's price. Compared
    # against what was quoted, not against the browser's price_changed flag.
    # A walk-in has no customer to remember a price for, so nothing to write.
    if not is_walk_in:
        for item in items:
            if item["unit_price"] != item["quoted"]:
                CustomerPrice.objects.update_or_create(
                    customer=customer,
                    product=item["product"],
                    defaults={"unit_price": item["unit_price"]},
                )

    return bill


@require_POST
@login_required
def bill_save(request):
    try:
        payload = json.loads(request.body or b"{}")
    except json.JSONDecodeError:
        return JsonResponse({"success": False, "error": MALFORMED}, status=400)
    if not isinstance(payload, dict):
        return JsonResponse({"success": False, "error": MALFORMED}, status=400)

    try:
        bill = _save_bill(request.user, payload)
    except BillError as exc:
        # _save_bill is atomic, so nothing it wrote survives this.
        return JsonResponse({"success": False, "error": str(exc)}, status=400)
    except Exception as exc:
        # Anything else — an IntegrityError from a stale DB schema, a bug in
        # this view — used to fall through to Django's 500 handler, which
        # answers with HTML. That leaves the biller with a generic "Could not
        # save the bill" because the browser can't parse the reply as JSON.
        # Report the exception's class and message so the biller can see what
        # actually broke; _save_bill is atomic, so nothing here survived it.
        import logging
        logging.exception("bill_save failed for user %s", request.user)
        return JsonResponse(
            {"success": False, "error": f"{type(exc).__name__}: {exc}"},
            status=500,
        )

    # If this bill was recalled from a held draft, retire the draft. Done
    # after _save_bill returns so a validation failure on the way in leaves
    # the draft alone — the biller can still recall it and try again.
    held_id = request.GET.get("held", "").strip()
    if held_id.isdigit():
        HeldBill.objects.filter(pk=int(held_id)).delete()

    who = bill.walk_in_name if bill.is_walk_in else bill.customer.name
    messages.success(request, f"Bill #{bill.pk} for {who} was saved.")
    return JsonResponse(
        {
            "success": True,
            "bill_id": bill.pk,
            "redirect": reverse("core:bill_detail", args=[bill.pk]),
        }
    )


@login_required
def bill_detail(request, pk):
    bill = get_object_or_404(_bills_with_counts(), pk=pk)

    # A walk-in has no customer behind it to have a balance at all — there is
    # nothing to recover here, and the template shows a plain "paid in full"
    # note instead of a balance breakdown.
    if bill.customer_id:
        balance_before = bill.customer.balance - bill.balance_change
        owed_now = -bill.customer.balance if bill.customer.balance < ZERO else ZERO
    else:
        balance_before = ZERO
        owed_now = ZERO

    import re
    items_qs = bill.items.select_related("product")
    items = []
    for item in items_qs:
        size_str = item.product.size or ""
        clean_name = item.product.name
        if size_str and clean_name.upper().startswith(size_str.upper()):
            clean_name = clean_name[len(size_str):].strip()
        clean_name = re.sub(r'\s*-\s*(SENOVKA|KRISHAN|SURESH)$', '', clean_name, flags=re.IGNORECASE).strip()
        item.clean_name = clean_name
        items.append(item)

    return render(
        request,
        "core/bill_detail.html",
        {
            "bill": bill,
            "reverses": json.dumps(_reversal_summary(bill)),
            "items": items,
            "payments": bill.payments.prefetch_related("cheques", "transfers"),
            # The bill records how it moved the balance, so the reading at the
            # time it was saved can be recovered without a full ledger replay.
            "balance_before": balance_before,
            # Templates can't take an absolute value, and a debt reads better
            # as a positive figure.
            "owed_now": owed_now,
        },
    )


@login_required
def bill_print(request, pk):
    """Formal A4 print-ready view of one bill.

    Renders a standalone HTML page with the current `BillingSettings` painted
    across the header and footer, the company logo top-left, and the items /
    totals / payments block below. The template auto-fires `window.print()` on
    load unless the URL carries `?noprint=1`, which is what the "Preview"
    button on the settings page uses to render without launching the print
    dialog.

    Not a PDF endpoint — the browser's print pipeline handles A4 sizing and
    lets the operator send it to whatever printer they have set up, which is
    the same story the bill_pdf endpoints tell for reports.
    """
    bill = get_object_or_404(
        Bill.objects.select_related("customer"), pk=pk
    )
    items = list(bill.items.select_related("product"))
    payments = list(bill.payments.prefetch_related("cheques", "transfers"))
    settings_row = BillingSettings.load()

    if bill.customer_id:
        balance_before = bill.customer.balance - bill.balance_change
    else:
        balance_before = ZERO

    return render(
        request,
        "core/bill_print.html",
        {
            "bill": bill,
            "items": items,
            "payments": payments,
            "billing": settings_row,
            "balance_before": balance_before,
            "auto_print": request.GET.get("noprint") != "1",
        },
    )


def _refresh_bill_status(bill):
    """Set `bill.status` from what has now been collected and settled.

    The same shape the bill-write path computes at save time, kept in one place
    so a follow-up payment or a settlement lands on the same status as if the
    bill had been written that way to begin with. A cancelled bill is left
    alone — payments against a cancelled bill would already have been refused
    upstream.
    """
    if bill.status == Bill.Status.CANCELLED:
        return
    covered = bill.paid_amount + bill.settled_amount
    if covered >= bill.total_amount:
        new_status = Bill.Status.PAID
    elif covered > ZERO:
        new_status = Bill.Status.PARTIAL
    else:
        new_status = Bill.Status.UNPAID
    if new_status != bill.status:
        Bill.objects.filter(pk=bill.pk).update(status=new_status)
        bill.status = new_status


@login_required
def bill_add_payment(request, pk):
    """Record one follow-up payment against a bill that still owes money.

    A Pay Later bill (or a Partial one) leaves an outstanding balance the
    customer still has to hand over. This view is that follow-up: pick cash or
    a single cheque, enter the amount, save. The plumbing is deliberately the
    same as the bill-write path — `_record_payments` writes the Payment / Cheque
    / CashDrawer / CashTransfer rows — so the drawer, the cheque list and the
    ledger see nothing they haven't seen before.

    Refused for:
      - walk-in bills: no customer to owe anything, and they are paid at the
        till by construction.
      - cancelled bills: nothing is owed on a cancelled sale.
      - bills already fully covered by payments and settlements.
    """
    bill = get_object_or_404(
        Bill.objects.select_related("customer"), pk=pk
    )

    if bill.is_walk_in or bill.customer_id is None:
        messages.error(
            request,
            "Walk-in bills are paid at the till and can't take a follow-up payment.",
        )
        return redirect("core:bill_detail", pk=pk)
    if bill.status == Bill.Status.CANCELLED:
        messages.error(request, "Cancelled bills can't take new payments.")
        return redirect("core:bill_detail", pk=pk)
    if bill.remaining_balance <= ZERO:
        messages.info(request, f"Bill #{bill.pk} is already settled.")
        return redirect("core:bill_detail", pk=pk)

    form = BillPaymentForm(request.POST or None, bill=bill)

    if request.method == "POST" and form.is_valid():
        data = form.cleaned_data
        amount = data["amount"]

        # Shape the payload _record_payments expects. Only one leg is ever
        # populated on this form — cash OR one cheque — so the other side is
        # an empty default.
        if data["method"] == Payment.Method.CASH:
            parts = {
                "cash": amount,
                "cash_account": data.get("cash_account") or "",
                "cheques": [],
            }
        else:
            parts = {
                "cash": ZERO,
                "cash_account": "",
                "cheques": [
                    {
                        "cheque_no": data["cheque_no"],
                        "bank_name": data["bank_name"],
                        "branch": data.get("branch") or "",
                        "acc_no": data.get("acc_no") or "",
                        "amount": amount,
                        "received_date": data["received_date"],
                        "maturity_date": data["maturity_date"],
                    }
                ],
            }

        with transaction.atomic():
            _record_payments(bill, bill.customer, parts)

            # Two counters move by the same figure: the bill's paid_amount so
            # remaining_balance drops, and balance_change so the pre-edit
            # figure the bill stores stays a true summary of everything it has
            # ever moved. Customer.balance rises by `amount` — cash coming in
            # settles debt, which reads as a positive move in the balance
            # column.
            Bill.objects.filter(pk=bill.pk).update(
                paid_amount=F("paid_amount") + amount,
                balance_change=F("balance_change") + amount,
            )
            Customer.objects.filter(pk=bill.customer_id).update(
                balance=F("balance") + amount
            )

            # Read the fresh figures back and let the status helper decide.
            bill.refresh_from_db()
            _refresh_bill_status(bill)

        messages.success(
            request,
            f"Recorded {amount:,.2f} against Bill #{bill.pk}. "
            f"Remaining: {bill.remaining_balance:,.2f}.",
        )
        return redirect("core:bill_detail", pk=bill.pk)

    if request.method == "POST":
        messages.error(request, f"Payment not saved: {form.first_error()}")

    return render(
        request,
        "core/bill_add_payment.html",
        {
            "bill": bill,
            "form": form,
            "account_choices": Payment.Account.choices,
        },
    )


def _outstanding_bills_for(customer):
    """The customer's unpaid or partially paid bills, oldest first.

    Ordering is the settlement contract: cash and cheques both flow into
    these bills in this order. Cancelled bills are excluded — nothing is
    owed on a cancelled sale. Walk-in flag is irrelevant here because a
    customer-scoped settlement only makes sense for a real account anyway.
    """
    bills = (
        Bill.objects.filter(customer=customer)
        .exclude(status=Bill.Status.CANCELLED)
        .order_by("bill_date", "pk")
    )
    return [b for b in bills if b.remaining_balance > ZERO]


def _allocate_settlement(customer, cash, cash_account, cheques, user, when=None):
    """Fan a lump settlement out across the customer's outstanding bills,
    then spill anything left over into detached payments against the
    customer's account.

    Order of operations:
      1. Cash pass — split cash oldest→newest across bills, up to each
         bill's remaining balance.
      2. Cheque pass — each cheque is one physical instrument, attached
         whole to the oldest bill still owing.
      3. Spillover — any remaining cash and any remaining cheques land on
         detached Payment rows (Payment.bill = NULL, Payment.customer set).
         This is what covers a payment against an opening balance, or
         against a customer with no bills at all.
      4. Customer.balance rises by the whole lump. If bills didn't need
         all of it, the extra sits as credit for the next bill; if there
         weren't any bills, the balance moves directly toward zero.

    Must run inside a transaction — a half-allocated settlement would leave
    orphan Payment rows against bills whose paid_amount was never updated.
    """
    outstanding = _outstanding_bills_for(customer)
    # Snapshot so the local remaining tracks alongside the DB update — the
    # F() update on Bill.paid_amount leaves the in-memory object stale.
    remaining = {b.pk: b.remaining_balance for b in outstanding}
    allocations = []

    # --- cash pass across bills ---
    cash_left = cash
    for bill in outstanding:
        if cash_left <= ZERO:
            break
        take = min(cash_left, remaining[bill.pk])
        if take <= ZERO:
            continue
        parts = {"cash": take, "cash_account": cash_account, "cheques": []}
        _record_payments(bill, customer, parts)
        Bill.objects.filter(pk=bill.pk).update(
            paid_amount=F("paid_amount") + take,
            balance_change=F("balance_change") + take,
        )
        remaining[bill.pk] -= take
        cash_left -= take
        allocations.append(("cash", bill.pk, take))

    # --- cheque pass across bills ---
    # Each cheque is a whole physical instrument. Attach it to the oldest
    # bill still owing; if all bills are settled, drop it into the
    # spillover below rather than over-paying an already-settled bill.
    cheques_left = []
    for cheque in cheques:
        target = None
        for bill in outstanding:
            if remaining[bill.pk] > ZERO:
                target = bill
                break
        if target is None:
            cheques_left.append(cheque)
            continue

        parts = {"cash": ZERO, "cash_account": "", "cheques": [cheque]}
        _record_payments(target, customer, parts)
        Bill.objects.filter(pk=target.pk).update(
            paid_amount=F("paid_amount") + cheque["amount"],
            balance_change=F("balance_change") + cheque["amount"],
        )
        remaining[target.pk] -= cheque["amount"]
        allocations.append(("cheque", target.pk, cheque["amount"]))

    # --- spillover into detached payments against the customer ---
    # Either bills were absent, or the payment overshot what they owed.
    # Anything landing here reduces the customer's opening/prior balance.
    if cash_left > ZERO or cheques_left:
        parts = {
            "cash": cash_left,
            "cash_account": cash_account,
            "cheques": cheques_left,
        }
        _record_payments(None, customer, parts, when=when)
        allocations.append(("detached-cash", None, cash_left))
        for c in cheques_left:
            allocations.append(("detached-cheque", None, c["amount"]))

    cheque_total = sum((c["amount"] for c in cheques), ZERO)
    total_paid = cash + cheque_total

    # Customer.balance rises by the whole lump — this is the one place
    # that reconciles the amount collected with what the account was
    # standing at. Bills the money hit already had their own
    # balance_change bumped inside the loop, so the sum still nets right.
    Customer.objects.filter(pk=customer.pk).update(
        balance=F("balance") + total_paid
    )

    # Status pass — a bill that has been settled or partially covered by
    # this call should read that way on the next page load.
    for bill in outstanding:
        bill.refresh_from_db()
        _refresh_bill_status(bill)

    return {"allocations": allocations, "total_paid": total_paid}


@login_required
def customer_settle(request, pk):
    """Settle a customer's outstanding bills with one lump payment.

    Reused from the bill-payment view but customer-scoped: the operator
    enters cash and/or cheques once, and _allocate_settlement fans the
    amount out across the unpaid bills FIFO. Every allocation goes through
    _record_payments, so the cash drawer, the cheque list and the ledger
    all see the same rows they would from a per-bill payment.
    """
    customer = get_object_or_404(Customer, pk=pk)

    outstanding = _outstanding_bills_for(customer)
    total_owed = sum((b.remaining_balance for b in outstanding), ZERO)

    form = CustomerSettlementForm(request.POST or None, customer=customer)

    if request.method == "POST" and form.is_valid():
        data = form.cleaned_data
        cash = data["_cash_amount"]
        account = data.get("cash_account") or ""
        cheques = form.parsed_cheques
        total_paid = data["_total_paid"]

        # Settlement is allowed whenever the operator wants to record a
        # payment — no bills is fine (the payment sits as detached against
        # the customer, chipping away at an opening balance or piling up
        # as credit if the account is already square).
        with transaction.atomic():
            _allocate_settlement(customer, cash, account, cheques, request.user)

        if outstanding:
            excess = total_paid - total_owed
            if excess > ZERO:
                messages.success(
                    request,
                    f"Settled {total_owed:,.2f} across {len(outstanding)} bill"
                    f"{'' if len(outstanding) == 1 else 's'}. "
                    f"{excess:,.2f} kept as credit on {customer.name}'s account.",
                )
            else:
                messages.success(
                    request,
                    f"Recorded {total_paid:,.2f} against {customer.name}'s bills.",
                )
        else:
            # No bills at all — the whole lump sat down on the account.
            customer.refresh_from_db()
            if customer.balance > ZERO:
                messages.success(
                    request,
                    f"Recorded {total_paid:,.2f} for {customer.name}. "
                    f"Account now stands at +{customer.balance:,.2f} credit.",
                )
            elif customer.balance == ZERO:
                messages.success(
                    request,
                    f"Recorded {total_paid:,.2f} for {customer.name}. "
                    f"Account cleared.",
                )
            else:
                messages.success(
                    request,
                    f"Recorded {total_paid:,.2f} for {customer.name}. "
                    f"Still owes {-customer.balance:,.2f}.",
                )
        return redirect("core:customer_ledger", pk=customer.pk)

    if request.method == "POST":
        messages.error(request, f"Settlement not saved: {form.first_error()}")

    return render(
        request,
        "core/customer_settle.html",
        {
            "customer": customer,
            "form": form,
            "outstanding_bills": outstanding,
            "total_owed": total_owed,
            "account_choices": Payment.Account.choices,
        },
    )


def _bills_with_counts():
    """Bills carrying what the delete modal has to describe."""
    return (
        Bill.objects.select_related("customer")
        .annotate(
            # distinct=True: without it these joins multiply each other's rows.
            item_count=Count("items", distinct=True),
            payment_count=Count("payments", distinct=True),
            drawer_count=Count("cash_drawer_entries", distinct=True),
        )
        # Spelled out even though it only repeats Bill.Meta.ordering: annotate()
        # groups, and Django drops Meta.ordering from a grouped query, leaving
        # no ORDER BY at all. Unordered is merely untidy when the whole list is
        # on one screen, but it is wrong once it is paginated — LIMIT/OFFSET
        # over an unordered read may hand the same bill to two pages and never
        # show another.
        .order_by("-bill_date", "-id")
    )


def _reversal_summary(bill):
    """Plain sentences for what undoing this bill puts back.

    Built here rather than in the modal so the warning and the reversal are
    read off the same numbers.
    """
    lines = []

    if bill.item_count:
        lines.append(
            f"{bill.item_count} line{'' if bill.item_count == 1 else 's'} "
            f"of stock returned"
        )

    if bill.balance_change:
        restored = bill.customer.balance - bill.balance_change
        lines.append(f"{bill.customer.name}'s balance returns to {restored:.2f}")

    if bill.payment_count:
        lines.append(
            f"{bill.payment_count} payment record{'' if bill.payment_count == 1 else 's'} "
            f"removed, with any cheque or transfer on them"
        )

    if bill.drawer_count:
        lines.append(
            f"{bill.drawer_count} cash drawer "
            f"entr{'y' if bill.drawer_count == 1 else 'ies'} removed"
        )

    return lines or ["Nothing — this bill moved no stock or money."]


def _bill_initial(bill):
    """The bill as the create page's own payload shape, for rehydrating it."""
    return {
        "customer_id": bill.customer_id,
        "bill_date": bill.bill_date.isoformat(),
        "is_walk_in": bill.is_walk_in,
        "walk_in_name": bill.walk_in_name,
        "delivery_charge": f"{bill.delivery_charge:.2f}",
        "discount_amount": f"{bill.discount_amount:.2f}",
        "discount_reason": bill.discount_reason,
        "lines": [
            {
                "product_id": item.product_id,
                "qty": f"{item.qty:.3f}".rstrip("0").rstrip("."),
                "unit_price": f"{item.unit_price:.2f}",
            }
            for item in bill.items.all()
        ],
        "payment": _payment_initial(bill),
    }


def _payment_initial(bill):
    """Unpick the payment rows back into the form's fields.

    Cheques are returned as an array — a bill can carry any number, and the
    page rebuilds one cheque row per entry. The legacy transfer amount is
    folded into the cash row's account so a re-edit of a Mixed bill from the
    old shape still hydrates sensibly.
    """
    payment = {"type": bill.payment_type, "cheques": []}

    for row in bill.payments.prefetch_related("cheques"):
        if row.method == Payment.Method.CASH:
            payment["cash"] = f"{row.amount:.2f}"
            payment["account"] = row.account
        elif row.method == Payment.Method.TRANSFER:
            # Legacy MIXED: the old shape had a separate transfer leg. Treat
            # its account as the cash account so the new Mixed panel finds
            # somewhere to put the destination.
            payment["cash"] = f"{row.amount:.2f}"
            payment["account"] = row.account
        elif row.method == Payment.Method.CHEQUE:
            cheque = row.cheques.first()
            if cheque:
                payment["cheques"].append(
                    {
                        "cheque_no": cheque.cheque_no,
                        "bank_name": cheque.bank_name,
                        "branch": cheque.branch,
                        "acc_no": cheque.acc_no,
                        "amount": f"{cheque.amount:.2f}",
                        "received_date": cheque.received_date.isoformat(),
                        "maturity_date": cheque.maturity_date.isoformat(),
                    }
                )
    return payment


def _edit_gate_key(pk):
    return f"bill_edit_gate:{pk}"


def _read_edit_gate(request, pk):
    """The date and reason this edit was gated on, or None if it wasn't.

    Kept in the session rather than posted with the bill: a hidden field is
    editable by whoever is on the page, and the whole point of the gate is that
    the reason recorded is the one that was confirmed. Keyed by bill, so two
    tabs editing two bills don't wear each other's reason.
    """
    gate = request.session.get(_edit_gate_key(pk))
    if not isinstance(gate, dict):
        return None

    edit_date = _parse_date(gate.get("edit_date"))
    reason = str(gate.get("reason") or "").strip()[:500]
    if edit_date is None or not reason:
        # Half a gate is no gate — an old or hand-made session value gets sent
        # back through the form rather than saved as a blank reason.
        return None
    return {"edit_date": edit_date, "reason": reason}


@login_required
def bill_edit(request, pk):
    bill = get_object_or_404(Bill.objects.select_related("customer"), pk=pk)
    gate = _read_edit_gate(request, pk)

    # Step 1. The reason gate stands in front of the form: it is a normal HTML
    # POST, where the save below is the form's JSON one, which is what tells
    # the two apart on the one URL.
    if request.method == "POST" and request.content_type != "application/json":
        form = BillEditReasonForm(request.POST)
        if form.is_valid():
            request.session[_edit_gate_key(pk)] = {
                "edit_date": form.cleaned_data["edit_date"].isoformat(),
                "reason": form.cleaned_data["reason"],
            }
            return redirect("core:bill_edit", pk=pk)
        return render(
            request, "core/bill_edit_reason.html", {"bill": bill, "form": form}
        )

    if request.method == "POST":
        # Step 2. The save. The gate is enforced here and not only on the way
        # in: the page posts JSON to this URL, so a save that skipped the gate
        # would otherwise be a bill edited for no recorded reason.
        if gate is None:
            return JsonResponse(
                {
                    "success": False,
                    "error": "This edit needs a date and reason. Reload the page and confirm them.",
                },
                status=400,
            )

        try:
            payload = json.loads(request.body or b"{}")
        except json.JSONDecodeError:
            return JsonResponse({"success": False, "error": MALFORMED}, status=400)
        if not isinstance(payload, dict):
            return JsonResponse({"success": False, "error": MALFORMED}, status=400)

        try:
            bill = _update_bill(
                bill,
                request.user,
                payload,
                gate["edit_date"],
                gate["reason"],
            )
        except BillError as exc:
            # _update_bill is atomic, so the reversal it started is undone too.
            # The gate stays put: the biller is being sent back to the same
            # form to fix the figure, not to re-justify the same edit.
            return JsonResponse({"success": False, "error": str(exc)}, status=400)

        # Spent. The next edit of this bill is a new one and asks again.
        request.session.pop(_edit_gate_key(pk), None)

        messages.success(request, f"Bill #{bill.pk} was updated.")
        return JsonResponse(
            {
                "success": True,
                "bill_id": bill.pk,
                "redirect": reverse("core:bill_detail", args=[bill.pk]),
            }
        )

    if gate is None or "change" in request.GET:
        # Nothing confirmed yet, or the biller asked to revisit what they said.
        return render(
            request,
            "core/bill_edit_reason.html",
            {
                "bill": bill,
                "form": BillEditReasonForm(
                    initial={
                        "edit_date": (
                            gate["edit_date"] if gate else timezone.localdate()
                        ),
                        "reason": gate["reason"] if gate else "",
                    }
                ),
            },
        )

    # The page prices this bill as though it had never been saved, so the
    # customer it belongs to is offered the balance it would have without it.
    # Every other customer's balance is already free of this bill.
    customers = _billable_customers()
    # A walk-in bill has no customer at all — nothing to insert or price, the
    # walk-in toggle and name are hydrated straight from the bill instead.
    if not bill.is_walk_in and bill.customer not in customers:
        # Retired or turned supplier: whichever this bill was made out to, it
        # still has to be editable.
        customers.insert(0, bill.customer)
    for customer in customers:
        customer.balance_for_bill = (
            customer.balance - bill.balance_change
            if customer.pk == bill.customer_id
            else customer.balance
        )

    context = _bill_form_context(request, customers)
    context.update(
        {
            "bill": bill,
            "save_url": reverse("core:bill_edit", args=[bill.pk]),
            "initial": _bill_initial(bill),
            "is_edit": True,
            "edit_date": gate["edit_date"],
            "edit_reason": gate["reason"],
        }
    )
    return render(request, "core/bill_edit.html", context)


@require_POST
@super_admin_required
def bill_delete(request, pk):
    bill = get_object_or_404(Bill.objects.select_related("customer"), pk=pk)
    label = f"Bill #{bill.pk}"
    customer = bill.walk_in_name if bill.is_walk_in else bill.customer.name

    with transaction.atomic():
        _reverse_bill(bill)
        bill.delete()

    messages.success(request, f"{label} for {customer} was deleted and reversed.")
    return redirect("core:bill_list")


# ---------------------------------------------------------------- held bills
# A held bill is dormant: nothing is written to Bill, no stock moves, no
# balance changes. It stores the raw form payload verbatim, so recalling it is
# the same as opening the create page pre-filled with everything the biller
# already typed. Saving from the recall page goes through the normal
# bill_save path and drops the held record.


def _held_bill_label(payload, customer):
    """A short who-and-what for the held bills list."""
    if bool(payload.get("is_walk_in")):
        name = str(payload.get("walk_in_name") or "").strip()
        return (name or "Walk-in") + " (walk-in)"
    if customer is not None:
        return customer.name
    return "Unknown customer"


def _held_bill_snapshot(payload):
    """Item count and subtotal, cached for the list without a JSON parse each
    time. Ignores anything malformed rather than refusing to hold the draft —
    a held bill is a scratchpad, not a submitted one.
    """
    lines = payload.get("lines") if isinstance(payload, dict) else None
    if not isinstance(lines, list):
        return 0, ZERO

    count = 0
    subtotal = ZERO
    for raw in lines:
        if not isinstance(raw, dict):
            continue
        try:
            qty = Decimal(str(raw.get("qty") or "0"))
            price = Decimal(str(raw.get("unit_price") or "0"))
        except InvalidOperation:
            continue
        if qty <= ZERO or price < ZERO:
            continue
        count += 1
        subtotal += (qty * price).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
    return count, subtotal


@require_POST
@login_required
def held_bill_save(request):
    """Park the current form to be recalled and finished later.

    Deliberately forgiving: a held bill is a scratchpad, so anything short of
    unreadable JSON is stored as-is. Full validation waits for the real save
    step, when the biller has actually decided this is the bill they want.
    """
    try:
        payload = json.loads(request.body or b"{}")
    except json.JSONDecodeError:
        return JsonResponse({"success": False, "error": MALFORMED}, status=400)
    if not isinstance(payload, dict):
        return JsonResponse({"success": False, "error": MALFORMED}, status=400)

    lines = payload.get("lines")
    if not isinstance(lines, list) or not lines:
        return JsonResponse(
            {"success": False, "error": "Add at least one product before holding the bill."},
            status=400,
        )

    customer = None
    raw_id = payload.get("customer_id")
    if raw_id not in (None, "", 0, "0"):
        try:
            customer = Customer.objects.filter(pk=int(raw_id)).first()
        except (TypeError, ValueError):
            customer = None

    walk_in_name = str(payload.get("walk_in_name") or "").strip()[:255]
    label = _held_bill_label(payload, customer)
    item_count, subtotal = _held_bill_snapshot(payload)

    held = HeldBill.objects.create(
        customer=customer,
        walk_in_name=walk_in_name if bool(payload.get("is_walk_in")) else "",
        payload=payload,
        label=label,
        item_count=item_count,
        subtotal=subtotal,
        created_by=request.user,
    )
    messages.success(request, f"Held bill for {label} saved. Recall it from Held Bills.")
    return JsonResponse(
        {
            "success": True,
            "held_id": held.pk,
            "redirect": reverse("core:held_bill_list"),
        }
    )


@login_required
def held_bill_list(request):
    held = HeldBill.objects.select_related("customer", "created_by")
    return render(request, "core/held_bills.html", {"held_bills": held})


@login_required
def held_bill_recall(request, pk):
    """Open bill_create with a held bill pre-loaded, as if the biller had
    just finished picking their products. The held record stays put until
    the recalled bill is actually saved, so a mis-click on the recall link
    doesn't lose the draft.
    """
    held = get_object_or_404(HeldBill.objects.select_related("customer"), pk=pk)

    customers = _billable_customers()
    if held.customer is not None and held.customer not in customers:
        # Retired, turned supplier, or the walk-in holding account — the draft
        # still has to be recallable to whichever account it was made against.
        customers.insert(0, held.customer)
    for customer in customers:
        customer.balance_for_bill = customer.balance

    context = _bill_form_context(request, customers)
    # bill_save deletes this held record on success — see bill_save below.
    context.update(
        {
            "save_url": reverse("core:bill_save") + f"?held={held.pk}",
            "initial": held.payload,
            "is_edit": False,
            "held_bill": held,
        }
    )
    return render(request, "core/bill_create.html", context)


@require_POST
@login_required
def held_bill_delete(request, pk):
    held = get_object_or_404(HeldBill, pk=pk)
    label = held.label or f"#{held.pk}"
    held.delete()
    messages.success(request, f"Held bill for {label} was dropped.")
    return redirect("core:held_bill_list")


def _filtered_bills(request):
    from_date = _parse_date(request.GET.get("from_date"))
    to_date = _parse_date(request.GET.get("to_date"))

    customer_id = request.GET.get("customer", "").strip()
    selected_customer = int(customer_id) if customer_id.isdigit() else None

    payment_type = request.GET.get("payment_type", "").strip()
    if payment_type not in {value for value, _ in Bill.PaymentType.choices}:
        payment_type = ""
    status = request.GET.get("status", "").strip()
    if status not in {value for value, _ in Bill.Status.choices}:
        status = ""

    bills = _bills_with_counts().annotate(
        # paid_amount can run past total_amount when a payment also clears old
        # debt, and a bill can't owe less than nothing.
        outstanding=Greatest(
            F("total_amount") - F("paid_amount"), Value(ZERO), output_field=MONEY
        )
    )

    if from_date:
        bills = bills.filter(bill_date__gte=from_date)
    if to_date:
        bills = bills.filter(bill_date__lte=to_date)
    if selected_customer:
        bills = bills.filter(customer_id=selected_customer)
    if payment_type:
        bills = bills.filter(payment_type=payment_type)
    if status:
        bills = bills.filter(status=status)

    return bills, from_date, to_date, selected_customer, payment_type, status


@login_required
def bill_list(request):
    bills, from_date, to_date, selected_customer, payment_type, status = _filtered_bills(request)

    # Paginate before the per-row work below: _reversal_summary queries per
    # bill, so priced over the whole filtered set it would cost a page's worth
    # of queries for every bill the operator cannot see.
    page_obj = _paginate(request, bills)
    for bill in page_obj:
        bill.reverses = json.dumps(_reversal_summary(bill))

    return render(
        request,
        "core/bill_list.html",
        {
            "page_obj": page_obj,
            # The page's rows. The template iterates this, so it never has to
            # know whether it was handed a page or a plain list.
            "bills": page_obj.object_list,
            "customers": Customer.objects.filter(is_walk_in_account=False),
            "from_date": from_date,
            "to_date": to_date,
            "selected_customer": selected_customer,
            "payment_type": payment_type,
            "status": status,
            "payment_types": Bill.PaymentType.choices,
            "statuses": Bill.Status.choices,
            "is_filtered": bool(
                from_date or to_date or selected_customer or payment_type or status
            ),
        },
    )


@login_required
def bill_list_excel(request):
    from openpyxl import Workbook
    from openpyxl.styles import Font
    bills, _, _, _, _, _ = _filtered_bills(request)
    # Re-order the queryset for excel output if needed, though they should be in default order
    # Let's ensure it's ordered properly
    bills = bills.order_by("-bill_date", "-pk")

    wb = Workbook()
    ws = wb.active
    ws.title = "Bills"

    headers = [
        "Bill No",
        "Date",
        "Customer Name",
        "Status",
        "Payment Type",
        "Total Amount",
        "Paid Amount",
        "Outstanding",
    ]
    
    header_font = Font(bold=True)
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font

    total_amount_sum = Decimal("0.00")
    paid_amount_sum = Decimal("0.00")
    outstanding_sum = Decimal("0.00")

    row = 2
    for bill in bills:
        customer_name = bill.walk_in_name if bill.is_walk_in else (bill.customer.name if bill.customer else "")
        total_amount = bill.total_amount
        paid_amount = bill.paid_amount
        outstanding = bill.outstanding

        total_amount_sum += total_amount
        paid_amount_sum += paid_amount
        outstanding_sum += outstanding

        ws.cell(row=row, column=1, value=f"#{bill.pk:04d}")
        ws.cell(row=row, column=2, value=bill.bill_date.strftime("%Y-%m-%d") if bill.bill_date else "")
        ws.cell(row=row, column=3, value=customer_name)
        ws.cell(row=row, column=4, value=bill.get_status_display())
        ws.cell(row=row, column=5, value=bill.get_payment_type_display())
        ws.cell(row=row, column=6, value=float(total_amount))
        ws.cell(row=row, column=7, value=float(paid_amount))
        ws.cell(row=row, column=8, value=float(outstanding))
        row += 1

    # Add totals row
    ws.cell(row=row, column=5, value="Totals:")
    ws.cell(row=row, column=5).font = header_font
    
    total_cell = ws.cell(row=row, column=6, value=float(total_amount_sum))
    total_cell.font = header_font
    
    paid_cell = ws.cell(row=row, column=7, value=float(paid_amount_sum))
    paid_cell.font = header_font
    
    out_cell = ws.cell(row=row, column=8, value=float(outstanding_sum))
    out_cell.font = header_font

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="bills.xlsx"'
    wb.save(response)
    return response


# ------------------------------------------------------------------ cheques
# A cheque is money we are counting on but do not have. Pending and deposited
# both mean we still expect it, so the customer keeps the credit. Held and
# bounced mean we don't, so the debt comes back.

#: Statuses where the cheque's amount is still credited to the customer.
CREDITED_CHEQUE_STATUSES = {Cheque.Status.PENDING, Cheque.Status.DEPOSITED}


def _cheque_credit(status, amount):
    """What a cheque in this state contributes to its customer's balance."""
    return amount if status in CREDITED_CHEQUE_STATUSES else ZERO


def _move_balance_for_cheque(cheque, was_status, was_amount):
    """Move the customer's balance by the difference the change makes.

    One rule covers every transition, including an amount correction:

        pending  -> deposited   nothing moves, we still expect the money
        pending  -> held        credit comes off, the debt is back
        deposited-> bounced     same, even though it had cleared
        bounced  -> pending     re-presented, credit goes back on
        amount 100 -> 150       while credited, 50 more is owed to us

    Returns the signed move, for the message.
    """
    delta = _cheque_credit(cheque.status, cheque.amount) - _cheque_credit(
        was_status, was_amount
    )
    if delta:
        # F() so a balance moved elsewhere in the meantime is adjusted rather
        # than overwritten.
        Customer.objects.filter(pk=cheque.customer_id).update(
            balance=F("balance") + delta
        )
    return delta


def _cheque_balance_note(cheque, delta):
    """Say what the balance did, since the operator can't see it happen."""
    if not delta:
        return ""
    cheque.customer.refresh_from_db()
    if delta < 0:
        return (
            f" {cheque.customer.name} owes {abs(delta):.2f} again — "
            f"balance is now {cheque.customer.balance:.2f}."
        )
    return (
        f" {cheque.customer.name} is credited {delta:.2f} — "
        f"balance is now {cheque.customer.balance:.2f}."
    )


def _set_cheque_status(request, pk, status, bounce_new_date=None):
    cheque = get_object_or_404(Cheque.objects.select_related("customer"), pk=pk)
    was_status, was_amount = cheque.status, cheque.amount

    if cheque.status == status:
        messages.info(
            request,
            f"Cheque {cheque.cheque_no} is already {cheque.get_status_display().lower()}.",
        )
        return redirect("core:cheque_list")

    fields = ["status"]
    with transaction.atomic():
        cheque.status = status
        if bounce_new_date is not None:
            cheque.bounce_new_date = bounce_new_date
            fields.append("bounce_new_date")
        cheque.save(update_fields=fields)
        delta = _move_balance_for_cheque(cheque, was_status, was_amount)

    messages.success(
        request,
        f"Cheque {cheque.cheque_no} marked {cheque.get_status_display().lower()}."
        + _cheque_balance_note(cheque, delta),
    )
    return redirect("core:cheque_list")


@require_POST
@login_required
def cheque_deposit(request, pk):
    # Deliberately no balance change: the credit went on when the cheque was
    # taken, and clearing the bank only confirms it.
    return _set_cheque_status(request, pk, Cheque.Status.DEPOSITED)


@require_POST
@login_required
def cheque_hold(request, pk):
    return _set_cheque_status(request, pk, Cheque.Status.HELD)


@require_POST
@login_required
def cheque_bounce(request, pk):
    new_date = _parse_date(request.POST.get("bounce_new_date"))
    if new_date is None:
        messages.error(
            request, "Enter the date the cheque is expected to be re-presented."
        )
        return redirect("core:cheque_list")
    return _set_cheque_status(request, pk, Cheque.Status.BOUNCED, new_date)


@require_POST
@super_admin_required
def cheque_delete(request, pk):
    """Remove a cheque that should never have been recorded.

    Not a bounce — that is a real event with its own status. This is for a
    cheque entered in error, so it takes the payment with it.
    """
    cheque = get_object_or_404(Cheque.objects.select_related("customer"), pk=pk)

    if cheque.status == Cheque.Status.DEPOSITED:
        messages.error(
            request,
            f"Cheque {cheque.cheque_no} has been deposited, so it can't be deleted. "
            f"The money is in the bank — mark it bounced if it came back.",
        )
        return redirect("core:cheque_list")

    number = cheque.cheque_no
    customer = cheque.customer

    with transaction.atomic():
        # Only a credited cheque has anything to take back; a held or bounced
        # one was already reversed when it got that status.
        delta = -_cheque_credit(cheque.status, cheque.amount)
        if delta:
            Customer.objects.filter(pk=customer.pk).update(balance=F("balance") + delta)

        # The cheque hangs off the payment by CASCADE, so removing the payment
        # removes both — the payment only ever existed to carry this cheque.
        payment = cheque.payment
        payment.delete()

    customer.refresh_from_db()
    note = (
        f" {customer.name} owes {abs(delta):.2f} again — "
        f"balance is now {customer.balance:.2f}."
        if delta
        else ""
    )
    messages.success(request, f"Cheque {number} was deleted." + note)
    return redirect("core:cheque_list")


@login_required
def cheque_edit(request, pk):
    cheque = get_object_or_404(Cheque.objects.select_related("customer"), pk=pk)
    was_status, was_amount = cheque.status, cheque.amount

    form = ChequeForm(request.POST or None, instance=cheque)
    if request.method == "POST" and form.is_valid():
        with transaction.atomic():
            cheque = form.save()
            delta = _move_balance_for_cheque(cheque, was_status, was_amount)

        messages.success(
            request,
            f"Cheque {cheque.cheque_no} was updated." + _cheque_balance_note(cheque, delta),
        )
        return redirect("core:cheque_list")

    return render(
        request,
        "core/cheque_edit.html",
        {"form": form, "cheque": cheque, "credited": was_status in CREDITED_CHEQUE_STATUSES},
    )


@login_required
@login_required
def cheque_list_excel(request):
    """Download the cheque list as an .xlsx, honouring the same status /
    customer / maturity-range filters as the page."""
    from io import BytesIO

    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    status = request.GET.get("status", "").strip()
    if status not in {value for value, _ in Cheque.Status.choices}:
        status = ""
    customer_id = request.GET.get("customer", "").strip()
    selected_customer = int(customer_id) if customer_id.isdigit() else None
    from_date = _parse_date(request.GET.get("from_date"))
    to_date = _parse_date(request.GET.get("to_date"))

    cheques = Cheque.objects.select_related("customer")
    if status:
        cheques = cheques.filter(status=status)
    if selected_customer:
        cheques = cheques.filter(customer_id=selected_customer)
    if from_date:
        cheques = cheques.filter(maturity_date__gte=from_date)
    if to_date:
        cheques = cheques.filter(maturity_date__lte=to_date)

    wb = Workbook()
    ws = wb.active
    ws.title = "Cheques"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1F2937")
    right = Alignment(horizontal="right")
    center = Alignment(horizontal="center")

    ws["A1"] = "Senovka Plastics — Cheques"
    ws["A1"].font = Font(bold=True, size=14)
    ws.merge_cells("A1:I1")

    # Filter summary
    filter_bits = []
    if status:
        label = dict(Cheque.Status.choices).get(status, status.title())
        filter_bits.append(f"Status: {label}")
    if selected_customer:
        cust = Customer.objects.filter(pk=selected_customer).first()
        if cust:
            filter_bits.append(f"Customer: {cust.name}")
    if from_date:
        filter_bits.append(f"Maturity from {from_date.strftime('%d %b %Y')}")
    if to_date:
        filter_bits.append(f"Maturity to {to_date.strftime('%d %b %Y')}")

    ws["A2"] = "Filters"
    ws["A2"].font = Font(bold=True)
    ws["B2"] = " · ".join(filter_bits) if filter_bits else "None"
    ws.merge_cells("B2:F2")
    ws["G2"] = "Generated"
    ws["G2"].font = Font(bold=True)
    ws["H2"] = timezone.localtime().strftime("%d %b %Y %H:%M")
    ws.merge_cells("H2:I2")

    HEADERS = [
        "No", "Cheque No", "Customer", "Bank", "Branch",
        "Amount", "Received", "Maturity", "Status",
    ]
    header_row = 4
    for idx, name in enumerate(HEADERS, start=1):
        cell = ws.cell(row=header_row, column=idx, value=name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center

    row = header_row + 1
    total = Decimal("0.00")
    for i, cheque in enumerate(cheques, start=1):
        ws.cell(row=row, column=1, value=i).alignment = center
        ws.cell(row=row, column=2, value=cheque.cheque_no)
        ws.cell(row=row, column=3, value=cheque.customer.name)
        ws.cell(row=row, column=4, value=cheque.bank_name)
        ws.cell(row=row, column=5, value=cheque.branch or "")
        c = ws.cell(row=row, column=6, value=float(cheque.amount))
        c.number_format = "#,##0.00"; c.alignment = right
        ws.cell(row=row, column=7, value=cheque.received_date.strftime("%Y-%m-%d"))
        ws.cell(row=row, column=8, value=cheque.maturity_date.strftime("%Y-%m-%d"))
        ws.cell(row=row, column=9, value=cheque.get_status_display()).alignment = center
        total += cheque.amount
        row += 1

    # Total
    ws.cell(row=row, column=5, value="Total").font = Font(bold=True)
    ws.cell(row=row, column=5).alignment = right
    c = ws.cell(row=row, column=6, value=float(total))
    c.number_format = "#,##0.00"; c.alignment = right; c.font = Font(bold=True)

    ws.freeze_panes = f"A{header_row + 1}"

    widths = {"A": 5, "B": 14, "C": 26, "D": 20, "E": 16, "F": 14, "G": 12, "H": 12, "I": 12}
    for letter, width in widths.items():
        ws.column_dimensions[letter].width = width

    today = timezone.localdate()
    stamp = today.strftime("%Y-%m-%d")

    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    response = HttpResponse(
        stream.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = (
        f'attachment; filename="cheques_{stamp}.xlsx"'
    )
    return response


def cheque_list(request):
    today = timezone.localdate()
    horizon = today + timedelta(days=CHEQUE_WARNING_DAYS)

    status = request.GET.get("status", "").strip()
    if status not in {value for value, _ in Cheque.Status.choices}:
        status = ""

    customer_id = request.GET.get("customer", "").strip()
    selected_customer = int(customer_id) if customer_id.isdigit() else None

    from_date = _parse_date(request.GET.get("from_date"))
    to_date = _parse_date(request.GET.get("to_date"))

    cheques = Cheque.objects.select_related("customer")
    if status:
        cheques = cheques.filter(status=status)
    if selected_customer:
        cheques = cheques.filter(customer_id=selected_customer)
    if from_date:
        cheques = cheques.filter(maturity_date__gte=from_date)
    if to_date:
        cheques = cheques.filter(maturity_date__lte=to_date)

    # Counted off the filtered set rather than the page: this banner warns the
    # operator what is waiting on them across the whole filter, and a count
    # that only saw page 1 would quietly under-report it.
    due_count = cheques.filter(
        status=Cheque.Status.PENDING, maturity_date__lte=horizon
    ).count()

    page_obj = _paginate(request, cheques)
    for cheque in page_obj:
        # Maturing on us and still not banked: the row the operator is meant
        # to act on today. Anything already overdue counts too.
        cheque.is_due_soon = (
            cheque.status == Cheque.Status.PENDING and cheque.maturity_date <= horizon
        )

    return render(
        request,
        "core/cheque_list.html",
        {
            "page_obj": page_obj,
            "cheques": page_obj.object_list,
            "customers": Customer.objects.filter(cheques__isnull=False).distinct(),
            "status": status,
            "selected_customer": selected_customer,
            "from_date": from_date,
            "to_date": to_date,
            "statuses": Cheque.Status.choices,
            "is_filtered": bool(status or selected_customer or from_date or to_date),
            "due_count": due_count,
            "warning_days": CHEQUE_WARNING_DAYS,
            "today": today,
        },
    )


# -------------------------------------------------------------- cash drawer
# Money comes in only by saving a bill; this page is where it leaves.


def _account_banked(account):
    """What bill payments have put into one account.

    Read off CashTransfer, which is the only place an account is recorded.
    Manual transfers on this page are CashDrawer rows with no account column,
    so they lower the drawer without ever reaching this figure.
    """
    return CashTransfer.objects.filter(to_account=account).aggregate(
        total=Coalesce(Sum("amount"), ZERO, output_field=MONEY)
    )["total"]


def _is_manual(entry):
    """Whether this row was typed by hand rather than written by a bill.

    A bill-linked row belongs to that bill's payment: editing or deleting it
    here would put the drawer out of step with the bill it came from, and the
    bill would go on insisting the cash arrived. Those are corrected by editing
    the bill.
    """
    return entry.bill_id is None


def _cash_drawer_page(request, out_form, edit_form=None, edit_entry=None, in_form=None):
    """Render the drawer log.

    Shared by the list, cash_drawer_edit (which re-renders this whole page
    when a correction fails validation), and cash_drawer_insert (same story
    for a failed top-up) — the running balance, the totals and the filters
    all have to come back with it, and rebuilding them is this function.
    """
    # The edit modal is one form reused by every row, filled in by JS from the
    # row's data attributes. Even with nothing being edited it has to render its
    # widgets, or there would be no fields for that script to fill.
    if edit_form is None:
        edit_form = CashDrawerEditForm()
    if in_form is None:
        in_form = CashDrawerInForm(initial={"txn_date": timezone.localdate()})

    balance = _cash_drawer_balance()

    # Monthly filter — the same shape used across the rest of the app.
    # Defaults to the current month so a first page load is scoped, with the
    # All time toggle to see everything. From/to still accepted as an
    # optional finer-grain override (used by exports/bookmarks).
    month_filter = get_month_filter(request)
    from_date = _parse_date(request.GET.get("from_date"))
    to_date = _parse_date(request.GET.get("to_date"))

    entries = CashDrawer.objects.select_related("bill", "bill__customer")
    if from_date:
        entries = entries.filter(txn_date__gte=from_date)
    if to_date:
        entries = entries.filter(txn_date__lte=to_date)
    if not from_date and not to_date:
        entries = month_filter.apply(entries, field="txn_date")

    # Oldest first: a running balance read newest-first counts backwards.
    entries = entries.order_by("txn_date", "id")

    # Everything before the range still happened, so the running column starts
    # where the drawer actually stood — not at zero. Whether the range came
    # from the month filter or from explicit from/to, the same "everything
    # earlier still counts" rule applies — an undeposited float from last
    # month rolls forward into this month's opening balance automatically.
    if from_date:
        opening = _cash_drawer_balance(
            CashDrawer.objects.filter(txn_date__lt=from_date)
        )
    elif not month_filter.is_all_time:
        opening = _cash_drawer_balance(
            CashDrawer.objects.filter(txn_date__lt=month_filter.start)
        )
    else:
        opening = ZERO

    rows = []
    running = opening
    total_in = ZERO
    total_out = ZERO
    for entry in entries:
        is_in = entry.txn_type == CashDrawer.TxnType.IN
        running += entry.amount if is_in else -entry.amount
        if is_in:
            total_in += entry.amount
        else:
            total_out += entry.amount
        rows.append(
            {
                "entry": entry,
                "is_in": is_in,
                "running": running,
                # Drives the Actions column: only a hand-typed row gets buttons.
                "is_manual": _is_manual(entry),
            }
        )

    # The rows are paginated, the arithmetic above is not. Every row's running
    # balance depends on every row before it, and opening/closing/totals
    # describe the whole filtered range — so the sums are taken over all of it
    # and only the display is cut into pages. Slicing the queryset instead
    # would restart the running balance at each page and make the column lie.
    #
    # Fifty to a page, not the usual twenty-five: the drawer takes a row for
    # every cash bill, so its log is long, and it is read as a run of figures
    # down the running-balance column rather than scanned for a single row.
    page_obj = _paginate(request, rows, settings.PAGINATE_BY_REPORTS)

    return render(
        request,
        "core/cash_drawer.html",
        {
            "form": out_form,
            "in_form": in_form,
            "edit_form": edit_form,
            "edit_entry": edit_entry,
            # The drawer as it stands now, whatever the filter shows.
            "balance": balance,
            "senovka_banked": _account_banked(CashTransfer.Account.SENOVKA),
            "dinusha_banked": _account_banked(CashTransfer.Account.DINUSHA),
            "page_obj": page_obj,
            "rows": page_obj.object_list,
            "opening": opening,
            "closing": running,
            "total_in": total_in,
            "total_out": total_out,
            "from_date": from_date,
            "to_date": to_date,
            "month_filter": month_filter,
            # The opening is a *carry-forward* whenever the current view is
            # scoped by month or a from-date — the template surfaces it as
            # such so the operator can see undeposited cash from earlier
            # rolling into this scope.
            "carried_forward": (
                opening if (not month_filter.is_all_time or from_date) else ZERO
            ),
            "is_filtered": bool(
                from_date or to_date or not month_filter.is_all_time
            ),
            "kind_choices": CashDrawerOutForm.KIND_CHOICES,
        },
    )


@login_required
def cash_drawer_excel(request):
    """Download the cash drawer log as an .xlsx, honouring the from/to date
    filters.

    Same running-balance / totals arithmetic as the page, so the sheet reads
    as a printable copy of what is on screen — including the opening balance
    row when a date range is applied.
    """
    from io import BytesIO

    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    from_date = _parse_date(request.GET.get("from_date"))
    to_date = _parse_date(request.GET.get("to_date"))

    entries = CashDrawer.objects.select_related("bill", "bill__customer")
    if from_date:
        entries = entries.filter(txn_date__gte=from_date)
    if to_date:
        entries = entries.filter(txn_date__lte=to_date)
    entries = entries.order_by("txn_date", "id")

    opening = (
        _cash_drawer_balance(CashDrawer.objects.filter(txn_date__lt=from_date))
        if from_date
        else ZERO
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Cash Drawer"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1F2937")
    right = Alignment(horizontal="right")
    center = Alignment(horizontal="center")

    ws["A1"] = "Senovka Plastics — Cash Drawer Log"
    ws["A1"].font = Font(bold=True, size=14)
    ws.merge_cells("A1:F1")

    ws["A2"] = "From"
    ws["A2"].font = Font(bold=True)
    ws["B2"] = from_date.strftime("%d %b %Y") if from_date else "—"
    ws["C2"] = "To"
    ws["C2"].font = Font(bold=True)
    ws["D2"] = to_date.strftime("%d %b %Y") if to_date else "—"
    ws["E2"] = "Generated"
    ws["E2"].font = Font(bold=True)
    ws["F2"] = timezone.localtime().strftime("%d %b %Y %H:%M")

    HEADERS = ["Date", "Description", "In (+)", "Out (−)", "Running Balance", "Source"]
    header_row = 4
    for idx, name in enumerate(HEADERS, start=1):
        cell = ws.cell(row=header_row, column=idx, value=name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center

    row = header_row + 1

    # Opening balance line — only when a range is applied. Everything before
    # the range still happened, so the running column starts where the drawer
    # actually stood.
    if from_date:
        ws.cell(row=row, column=1, value=from_date.strftime("%Y-%m-%d"))
        ws.cell(row=row, column=2, value="Opening balance").font = Font(italic=True)
        c = ws.cell(row=row, column=5, value=float(opening))
        c.number_format = "#,##0.00"; c.alignment = right; c.font = Font(bold=True)
        row += 1

    running = opening
    total_in = ZERO
    total_out = ZERO

    for entry in entries:
        is_in = entry.txn_type == CashDrawer.TxnType.IN
        running += entry.amount if is_in else -entry.amount
        if is_in:
            total_in += entry.amount
        else:
            total_out += entry.amount

        # Description mirrors what the page shows.
        if entry.bill_id:
            desc = f"Bill #{entry.bill_id}"
            if entry.bill and entry.bill.customer_id:
                desc += f" · {entry.bill.customer.name}"
            if entry.reason:
                desc += f" — {entry.reason}"
        elif entry.reason:
            desc = entry.reason
        else:
            desc = entry.get_txn_type_display()

        source = f"Bill #{entry.bill_id}" if entry.bill_id else "Manual"

        ws.cell(row=row, column=1, value=entry.txn_date.strftime("%Y-%m-%d"))
        ws.cell(row=row, column=2, value=desc)

        if is_in:
            c = ws.cell(row=row, column=3, value=float(entry.amount))
            c.number_format = "#,##0.00"; c.alignment = right
        else:
            c = ws.cell(row=row, column=4, value=float(entry.amount))
            c.number_format = "#,##0.00"; c.alignment = right

        c = ws.cell(row=row, column=5, value=float(running))
        c.number_format = "#,##0.00"; c.alignment = right; c.font = Font(bold=True)

        ws.cell(row=row, column=6, value=source).alignment = center
        row += 1

    # Totals + closing balance.
    row += 1
    ws.cell(row=row, column=2, value="Totals").font = Font(bold=True)
    c = ws.cell(row=row, column=3, value=float(total_in))
    c.number_format = "#,##0.00"; c.alignment = right; c.font = Font(bold=True)
    c = ws.cell(row=row, column=4, value=float(total_out))
    c.number_format = "#,##0.00"; c.alignment = right; c.font = Font(bold=True)
    row += 1
    ws.cell(row=row, column=2, value="Closing balance").font = Font(bold=True)
    c = ws.cell(row=row, column=5, value=float(running))
    c.number_format = "#,##0.00"; c.alignment = right; c.font = Font(bold=True)

    ws.freeze_panes = f"A{header_row + 1}"

    widths = {"A": 12, "B": 42, "C": 14, "D": 14, "E": 16, "F": 18}
    for letter, width in widths.items():
        ws.column_dimensions[letter].width = width

    today = timezone.localdate()
    stamp = today.strftime("%Y-%m-%d")

    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    response = HttpResponse(
        stream.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = (
        f'attachment; filename="cash_drawer_{stamp}.xlsx"'
    )
    return response


@login_required
def cash_drawer(request):
    form = CashDrawerOutForm(
        request.POST or None,
        drawer_balance=_cash_drawer_balance(),
        initial={"txn_date": timezone.localdate()},
    )
    if request.method == "POST":
        if form.is_valid():
            entry = form.save()
            messages.success(
                request,
                f"{entry.reason} — {entry.amount:,.2f} out of the drawer. "
                f"{_cash_drawer_balance():,.2f} left.",
            )
            return redirect("core:cash_drawer")
        messages.error(request, "That entry couldn't be saved — see the form.")

    return _cash_drawer_page(request, form)


@require_POST
@login_required
def cash_drawer_insert(request):
    """Record cash coming into the drawer by hand.

    Mirror of `cash_drawer` for the OUT form. On success the running
    balance updates immediately (it is summed live from the CashDrawer
    rows on the next render). On failure the page comes back with the
    Insert modal's bound form so the operator can fix it in place.
    """
    form = CashDrawerInForm(request.POST)
    if form.is_valid():
        with transaction.atomic():
            entry = form.save()
        messages.success(
            request,
            f"{entry.reason} — {entry.amount:,.2f} into the drawer. "
            f"Balance: {_cash_drawer_balance():,.2f}.",
        )
        return redirect("core:cash_drawer")

    messages.error(request, "That insert couldn't be saved — see the form.")
    # Re-render the whole page so the Insert modal opens on the errors,
    # matching how cash_drawer_edit handles a failed edit.
    return _cash_drawer_page(
        request,
        CashDrawerOutForm(drawer_balance=_cash_drawer_balance()),
        in_form=form,
    )


def _drawer_balance_without(entry):
    """The drawer as it would stand if `entry` had never been written.

    What an edited entry has to be judged against: raising a 500 withdrawal to
    5000 is only affordable if the original 500 is put back first.
    """
    return _cash_drawer_balance(CashDrawer.objects.exclude(pk=entry.pk))


@login_required
def cash_drawer_edit(request, pk):
    entry = get_object_or_404(CashDrawer, pk=pk)
    if not _is_manual(entry):
        messages.error(
            request,
            f"That entry came from Bill #{entry.bill_id} and is part of its "
            f"payment. Edit the bill instead.",
        )
        return redirect("core:cash_drawer")

    # The form lives in a modal on the list, so there is nothing to GET.
    if request.method != "POST":
        return redirect("core:cash_drawer")

    form = CashDrawerEditForm(
        request.POST,
        instance=entry,
        drawer_balance=_drawer_balance_without(entry),
    )
    if form.is_valid():
        edited = form.save(commit=False)
        edited.edited_at = timezone.now()
        edited.edited_by = request.user
        edited.save()

        messages.success(
            request,
            f"The {edited.txn_date:%d %b %Y} entry was updated. "
            f"The drawer now holds {_cash_drawer_balance():,.2f}.",
        )
        return redirect("core:cash_drawer")

    # Straight back to the list with the modal open on the errors, rather than
    # a redirect that would throw away what was typed.
    messages.error(request, "That correction couldn't be saved — see the form.")
    return _cash_drawer_page(
        request, CashDrawerOutForm(drawer_balance=_cash_drawer_balance()), form, entry
    )


@require_POST
@login_required
def cash_drawer_delete(request, pk):
    entry = get_object_or_404(CashDrawer, pk=pk)
    if not _is_manual(entry):
        messages.error(
            request,
            f"That entry came from Bill #{entry.bill_id} and is part of its "
            f"payment. Delete the bill instead.",
        )
        return redirect("core:cash_drawer")

    label = f"{entry.get_txn_type_display()} of {entry.amount:,.2f} on {entry.txn_date:%d %b %Y}"
    # No balance to reverse: nothing stores the drawer total — every figure on
    # the page is summed from the rows on each render, so a row that is gone is
    # simply no longer counted.
    entry.delete()

    messages.success(
        request,
        f"{label} was deleted. The drawer now holds {_cash_drawer_balance():,.2f}.",
    )
    return redirect("core:cash_drawer")


# ----------------------------------------------------------- supplier bills
# The mirror of a sales bill: stock comes in and the balance moves the other
# way. A positive balance is what we owe them.


class SupplierBillError(BillError):
    """A reason to roll the save back, worded for the operator.

    Subclasses BillError because the two paths share their parsing helpers —
    _decimal raises BillError, and a supplier bill has to catch that as readily
    as its own complaints rather than let it escape as a 500.
    """


def _supplier_products():
    """Everything a supplier bill may receive, for the line dropdown."""
    return [
        {
            "id": product.pk,
            "name": product.name,
            "size": product.size,
            "label": str(product),
            "default_price": f"{product.default_price:.2f}",
        }
        for product in Product.objects.filter(is_active=True).order_by("name", "size")
    ]


def _read_supplier_lines(payload):
    raw_lines = payload.get("lines")
    if not isinstance(raw_lines, list) or not raw_lines:
        raise SupplierBillError("Add at least one product line.")

    products = {
        product.pk: product
        for product in Product.objects.filter(
            pk__in=[raw.get("product_id") for raw in raw_lines if isinstance(raw, dict)]
        )
    }

    items = []
    seen = set()
    for raw in raw_lines:
        if not isinstance(raw, dict):
            raise SupplierBillError(MALFORMED)

        product = products.get(raw.get("product_id"))
        if product is None:
            raise SupplierBillError("A product on this bill no longer exists.")
        if product.pk in seen:
            raise SupplierBillError(f"{product} is on the bill twice.")
        seen.add(product.pk)

        qty = _decimal(raw.get("qty"), "Quantity", 3)
        if qty <= ZERO:
            raise SupplierBillError(f"Quantity for {product} must be above 0.")
        unit_price = _decimal(raw.get("unit_price"), "Unit price", 2)

        # Recomputed: a line total off the browser is a number someone typed.
        line_total = (qty * unit_price).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
        items.append(
            {
                "product": product,
                "qty": qty,
                "unit_price": unit_price,
                "line_total": line_total,
            }
        )
    return items


def _reverse_supplier_bill(bill):
    """Undo a supplier bill: stock back out, and the debt to them cancelled.

    Guarded, because received stock may already have been sold on. Taking it
    back out regardless would leave a product holding a negative quantity, and
    that product then vanishes from every sales screen.
    """
    for item in bill.items.all():
        moved = Product.objects.filter(
            pk=item.product_id, qty__gte=item.qty
        ).update(qty=F("qty") - item.qty)
        if not moved:
            product = Product.objects.get(pk=item.product_id)
            raise SupplierBillError(
                f"Can't reverse {product}: {item.qty:.3f} was received but only "
                f"{product.qty:.3f} is left, so some of it has been sold on."
            )

    Customer.objects.filter(pk=bill.supplier_id).update(
        balance=F("balance") - bill.total_amount
    )
    bill.items.all().delete()


def _write_supplier_bill(bill, payload):
    supplier = Customer.objects.filter(
        pk=payload.get("supplier_id"), is_supplier=True
    ).first()
    if supplier is None:
        raise SupplierBillError("Choose a supplier.")

    items = _read_supplier_lines(payload)
    total = sum((item["line_total"] for item in items), ZERO)

    # 1. header. An edit keeps the date the goods actually arrived.
    bill.supplier = supplier
    if bill.pk is None:
        bill.bill_date = timezone.localdate()
    bill.total_amount = total
    # Paying suppliers isn't built yet, so nothing has been paid on it.
    bill.paid_amount = ZERO
    bill.status = SupplierBill.Status.UNPAID
    bill.notes = str(payload.get("notes") or "").strip()
    bill.save()

    # 2. lines, and the stock they bring in
    for item in items:
        SupplierBillItem.objects.create(
            supplier_bill=bill,
            product=item["product"],
            qty=item["qty"],
            unit_price=item["unit_price"],
            line_total=item["line_total"],
        )
        Product.objects.filter(pk=item["product"].pk).update(qty=F("qty") + item["qty"])

    # 3. we owe them the lot. Positive is credit in their favour, so the sign
    # runs opposite to a sales bill. F() so a concurrent move is adjusted, not
    # overwritten.
    Customer.objects.filter(pk=supplier.pk).update(balance=F("balance") + total)

    return bill


@transaction.atomic
def _save_supplier_bill(payload):
    return _write_supplier_bill(SupplierBill(), payload)


@transaction.atomic
def _update_supplier_bill(bill, payload):
    _reverse_supplier_bill(bill)
    return _write_supplier_bill(bill, payload)


def _supplier_bill_payload(request):
    try:
        payload = json.loads(request.body or b"{}")
    except json.JSONDecodeError:
        return None
    return payload if isinstance(payload, dict) else None


@require_POST
@login_required
def supplier_quick_create(request):
    """Create a supplier without leaving the bill form."""
    form = SupplierQuickForm(request.POST)
    if not form.is_valid():
        return JsonResponse({"success": False, "errors": form.errors}, status=400)

    supplier = form.save()
    return JsonResponse(
        {"success": True, "supplier": {"id": supplier.pk, "name": supplier.name}}
    )


@require_POST
@login_required
def product_quick_create(request):
    """Create a product without leaving the bill form."""
    form = ProductQuickForm(request.POST)
    if not form.is_valid():
        return JsonResponse({"success": False, "errors": form.errors}, status=400)

    product = form.save()
    return JsonResponse(
        {
            "success": True,
            "product": {
                "id": product.pk,
                "name": product.name,
                "size": product.size,
                "label": str(product),
                "default_price": f"{product.default_price:.2f}",
            },
        }
    )


def _supplier_bill_form_context(request, bill=None):
    return {
        "bill": bill,
        "suppliers": Customer.objects.filter(is_supplier=True),
        "products_json": _supplier_products(),
        "categories": Category.objects.all(),
        "product_form": ProductQuickForm(),
        "supplier_form": SupplierQuickForm(),
        "is_edit": bill is not None,
        "save_url": (
            reverse("core:supplier_bill_edit", args=[bill.pk])
            if bill
            else reverse("core:supplier_bill_create")
        ),
        "initial": (
            {
                "supplier_id": bill.supplier_id,
                "lines": [
                    {
                        "product_id": item.product_id,
                        "qty": f"{item.qty:.3f}".rstrip("0").rstrip("."),
                        "unit_price": f"{item.unit_price:.2f}",
                    }
                    for item in bill.items.all()
                ],
                "notes": bill.notes,
            }
            if bill
            else None
        ),
    }


@login_required
def supplier_bill_create(request):
    if request.method == "POST":
        payload = _supplier_bill_payload(request)
        if payload is None:
            return JsonResponse({"success": False, "error": MALFORMED}, status=400)
        try:
            bill = _save_supplier_bill(payload)
        except BillError as exc:
            return JsonResponse({"success": False, "error": str(exc)}, status=400)

        messages.success(
            request, f"Supplier bill #{bill.pk} for {bill.supplier.name} was saved."
        )
        return JsonResponse(
            {
                "success": True,
                "redirect": reverse("core:supplier_bill_detail", args=[bill.pk]),
            }
        )

    return render(
        request, "core/supplier_bill_create.html", _supplier_bill_form_context(request)
    )


@login_required
def supplier_bill_edit(request, pk):
    bill = get_object_or_404(SupplierBill.objects.select_related("supplier"), pk=pk)

    if request.method == "POST":
        payload = _supplier_bill_payload(request)
        if payload is None:
            return JsonResponse({"success": False, "error": MALFORMED}, status=400)
        try:
            bill = _update_supplier_bill(bill, payload)
        except BillError as exc:
            # Atomic, so the reversal it began is undone with it.
            return JsonResponse({"success": False, "error": str(exc)}, status=400)

        messages.success(request, f"Supplier bill #{bill.pk} was updated.")
        return JsonResponse(
            {
                "success": True,
                "redirect": reverse("core:supplier_bill_detail", args=[bill.pk]),
            }
        )

    return render(
        request, "core/supplier_bill_edit.html", _supplier_bill_form_context(request, bill)
    )


@require_POST
@super_admin_required
def supplier_bill_delete(request, pk):
    bill = get_object_or_404(SupplierBill.objects.select_related("supplier"), pk=pk)
    label = f"Supplier bill #{bill.pk}"
    supplier = bill.supplier.name

    try:
        with transaction.atomic():
            _reverse_supplier_bill(bill)
            bill.delete()
    except BillError as exc:
        messages.error(request, f"Cannot delete {label} — {exc}")
        return redirect("core:supplier_bill_detail", pk=pk)

    messages.success(request, f"{label} for {supplier} was deleted and reversed.")
    return redirect("core:supplier_bill_list")


@login_required
def supplier_bill_detail(request, pk):
    bill = get_object_or_404(
        SupplierBill.objects.select_related("supplier").annotate(
            item_count=Count("items")
        ),
        pk=pk,
    )
    return render(
        request,
        "core/supplier_bill_detail.html",
        {"bill": bill, "items": bill.items.select_related("product")},
    )


@login_required
def supplier_bill_list(request):
    from_date = _parse_date(request.GET.get("from_date"))
    to_date = _parse_date(request.GET.get("to_date"))

    supplier_id = request.GET.get("supplier", "").strip()
    selected_supplier = int(supplier_id) if supplier_id.isdigit() else None

    status = request.GET.get("status", "").strip()
    if status not in {value for value, _ in SupplierBill.Status.choices}:
        status = ""

    # order_by repeats SupplierBill.Meta.ordering, which the annotate() below
    # would otherwise drop — see _bills_with_counts.
    bills = (
        SupplierBill.objects.select_related("supplier")
        .annotate(item_count=Count("items"))
        .order_by("-bill_date", "-id")
    )
    if from_date:
        bills = bills.filter(bill_date__gte=from_date)
    if to_date:
        bills = bills.filter(bill_date__lte=to_date)
    if selected_supplier:
        bills = bills.filter(supplier_id=selected_supplier)
    if status:
        bills = bills.filter(status=status)

    page_obj = _paginate(request, bills)

    return render(
        request,
        "core/supplier_bill_list.html",
        {
            "page_obj": page_obj,
            "bills": page_obj.object_list,
            "suppliers": Customer.objects.filter(is_supplier=True),
            "from_date": from_date,
            "to_date": to_date,
            "selected_supplier": selected_supplier,
            "status": status,
            "statuses": SupplierBill.Status.choices,
            "is_filtered": bool(
                from_date or to_date or selected_supplier or status
            ),
        },
    )


# ------------------------------------------------------------- production
# Stock made in-house. The only other thing that puts stock on the shelf is a
# supplier bill; everything else takes it off.


class ProductionError(BillError):
    """A reason to roll the save back, worded for the operator.

    Subclasses BillError so the shared _decimal parser's complaints are caught
    here too rather than escaping as a 500.
    """


def _move_stock(product, delta):
    """Apply a signed change to a product's stock.

    Guarded downwards: correcting or removing production takes stock back off
    the shelf, and it may already have been sold. Letting a product hold a
    negative quantity would drop it out of every sales screen, which filters on
    qty > 0.
    """
    if delta >= ZERO:
        Product.objects.filter(pk=product.pk).update(qty=F("qty") + delta)
        return

    needed = -delta
    moved = Product.objects.filter(pk=product.pk, qty__gte=needed).update(
        qty=F("qty") - needed
    )
    if not moved:
        product.refresh_from_db()
        raise ProductionError(
            f"Can't take {needed:.3f} back off {product} — only {product.qty:.3f} "
            f"is left, so some of it has been sold."
        )


@transaction.atomic
def _save_production(payload):
    """Write one day's production, or none of it."""
    production_date = _parse_date(payload.get("production_date"))
    if production_date is None:
        raise ProductionError("Choose a production date.")
    if production_date > timezone.localdate():
        raise ProductionError("Production can't be dated in the future.")

    raw_lines = payload.get("lines")
    if not isinstance(raw_lines, list):
        raise ProductionError(MALFORMED)

    products = {
        product.pk: product
        for product in Product.objects.filter(
            pk__in=[raw.get("product_id") for raw in raw_lines if isinstance(raw, dict)],
            is_active=True,
        )
    }

    entries = []
    seen = set()
    for raw in raw_lines:
        if not isinstance(raw, dict):
            raise ProductionError(MALFORMED)

        qty = _decimal(raw.get("qty_produced"), "Quantity produced", 3)
        # Only rows with something on them are saved; the rest of the table is
        # just the shelf, sitting there at zero.
        if qty == ZERO:
            continue

        product = products.get(raw.get("product_id"))
        if product is None:
            raise ProductionError("A product on this sheet is no longer available.")
        if product.pk in seen:
            raise ProductionError(f"{product} is on the sheet twice.")
        seen.add(product.pk)

        # Required on any row that produced something — a quantity with nothing
        # to explain it is what the reason field exists to prevent. Rows left at
        # zero were skipped above and are never asked for one.
        reason = str(raw.get("reason") or "").strip()
        if not reason:
            raise ProductionError(f"Give a reason for the {product} production.")
        if len(reason) > 500:
            raise ProductionError(
                f"The reason for {product} is too long — keep it under 500 characters."
            )

        entries.append((product, qty, reason))

    if not entries:
        raise ProductionError("Enter a quantity against at least one product.")

    written = []
    for product, qty, reason in entries:
        # Read inside the transaction: the snapshot has to be the shelf as this
        # entry found it, not as the page rendered it some minutes ago.
        product.refresh_from_db()
        before = product.qty

        written.append(
            ProductionEntry.objects.create(
                product=product,
                production_date=production_date,
                qty_produced=qty,
                reason=reason,
                stock_before=before,
                stock_after=before + qty,
            )
        )
        _move_stock(product, qty)

    return production_date, written


@transaction.atomic
def _update_production(entry, qty, reason, production_date):
    """Correct one entry, moving the shelf by the difference.

    The stored quantity is re-read rather than taken off `entry`: a bound
    ModelForm writes the submitted value onto its instance while validating,
    so by now entry.qty_produced is already the new figure and the difference
    against it would always be zero.
    """
    stored = ProductionEntry.objects.select_related("product").get(pk=entry.pk)
    diff = qty - stored.qty_produced
    _move_stock(stored.product, diff)

    entry.qty_produced = qty
    entry.reason = reason
    entry.production_date = production_date
    # stock_before stays: it is what this entry found, and no correction now
    # changes what was on the shelf then. What it left behind does change.
    entry.stock_after = stored.stock_before + qty
    entry.save(
        update_fields=["qty_produced", "reason", "production_date", "stock_after"]
    )
    return entry


@transaction.atomic
def _delete_production(entry):
    _move_stock(entry.product, -entry.qty_produced)
    entry.delete()


@login_required
def production_create(request):
    if request.method == "POST":
        try:
            payload = json.loads(request.body or b"{}")
        except json.JSONDecodeError:
            return JsonResponse({"success": False, "error": MALFORMED}, status=400)
        if not isinstance(payload, dict):
            return JsonResponse({"success": False, "error": MALFORMED}, status=400)

        try:
            production_date, written = _save_production(payload)
        except BillError as exc:
            return JsonResponse({"success": False, "error": str(exc)}, status=400)

        total = sum((entry.qty_produced for entry in written), ZERO)
        messages.success(
            request,
            f"{len(written)} product{'' if len(written) == 1 else 's'} produced on "
            f"{production_date:%d %b %Y} — {total:.3f} in total.",
        )
        return JsonResponse(
            {"success": True, "redirect": reverse("core:production_list")}
        )

    products = Product.objects.select_related("category")
    return render(
        request,
        "core/production_create.html",
        {
            "products": products.filter(is_active=True),
            "today": timezone.localdate(),
        },
    )


@login_required
def production_edit(request, pk):
    entry = get_object_or_404(
        ProductionEntry.objects.select_related("product"), pk=pk
    )

    # The shelf as it stands before any of this is applied, for the read-only
    # panel on the form. Read now: _update_production moves it.
    stock_now = entry.product.qty
    was_qty = entry.qty_produced

    form = ProductionEntryForm(request.POST or None, instance=entry)
    if request.method == "POST" and form.is_valid():
        try:
            _update_production(
                entry,
                form.cleaned_data["qty_produced"],
                form.cleaned_data["reason"],
                form.cleaned_data["production_date"],
            )
        except BillError as exc:
            form.add_error("qty_produced", str(exc))
        else:
            # The stock moved by an F() expression, so the product in memory
            # still holds the figure from before.
            entry.product.refresh_from_db()
            messages.success(
                request,
                f"{entry.product} production on {entry.production_date:%d %b %Y} "
                f"is now {entry.qty_produced:.3f}. Stock is {entry.product.qty:.3f}.",
            )
            return redirect("core:production_list")

    return render(
        request,
        "core/production_edit.html",
        {
            "form": form,
            "entry": entry,
            "stock_now": stock_now,
            # A bound form has already written the submitted qty onto `entry`,
            # so the template can't read the stored one off it.
            "was_qty": was_qty,
        },
    )


@require_POST
@super_admin_required
def production_delete(request, pk):
    entry = get_object_or_404(
        ProductionEntry.objects.select_related("product"), pk=pk
    )
    label = f"{entry.product} on {entry.production_date:%d %b %Y}"

    try:
        _delete_production(entry)
    except BillError as exc:
        messages.error(request, f"Cannot delete {label} — {exc}")
        return redirect("core:production_list")

    messages.success(request, f"Production of {label} was deleted and reversed.")
    return redirect("core:production_list")


@login_required
def production_list(request):
    from_date = _parse_date(request.GET.get("from_date"))
    to_date = _parse_date(request.GET.get("to_date"))

    product_id = request.GET.get("product", "").strip()
    selected_product = int(product_id) if product_id.isdigit() else None

    entries = ProductionEntry.objects.select_related("product")
    if from_date:
        entries = entries.filter(production_date__gte=from_date)
    if to_date:
        entries = entries.filter(production_date__lte=to_date)
    if selected_product:
        entries = entries.filter(product_id=selected_product)

    # Newest day first, and within a day the order they were entered.
    entries = entries.order_by("-production_date", "id")

    # Grouped here rather than by a second query per day: the rows are already
    # in hand and already in the right order.
    days = []
    for entry in entries:
        if not days or days[-1]["date"] != entry.production_date:
            days.append(
                {
                    "date": entry.production_date,
                    "entries": [],
                    "total_qty": ZERO,
                }
            )
        days[-1]["entries"].append(entry)
        days[-1]["total_qty"] += entry.qty_produced

    for day in days:
        day["product_count"] = len(day["entries"])

    # Paginated by day, not by entry: a day is one row here — the entries sit
    # inside it, behind the expander — and its product_count and total_qty
    # describe the whole day. Splitting one across a page boundary would leave
    # both halves reporting a total that was never produced.
    page_obj = _paginate(request, days)

    return render(
        request,
        "core/production_list.html",
        {
            "page_obj": page_obj,
            "days": page_obj.object_list,
            "products": Product.objects.filter(production_entries__isnull=False).distinct(),
            "from_date": from_date,
            "to_date": to_date,
            "selected_product": selected_product,
            "is_filtered": bool(from_date or to_date or selected_product),
            # Entries across every day the filter matched, not just this page's.
            # The pager counts days, which is what a row is here, so this is the
            # only figure that says how much production that adds up to.
            "entry_count": sum(day["product_count"] for day in days),
        },
    )


@login_required
def ledger_index(request):
    """The sidebar's Customer Ledger entry: pick a customer, get their ledger.

    Choosing one redirects to that customer's own ledger rather than drawing a
    second copy here. One ledger page, one set of date filters, one PDF export
    — two would only be two things to keep in step.
    """
    picked = request.GET.get("customer", "").strip()
    if picked.isdigit() and Customer.objects.filter(pk=picked).exists():
        return redirect("core:customer_ledger", pk=int(picked))

    query = request.GET.get("q", "").strip()
    customers = Customer.objects.annotate(
        owed=Case(
            When(balance__lt=0, then=Value(0) - F("balance")),
            default=Value(ZERO),
            output_field=MONEY,
        )
    )
    if query:
        customers = customers.filter(name__icontains=query)

    return render(
        request,
        "core/ledger_index.html",
        {
            # Biggest debt first: the accounts most likely to be looked up.
            "customers": customers.order_by("-owed", "name"),
            "query": query,
            "total_count": Customer.objects.count(),
        },
    )


# -------------------------------------------------------------- sales report
# Read-only. Every figure is derived from bills and their payments, so nothing
# here writes and nothing here should ever disagree with the bill it came from.


def _sales_report_context(request):
    """Everything both the page and the PDF report, from one set of filters."""
    from_date = _parse_date(request.GET.get("from_date"))
    to_date = _parse_date(request.GET.get("to_date"))

    customer_id = request.GET.get("customer_id", "").strip()
    selected_customer = int(customer_id) if customer_id.isdigit() else None

    payment_type = request.GET.get("payment_type", "").strip()
    if payment_type not in {value for value, _ in Bill.PaymentType.choices}:
        payment_type = ""

    # A cancelled bill is not a sale. Leaving them in would overstate every
    # card on the page.
    bills = Bill.objects.select_related("customer").exclude(
        status=Bill.Status.CANCELLED
    )
    if from_date:
        bills = bills.filter(bill_date__gte=from_date)
    if to_date:
        bills = bills.filter(bill_date__lte=to_date)
    if selected_customer:
        bills = bills.filter(customer_id=selected_customer)
    if payment_type:
        bills = bills.filter(payment_type=payment_type)

    bills = bills.annotate(
        # paid_amount can run past total_amount when a payment also clears old
        # debt, and a bill can't owe less than nothing.
        outstanding=Greatest(
            F("total_amount") - F("paid_amount"), Value(ZERO), output_field=MONEY
        )
    ).order_by("bill_date", "id")

    bills = list(bills)
    bill_ids = [bill.pk for bill in bills]

    totals = Bill.objects.filter(pk__in=bill_ids).aggregate(
        sales=Coalesce(Sum("total_amount"), ZERO, output_field=MONEY),
    )
    total_outstanding = sum((bill.outstanding for bill in bills), ZERO)

    # Payments are counted through their bill, so the same date range and the
    # same filters apply to both without asking twice.
    payments = Payment.objects.filter(bill_id__in=bill_ids).select_related(
        "bill", "bill__customer"
    )

    cash_rows = list(
        payments.filter(method=Payment.Method.CASH).order_by("bill__bill_date", "id")
    )
    cheque_rows = list(
        Cheque.objects.filter(payment__bill_id__in=bill_ids)
        .select_related("payment", "payment__bill", "customer")
        .order_by("payment__bill__bill_date", "id")
    )

    total_cash = sum((row.amount for row in cash_rows), ZERO)
    total_cheque = sum((row.amount for row in cheque_rows), ZERO)

    # Cash either stayed in the drawer or was banked; Payment.account says
    # which, and blank means it stayed.
    account_labels = dict(Payment.Account.choices)
    by_account = {"": ZERO}
    for value, _ in Payment.Account.choices:
        by_account[value] = ZERO
    for row in cash_rows:
        by_account[row.account] = by_account.get(row.account, ZERO) + row.amount

    account_totals = [("Physical", by_account.get("", ZERO))]
    for value, label in Payment.Account.choices:
        account_totals.append((f"{label} Acc", by_account.get(value, ZERO)))

    return {
        "bills": bills,
        "cash_rows": cash_rows,
        "cheque_rows": cheque_rows,
        "account_labels": account_labels,
        "account_totals": account_totals,
        "total_sales": totals["sales"],
        "total_cash": total_cash,
        "total_cheque": total_cheque,
        "total_outstanding": total_outstanding,
        "from_date": from_date,
        "to_date": to_date,
        "selected_customer": selected_customer,
        "payment_type": payment_type,
        "customers": Customer.objects.filter(is_walk_in_account=False),
        "payment_types": Bill.PaymentType.choices,
        "is_filtered": bool(
            from_date or to_date or selected_customer or payment_type
        ),
        "generated_at": timezone.localtime(),
    }


@login_required
def sales_report(request):
    context = _sales_report_context(request)
    # The PDF link carries the same filters, so it reports what is on screen.
    context["query"] = request.GET.urlencode()

    # Paged here and not in _sales_report_context, which sales_report_pdf also
    # calls: the totals on this page are struck over every bill in the range,
    # and the PDF is the whole report. Paging the shared builder would cut both
    # down to 50 bills.
    page_obj = _paginate(request, context["bills"], settings.PAGINATE_BY_REPORTS)
    context["page_obj"] = page_obj
    context["bills"] = page_obj.object_list
    return render(request, "core/sales_report.html", context)


def _pdf_response(request, template, context, filename):
    """Render a print template to PDF, or to itself when that isn't possible.

    Shared by every report. The templates are written to stand up unaided —
    WeasyPrint fetches nothing and runs no JavaScript — which is what lets the
    fallback hand the very same document to the browser to print.
    """
    html = render_to_string(template, context, request=request)

    try:
        from weasyprint import HTML
    except (ImportError, OSError):
        # OSError, not just ImportError: `pip install weasyprint` succeeds on
        # Windows and then importing it fails, because the GTK libraries it
        # binds to are not something pip can deliver. Rather than 500, say so
        # and hand back the document.
        messages.warning(
            request,
            "WeasyPrint can't run here, so this is the print view rather than a "
            "PDF download — use your browser's Print to PDF. To get real PDFs, "
            "install WeasyPrint's GTK libraries on the server.",
        )
        return HttpResponse(html)

    pdf = HTML(string=html, base_url=request.build_absolute_uri()).write_pdf()
    response = HttpResponse(pdf, content_type="application/pdf")
    response["Content-Disposition"] = f'inline; filename="{filename}"'
    return response


@login_required
def sales_report_pdf(request):
    stamp = timezone.localdate().isoformat()
    return _pdf_response(
        request,
        "core/sales_report_pdf.html",
        _sales_report_context(request),
        f"senovka-sales-{stamp}.pdf",
    )


# ------------------------------------------------------------- ledger report


@login_required
def customer_ledger_pdf(request, pk):
    """The per-customer ledger as a document, off the same rows as the page."""
    customer = get_object_or_404(_customers(), pk=pk)

    from_date = _parse_date(request.GET.get("from_date"))
    to_date = _parse_date(request.GET.get("to_date"))
    rows = _ledger_rows(customer, from_date, to_date)

    context = {
        "customer": customer,
        "rows": rows,
        "from_date": from_date,
        "to_date": to_date,
        "is_filtered": bool(from_date or to_date),
        "total_sale": sum((row["sale"] or ZERO for row in rows), ZERO),
        "total_credit": sum((row["credit"] or ZERO for row in rows), ZERO),
        "closing_balance": rows[-1]["balance"] if rows else ZERO,
        # What the closing balance is *as of*: the end of the range asked for,
        # or today when the range runs to now.
        "as_of": to_date or timezone.localdate(),
        "generated_at": timezone.localtime(),
    }

    # slugify: a customer name is free text, and a raw one in a filename header
    # is at best broken and at worst a way to inject a header.
    stamp = timezone.localdate().isoformat()
    filename = f"ledger_{slugify(customer.name) or customer.pk}_{stamp}.pdf"
    return _pdf_response(request, "core/ledger_pdf.html", context, filename)


def _write_customer_ledger_sheet(ws, customer, from_date=None, to_date=None):
    """Write one customer's ledger onto worksheet `ws`.

    Same six-column shape as the on-screen ledger: date, description, sale,
    credit (money in / discount off), running balance, remaining-on-bill. The
    sheet builder is deliberately its own function so the bulk export drops in
    without duplicating the styling, and a later single-customer Excel button
    can reuse it without any refactor.
    """
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    rows = _ledger_rows(customer, from_date, to_date)

    thin = Side(style="thin")
    border_all = Border(top=thin, bottom=thin, left=thin, right=thin)
    bold = Font(bold=True)
    bold_large = Font(bold=True, size=14)
    center = Alignment(horizontal="center", vertical="center")
    right = Alignment(horizontal="right", vertical="center")
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)

    header_fill = PatternFill("solid", fgColor="1F2937")
    header_font = Font(bold=True, color="FFFFFF")
    fill_note = PatternFill("solid", fgColor="F1F5F9")
    fill_sale = PatternFill("solid", fgColor="FFF1F2")
    fill_credit = PatternFill("solid", fgColor="ECFDF5")

    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 44
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 18
    ws.column_dimensions["F"].width = 16

    # Title band.
    ws["A1"] = "Senovka Plastics — Customer Ledger"
    ws["A1"].font = bold_large
    ws.merge_cells("A1:F1")

    ws["A2"] = "Customer:"; ws["B2"] = customer.name; ws["B2"].font = bold
    ws["C2"] = "Phone:"; ws["D2"] = customer.phone or "—"; ws["D2"].font = bold

    period = "All time"
    if from_date and to_date:
        period = f"{from_date:%d %b %Y} → {to_date:%d %b %Y}"
    elif from_date:
        period = f"From {from_date:%d %b %Y}"
    elif to_date:
        period = f"Up to {to_date:%d %b %Y}"
    ws["E2"] = "Period:"; ws["F2"] = period; ws["F2"].font = bold

    for cell_addr in ["A2", "C2", "E2"]:
        ws[cell_addr].font = bold

    # Summary row.
    total_sale = sum((r["sale"] or ZERO for r in rows), ZERO)
    total_credit = sum((r["credit"] or ZERO for r in rows), ZERO)
    closing = rows[-1]["balance"] if rows else ZERO

    ws["A4"] = "Total Billed"; ws["B4"] = float(total_sale)
    ws["C4"] = "Total Received / Credited"; ws["D4"] = float(total_credit)
    ws["E4"] = "Closing Balance"; ws["F4"] = float(closing)
    for addr in ("B4", "D4", "F4"):
        ws[addr].font = bold
        ws[addr].number_format = "#,##0.00"
    for addr in ("A4", "C4", "E4"):
        ws[addr].font = bold

    # Table header.
    headers = ["Date", "Description", "Sale (+)", "Credit (-)", "Balance", "Bill #"]
    header_row = 6
    for idx, name in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=idx, value=name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border_all

    row_num = header_row + 1
    for r in rows:
        c_date = ws.cell(row=row_num, column=1, value=r["date"].strftime("%d %b %Y"))
        c_desc = ws.cell(row=row_num, column=2, value=r["description"])
        c_sale = ws.cell(
            row=row_num, column=3,
            value=float(r["sale"]) if r["sale"] is not None else "",
        )
        c_credit = ws.cell(
            row=row_num, column=4,
            value=float(r["credit"]) if r["credit"] is not None else "",
        )
        c_bal = ws.cell(row=row_num, column=5, value=float(r["balance"]))
        c_ref = ws.cell(
            row=row_num, column=6,
            value=(f"Bill #{r['bill_pk']}" if r.get("bill_pk") else ""),
        )

        c_date.alignment = center
        c_desc.alignment = left
        c_sale.alignment = right
        c_credit.alignment = right
        c_bal.alignment = right
        c_ref.alignment = center

        for cell in (c_sale, c_credit, c_bal):
            cell.number_format = "#,##0.00"

        row_fill = None
        if r.get("is_note"):
            row_fill = fill_note
        elif r.get("sale"):
            row_fill = fill_sale
        elif r.get("credit"):
            row_fill = fill_credit

        for c_idx in range(1, 7):
            cell = ws.cell(row=row_num, column=c_idx)
            cell.border = border_all
            if row_fill:
                cell.fill = row_fill
            if r.get("is_note"):
                cell.font = Font(italic=True, color="475569")

        row_num += 1

    if not rows:
        # Empty ledger still gets a "nothing to show" row so the file explains
        # itself when opened rather than looking like a corrupted export.
        cell = ws.cell(
            row=row_num, column=1,
            value="No ledger activity in the selected range.",
        )
        cell.font = Font(italic=True, color="94A3B8")
        ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=6)


@login_required
def customer_ledger_bulk_excel(request):
    """Multiple customers' ledgers in one workbook — one sheet each.

    Fired by the "Download ledger (Excel)" button on the customer list. Same
    date-range knobs as the on-screen ledger (`from_date`, `to_date` in the
    query string) so a filtered view can be exported without giving up the
    bounds.
    """
    from openpyxl import Workbook

    ids = _parse_id_list(request)
    if not ids:
        messages.error(request, "Pick at least one customer first.")
        return redirect("core:customer_list")

    from_date = _parse_date(request.GET.get("from_date"))
    to_date = _parse_date(request.GET.get("to_date"))

    customers = list(_customers().filter(pk__in=ids))
    if not customers:
        messages.error(request, "None of the picked customers exist.")
        return redirect("core:customer_list")

    wb = Workbook()
    default_ws = wb.active
    wb.remove(default_ws)

    used_titles = set()
    for customer in customers:
        ws = wb.create_sheet(title=_sheet_title_for(customer.name, used_titles))
        _write_customer_ledger_sheet(ws, customer, from_date, to_date)

    stamp = timezone.localdate().isoformat()
    return _xlsx_response(
        wb, f"customer_ledgers_{len(customers)}customers_{stamp}.xlsx"
    )


# -------------------------------------------------------- outstanding report


def _outstanding_context(request):
    """Every customer's account at a glance, worst debt first."""
    scope = request.GET.get("scope", "owing")
    if scope not in {"owing", "all"}:
        scope = "owing"

    # Subqueries, not Sum() over joins: totalling bills and payments in one
    # query would count each bill once per payment on it. These each aggregate
    # on their own and hand back a single figure.
    live_bills = Bill.objects.filter(customer=OuterRef("pk")).exclude(
        status=Bill.Status.CANCELLED
    )
    billed = (
        live_bills.values("customer").annotate(total=Sum("total_amount")).values("total")
    )
    received = (
        Payment.objects.filter(bill__customer=OuterRef("pk"))
        .exclude(bill__status=Bill.Status.CANCELLED)
        # Money that never arrived isn't received. Same rule as the ledger, so
        # the two reports can't tell different stories.
        .exclude(cheques__status__in=[Cheque.Status.BOUNCED, Cheque.Status.HELD])
        .values("bill__customer")
        .annotate(total=Sum("amount"))
        .values("total")
    )
    last_bill = live_bills.order_by("-bill_date").values("bill_date")[:1]
    last_payment = (
        Payment.objects.filter(bill__customer=OuterRef("pk"))
        .exclude(bill__status=Bill.Status.CANCELLED)
        .order_by("-paid_at")
        .values("paid_at")[:1]
    )

    customers = (
        Customer.objects.annotate(
            owed=Case(
                When(balance__lt=0, then=Value(0) - F("balance")),
                default=Value(ZERO),
                output_field=MONEY,
            )
        )
        .annotate(
            available_credit=Greatest(
                F("credit_limit") - F("owed"), Value(ZERO), output_field=MONEY
            ),
            total_billed=Coalesce(Subquery(billed), ZERO, output_field=MONEY),
            total_received=Coalesce(Subquery(received), ZERO, output_field=MONEY),
            last_bill_date=Subquery(last_bill),
            last_payment_at=Subquery(last_payment),
        )
    )

    if scope == "owing":
        customers = customers.filter(balance__lt=0)

    # Worst debt first. Name breaks the ties so the order never wobbles.
    customers = list(customers.order_by("-owed", "name"))

    for customer in customers:
        # The last time anything happened on the account, whichever side it was.
        dates = [
            stamp
            for stamp in (
                customer.last_bill_date,
                timezone.localtime(customer.last_payment_at).date()
                if customer.last_payment_at
                else None,
            )
            if stamp
        ]
        customer.last_transaction = max(dates) if dates else None

    return {
        "customers": customers,
        "scope": scope,
        "total_owed": sum((c.owed for c in customers), ZERO),
        "total_billed": sum((c.total_billed for c in customers), ZERO),
        "total_received": sum((c.total_received for c in customers), ZERO),
        "generated_at": timezone.localtime(),
    }


@login_required
def outstanding_report(request):
    context = _outstanding_context(request)
    context["query"] = request.GET.urlencode()

    # As with the sales report: paged on the page only, never in the builder
    # the PDF shares, and the totals stay struck over every customer in scope.
    page_obj = _paginate(request, context["customers"], settings.PAGINATE_BY_REPORTS)
    context["page_obj"] = page_obj
    context["customers"] = page_obj.object_list
    return render(request, "core/outstanding_report.html", context)


@login_required
def outstanding_report_pdf(request):
    stamp = timezone.localdate().isoformat()
    return _pdf_response(
        request,
        "core/outstanding_pdf.html",
        _outstanding_context(request),
        f"senovka-outstanding-{stamp}.pdf",
    )


# ------------------------------------------------------------- petty cash
# One PettyCashFund per calendar month, auto-created with the previous
# month's closing balance carried forward. Two kinds of movement:
# PettyCashEntry (expense out of the tin) and PettyCashReimbursement (top-up
# into the tin). The list page tabs between them and the fund's
# closing_balance is rewritten on every write via fund.recalculate().


def _petty_cash_context(request, active_tab="expenses"):
    """Everything the petty-cash page needs, whichever tab is showing.

    Called at page load and after every save so a POST error can re-render
    the same page state without a redirect losing the form.
    """
    month_filter = get_month_filter(request)
    # A specific month drives the fund; All time still needs *some* fund to
    # front the page (the "current month" one, so the balance card reads
    # sanely). The list below queries across all funds when in All Time.
    display_month = month_filter.month or timezone.localdate().replace(day=1)
    fund, carried_from = PettyCashFund.for_month(display_month)

    if month_filter.is_all_time:
        expense_qs = PettyCashEntry.objects.filter(
            entry_type=PettyCashEntry.EntryType.EXPENSE
        ).select_related("fund", "added_by").order_by("-date", "-id")
        reimb_qs = PettyCashReimbursement.objects.select_related(
            "fund", "added_by"
        ).order_by("-date", "-id")
    else:
        expense_qs = fund.entries.filter(
            entry_type=PettyCashEntry.EntryType.EXPENSE
        ).select_related("added_by").order_by("-date", "-id")
        reimb_qs = fund.reimbursements.select_related("added_by").order_by(
            "-date", "-id"
        )

    # Two paginators because the tabs are independent — ?page= applies to
    # whichever tab was clicked, so the two share the same page number and
    # only one tab is ever seen at once.
    expense_page = _paginate(request, expense_qs)
    reimb_page = _paginate(request, reimb_qs)

    return {
        "fund": fund,
        "carried_from": carried_from,
        "month_filter": month_filter,
        "expense_form": PettyCashExpenseForm(),
        "reimbursement_form": PettyCashReimbursementForm(),
        "categories": PettyCashEntry.Category.choices,
        "expense_page": expense_page,
        "expenses": expense_page.object_list,
        "reimbursement_page": reimb_page,
        "reimbursements": reimb_page.object_list,
        "active_tab": active_tab,
        "low_balance_threshold": Decimal("1000.00"),
    }


@login_required
def petty_cash(request):
    """The petty-cash page for one month (or all months when ?month=all).

    The fund for the requested month is auto-created with the previous
    month's closing balance if it does not exist yet — see
    PettyCashFund.for_month. If that happened just now, `carried_from`
    surfaces to the template so a notice can explain the seed balance.
    """
    return render(request, "core/petty_cash.html", _petty_cash_context(request))


def _petty_cash_redirect(request):
    """Where to land after a petty-cash write. Preserves ?month= so a
    correction made against last month does not send the operator back to
    the current month."""
    month = request.GET.get("month") or request.POST.get("month") or ""
    url = reverse("core:petty_cash")
    return redirect(f"{url}?month={month}" if month else url)


@require_POST
@login_required
def petty_cash_expense_create(request):
    form = PettyCashExpenseForm(request.POST)
    if not form.is_valid():
        messages.error(request, f"Expense not saved: {form.first_error()}")
        return _petty_cash_redirect(request)

    entry_date = form.cleaned_data["date"]
    with transaction.atomic():
        fund, _ = PettyCashFund.for_month(entry_date)
        entry = form.save(commit=False)
        entry.fund = fund
        entry.added_by = request.user
        entry.entry_type = PettyCashEntry.EntryType.EXPENSE
        entry.save()
        fund.recalculate()

    messages.success(
        request,
        f"Expense of {entry.amount:,.2f} recorded. "
        f"Available: {fund.closing_balance:,.2f}.",
    )
    return _petty_cash_redirect(request)


@require_POST
@login_required
def petty_cash_expense_edit(request, pk):
    entry = get_object_or_404(
        PettyCashEntry.objects.select_related("fund"),
        pk=pk,
        entry_type=PettyCashEntry.EntryType.EXPENSE,
    )
    form = PettyCashExpenseForm(request.POST, instance=entry, require_edit_reason=True)
    if not form.is_valid():
        messages.error(request, f"Edit not saved: {form.first_error()}")
        return _petty_cash_redirect(request)

    new_date = form.cleaned_data["date"]
    with transaction.atomic():
        old_fund = entry.fund
        edited = form.save(commit=False)
        edited.edit_date = timezone.localdate()

        # The date can move an entry between months. Repoint it to the new
        # month's fund and recalculate both funds so neither carries the
        # other's amount by accident.
        if new_date.replace(day=1) != old_fund.month:
            new_fund, _ = PettyCashFund.for_month(new_date)
            edited.fund = new_fund
            edited.save()
            old_fund.recalculate()
            new_fund.recalculate()
        else:
            edited.save()
            old_fund.recalculate()

    messages.success(request, "Expense updated.")
    return _petty_cash_redirect(request)


@require_POST
@login_required
def petty_cash_expense_delete(request, pk):
    entry = get_object_or_404(
        PettyCashEntry.objects.select_related("fund"),
        pk=pk,
        entry_type=PettyCashEntry.EntryType.EXPENSE,
    )
    with transaction.atomic():
        fund = entry.fund
        entry.delete()
        fund.recalculate()

    messages.success(request, "Expense removed.")
    return _petty_cash_redirect(request)


@require_POST
@login_required
def petty_cash_reimbursement_create(request):
    form = PettyCashReimbursementForm(request.POST)
    if not form.is_valid():
        messages.error(request, f"Reimbursement not saved: {form.first_error()}")
        return _petty_cash_redirect(request)

    entry_date = form.cleaned_data["date"]
    with transaction.atomic():
        fund, _ = PettyCashFund.for_month(entry_date)
        reimb = form.save(commit=False)
        reimb.fund = fund
        reimb.added_by = request.user
        reimb.save()
        fund.recalculate()

    messages.success(
        request,
        f"Reimbursement of {reimb.amount:,.2f} recorded. "
        f"Available: {fund.closing_balance:,.2f}.",
    )
    return _petty_cash_redirect(request)


@require_POST
@login_required
def petty_cash_reimbursement_edit(request, pk):
    reimb = get_object_or_404(
        PettyCashReimbursement.objects.select_related("fund"), pk=pk
    )
    form = PettyCashReimbursementForm(
        request.POST, instance=reimb, require_edit_reason=True
    )
    if not form.is_valid():
        messages.error(request, f"Edit not saved: {form.first_error()}")
        return _petty_cash_redirect(request)

    new_date = form.cleaned_data["date"]
    with transaction.atomic():
        old_fund = reimb.fund
        edited = form.save(commit=False)
        edited.edit_date = timezone.localdate()

        if new_date.replace(day=1) != old_fund.month:
            new_fund, _ = PettyCashFund.for_month(new_date)
            edited.fund = new_fund
            edited.save()
            old_fund.recalculate()
            new_fund.recalculate()
        else:
            edited.save()
            old_fund.recalculate()

    messages.success(request, "Reimbursement updated.")
    return _petty_cash_redirect(request)


@require_POST
@login_required
def petty_cash_reimbursement_delete(request, pk):
    reimb = get_object_or_404(
        PettyCashReimbursement.objects.select_related("fund"), pk=pk
    )
    with transaction.atomic():
        fund = reimb.fund
        reimb.delete()
        fund.recalculate()

    messages.success(request, "Reimbursement removed.")
    return _petty_cash_redirect(request)


@login_required
def petty_cash_excel(request):
    """Download the current fund's month (or all months) as an .xlsx.

    Honours the same ?month= filter the page uses, so the sheet matches
    exactly what the operator is looking at.
    """
    from io import BytesIO

    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    context = _petty_cash_context(request)
    fund = context["fund"]
    month_filter = context["month_filter"]

    if month_filter.is_all_time:
        expense_qs = PettyCashEntry.objects.filter(
            entry_type=PettyCashEntry.EntryType.EXPENSE
        ).select_related("fund", "added_by").order_by("-date", "-id")
        reimb_qs = PettyCashReimbursement.objects.select_related(
            "fund", "added_by"
        ).order_by("-date", "-id")
    else:
        expense_qs = fund.entries.filter(
            entry_type=PettyCashEntry.EntryType.EXPENSE
        ).select_related("added_by").order_by("-date", "-id")
        reimb_qs = fund.reimbursements.select_related("added_by").order_by(
            "-date", "-id"
        )

    wb = Workbook()
    ws = wb.active
    ws.title = "Petty Cash"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1F2937")
    section_font = Font(bold=True, size=12)
    right = Alignment(horizontal="right")
    center = Alignment(horizontal="center")

    # Masthead
    ws["A1"] = "Senovka Plastics — Petty Cash"
    ws["A1"].font = Font(bold=True, size=14)
    ws.merge_cells("A1:F1")

    ws["A2"] = "Period"
    ws["A2"].font = Font(bold=True)
    ws["B2"] = month_filter.label
    ws["D2"] = "Generated"
    ws["D2"].font = Font(bold=True)
    ws["E2"] = timezone.localtime().strftime("%d %b %Y %H:%M")

    # Summary block (only meaningful for a single month)
    row = 4
    if not month_filter.is_all_time:
        ws.cell(row=row, column=1, value="Opening balance").font = Font(bold=True)
        c = ws.cell(row=row, column=2, value=float(fund.opening_balance))
        c.number_format = "#,##0.00"; c.alignment = right
        row += 1
        ws.cell(row=row, column=1, value="Total reimbursements").font = Font(bold=True)
        c = ws.cell(row=row, column=2, value=float(fund.total_reimbursements))
        c.number_format = "#,##0.00"; c.alignment = right
        row += 1
        ws.cell(row=row, column=1, value="Total expenses").font = Font(bold=True)
        c = ws.cell(row=row, column=2, value=float(fund.total_expenses))
        c.number_format = "#,##0.00"; c.alignment = right
        row += 1
        ws.cell(row=row, column=1, value="Closing balance").font = Font(bold=True)
        c = ws.cell(row=row, column=2, value=float(fund.closing_balance))
        c.number_format = "#,##0.00"; c.alignment = right; c.font = Font(bold=True)
        row += 2

    # Expenses table
    ws.cell(row=row, column=1, value="Expenses").font = section_font
    row += 1
    EXP_HEADERS = ["No", "Date", "Category", "Description", "Receipt No", "Amount"]
    for idx, name in enumerate(EXP_HEADERS, start=1):
        cell = ws.cell(row=row, column=idx, value=name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
    row += 1

    expense_total = Decimal("0.00")
    for i, entry in enumerate(expense_qs, start=1):
        ws.cell(row=row, column=1, value=i).alignment = center
        ws.cell(row=row, column=2, value=entry.date.strftime("%Y-%m-%d"))
        ws.cell(row=row, column=3, value=entry.get_category_display())
        ws.cell(row=row, column=4, value=entry.description)
        ws.cell(row=row, column=5, value=entry.receipt_no or "")
        c = ws.cell(row=row, column=6, value=float(entry.amount))
        c.number_format = "#,##0.00"; c.alignment = right
        expense_total += entry.amount
        row += 1

    ws.cell(row=row, column=5, value="Total").font = Font(bold=True)
    ws.cell(row=row, column=5).alignment = right
    c = ws.cell(row=row, column=6, value=float(expense_total))
    c.number_format = "#,##0.00"; c.alignment = right; c.font = Font(bold=True)
    row += 2

    # Reimbursements table
    ws.cell(row=row, column=1, value="Reimbursements").font = section_font
    row += 1
    REIMB_HEADERS = ["No", "Date", "Given By", "Reason", "", "Amount"]
    for idx, name in enumerate(REIMB_HEADERS, start=1):
        cell = ws.cell(row=row, column=idx, value=name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
    row += 1

    reimb_total = Decimal("0.00")
    for i, reimb in enumerate(reimb_qs, start=1):
        ws.cell(row=row, column=1, value=i).alignment = center
        ws.cell(row=row, column=2, value=reimb.date.strftime("%Y-%m-%d"))
        ws.cell(row=row, column=3, value=reimb.given_by)
        ws.cell(row=row, column=4, value=reimb.reason)
        c = ws.cell(row=row, column=6, value=float(reimb.amount))
        c.number_format = "#,##0.00"; c.alignment = right
        reimb_total += reimb.amount
        row += 1

    ws.cell(row=row, column=5, value="Total").font = Font(bold=True)
    ws.cell(row=row, column=5).alignment = right
    c = ws.cell(row=row, column=6, value=float(reimb_total))
    c.number_format = "#,##0.00"; c.alignment = right; c.font = Font(bold=True)

    widths = {"A": 5, "B": 12, "C": 16, "D": 40, "E": 16, "F": 14}
    for letter, width in widths.items():
        ws.column_dimensions[letter].width = width

    if month_filter.is_all_time:
        stamp = "all-months"
    else:
        stamp = fund.month.strftime("%Y-%m")

    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    response = HttpResponse(
        stream.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = (
        f'attachment; filename="petty_cash_{stamp}.xlsx"'
    )
    return response


@login_required
def petty_cash_pdf(request):
    """Print the current fund's month as a PDF summary + full lists."""
    context = _petty_cash_context(request)
    fund = context["fund"]
    # The PDF is a complete statement of the month — no pagination there,
    # so it needs the full lists rather than a page of them.
    expense_qs = fund.entries.filter(
        entry_type=PettyCashEntry.EntryType.EXPENSE
    ).select_related("added_by").order_by("-date", "-id")
    reimb_qs = fund.reimbursements.select_related("added_by").order_by(
        "-date", "-id"
    )
    context.update({
        "expenses": expense_qs,
        "reimbursements": reimb_qs,
        "generated_at": timezone.localtime(),
    })
    stamp = fund.month.strftime("%Y-%m")
    return _pdf_response(
        request,
        "core/petty_cash_pdf.html",
        context,
        f"senovka-petty-cash-{stamp}.pdf",
    )


# =========================================================== material master
# Suppliers we buy raw material from and the raw materials themselves. Both
# CRUD flows are modal-based on the list page, POST-only edit/delete, and
# super-admin only — a manager can view but not change master data. See
# MaterialSupplier / Material for why these are separate from Customer and
# Product.


@super_admin_required
def material_supplier_list(request):
    suppliers = (
        MaterialSupplier.objects.annotate(
            purchase_count=Count("purchases", distinct=True)
        ).order_by("name")
    )
    page_obj = _paginate(request, suppliers)
    return render(
        request,
        "core/material_supplier_list.html",
        {
            "page_obj": page_obj,
            "suppliers": page_obj.object_list,
            "form": MaterialSupplierForm(),
        },
    )


@require_POST
@super_admin_required
def material_supplier_create(request):
    form = MaterialSupplierForm(request.POST)
    if not form.is_valid():
        messages.error(request, f"Supplier not saved: {form.first_error()}")
    else:
        with transaction.atomic():
            supplier = form.save()
        messages.success(request, f"Supplier '{supplier.name}' created.")
    return redirect("core:material_supplier_list")


@require_POST
@super_admin_required
def material_supplier_edit(request, pk):
    supplier = get_object_or_404(MaterialSupplier, pk=pk)
    form = MaterialSupplierForm(request.POST, instance=supplier)
    if not form.is_valid():
        messages.error(request, f"Supplier not saved: {form.first_error()}")
    else:
        with transaction.atomic():
            form.save()
        messages.success(request, "Supplier updated.")
    return redirect("core:material_supplier_list")


@require_POST
@super_admin_required
def material_supplier_delete(request, pk):
    supplier = get_object_or_404(
        MaterialSupplier.objects.annotate(
            purchase_count=Count("purchases", distinct=True)
        ),
        pk=pk,
    )
    if supplier.purchase_count:
        # PROTECT on the FK would raise anyway; say so first.
        messages.error(
            request,
            f"Cannot delete '{supplier.name}' — {supplier.purchase_count} "
            f"purchase{'' if supplier.purchase_count == 1 else 's'} still "
            f"reference it. Deactivate instead.",
        )
        return redirect("core:material_supplier_list")
    try:
        supplier.delete()
        messages.success(request, f"Supplier '{supplier.name}' deleted.")
    except ProtectedError:
        messages.error(
            request,
            f"Cannot delete '{supplier.name}' — other records reference it.",
        )
    return redirect("core:material_supplier_list")


@super_admin_required
def material_list(request):
    materials = (
        Material.objects.annotate(
            purchase_count=Count("purchase_items", distinct=True)
        ).order_by("name")
    )
    page_obj = _paginate(request, materials)
    return render(
        request,
        "core/material_list.html",
        {
            "page_obj": page_obj,
            "materials": page_obj.object_list,
            "form": MaterialForm(),
            "units": Material.Unit.choices,
        },
    )


@require_POST
@super_admin_required
def material_create(request):
    form = MaterialForm(request.POST)
    if not form.is_valid():
        messages.error(request, f"Material not saved: {form.first_error()}")
    else:
        with transaction.atomic():
            material = form.save()
        messages.success(request, f"Material '{material.name}' created.")
    return redirect("core:material_list")


@require_POST
@super_admin_required
def material_edit(request, pk):
    material = get_object_or_404(Material, pk=pk)
    form = MaterialForm(request.POST, instance=material)
    if not form.is_valid():
        messages.error(request, f"Material not saved: {form.first_error()}")
    else:
        with transaction.atomic():
            form.save()
        messages.success(request, "Material updated.")
    return redirect("core:material_list")


@require_POST
@super_admin_required
def material_delete(request, pk):
    material = get_object_or_404(
        Material.objects.annotate(
            purchase_count=Count("purchase_items", distinct=True)
        ),
        pk=pk,
    )
    if material.purchase_count:
        messages.error(
            request,
            f"Cannot delete '{material.name}' — {material.purchase_count} "
            f"purchase line{'' if material.purchase_count == 1 else 's'} "
            f"still reference it. Deactivate instead.",
        )
        return redirect("core:material_list")
    try:
        material.delete()
        messages.success(request, f"Material '{material.name}' deleted.")
    except ProtectedError:
        messages.error(
            request,
            f"Cannot delete '{material.name}' — other records reference it.",
        )
    return redirect("core:material_list")


# ======================================================== material purchases
# The main flow. A purchase collects several MaterialPurchaseItems (ordered
# quantities), which are then weighed in over one or more visits to the
# scale — each visit is a MaterialWeighEntry, and MaterialPurchaseItem
# caches the running weighed_qty via recalculate_weighed(). MaterialPurchase
# caches status via refresh_status(). All model helpers.


def _parse_purchase_items(raw_json):
    """Read the items JSON off a purchase POST and return validated dicts.

    Refuses malformed JSON, empty lists, unknown materials, and negative
    numbers. Raises BillError (reused: the message pipeline is the same).
    """
    try:
        rows = json.loads(raw_json or "[]")
    except (ValueError, TypeError):
        raise BillError("Item list is malformed.")
    if not isinstance(rows, list) or not rows:
        raise BillError("Add at least one item.")

    material_ids = [
        r.get("material_id") for r in rows if isinstance(r, dict)
    ]
    materials = {
        m.pk: m for m in Material.objects.filter(pk__in=material_ids)
    }

    seen = set()
    items = []
    for index, row in enumerate(rows, start=1):
        if not isinstance(row, dict):
            raise BillError(f"Item {index}: malformed row.")
        material = materials.get(row.get("material_id"))
        if material is None:
            raise BillError(f"Item {index}: material no longer exists.")
        if material.pk in seen:
            raise BillError(f"{material.name} is on this purchase twice.")
        seen.add(material.pk)

        ordered = _decimal(row.get("ordered_qty"), f"Item {index} qty", 3)
        if ordered <= ZERO:
            raise BillError(f"Item {index}: quantity must be above 0.")
        unit_price = _decimal(row.get("unit_price"), f"Item {index} price", 2)
        line_total = (ordered * unit_price).quantize(
            Decimal("0.01"), rounding=ROUND_HALF_UP
        )
        items.append(
            {
                "material": material,
                "ordered_qty": ordered,
                "unit_price": unit_price,
                "line_total": line_total,
            }
        )

    return items


@login_required
def material_purchase_list(request):
    """List of purchases, filtered by status and month."""
    status = request.GET.get("status", "").strip()
    valid_status = {v for v, _ in MaterialPurchase.Status.choices}
    if status not in valid_status:
        status = ""

    month_filter = get_month_filter(request)

    purchases = (
        MaterialPurchase.objects.select_related("supplier", "created_by")
        .annotate(item_count=Count("items", distinct=True))
        .order_by("-purchase_date", "-id")
    )
    if status:
        purchases = purchases.filter(status=status)
    purchases = month_filter.apply(purchases, field="purchase_date")

    page_obj = _paginate(request, purchases)

    return render(
        request,
        "core/material_purchase_list.html",
        {
            "page_obj": page_obj,
            "purchases": page_obj.object_list,
            "status": status,
            "statuses": MaterialPurchase.Status.choices,
            "month_filter": month_filter,
            "is_filtered": bool(status or not month_filter.is_all_time),
        },
    )


def _material_form_context(request, form, items=None, is_edit=False, purchase=None):
    return {
        "form": form,
        "materials": list(
            Material.objects.filter(is_active=True).order_by("name")
        ),
        # For a live "add item" grid on the create page.
        "materials_json": json.dumps([
            {
                "id": m.pk,
                "name": m.name,
                "unit": m.get_unit_display(),
                "unit_price": f"{m.default_unit_price:.2f}",
            }
            for m in Material.objects.filter(is_active=True).order_by("name")
        ]),
        "initial_items_json": json.dumps(items or []),
        "is_edit": is_edit,
        "purchase": purchase,
    }


@login_required
def material_purchase_create(request):
    """GET renders the form, POST saves the purchase and its items."""
    form = MaterialPurchaseHeaderForm(request.POST or None)

    if request.method == "POST":
        try:
            with transaction.atomic():
                if not form.is_valid():
                    raise BillError(form.first_error())
                items = _parse_purchase_items(request.POST.get("items_json"))

                purchase = form.save(commit=False)
                purchase.created_by = request.user
                purchase.save()
                for item in items:
                    MaterialPurchaseItem.objects.create(
                        purchase=purchase,
                        material=item["material"],
                        ordered_qty=item["ordered_qty"],
                        unit_price=item["unit_price"],
                        line_total=item["line_total"],
                    )
                purchase.refresh_status()
        except BillError as exc:
            messages.error(request, f"Purchase not saved: {exc}")
        else:
            messages.success(
                request,
                f"Purchase #{purchase.pk} created ({len(items)} item"
                f"{'' if len(items) == 1 else 's'}, "
                f"{purchase.total_amount:,.2f}).",
            )
            return redirect("core:material_purchase_detail", pk=purchase.pk)

    return render(
        request,
        "core/material_purchase_create.html",
        _material_form_context(request, form),
    )


@login_required
def material_purchase_edit(request, pk):
    purchase = get_object_or_404(MaterialPurchase, pk=pk)

    if request.method == "POST":
        form = MaterialPurchaseHeaderForm(
            request.POST, instance=purchase, require_edit_reason=True
        )
        try:
            with transaction.atomic():
                if not form.is_valid():
                    raise BillError(form.first_error())
                items = _parse_purchase_items(request.POST.get("items_json"))

                # A rewrite: drop the old items (and their weigh entries via
                # CASCADE) and re-create from the new list. Simpler than
                # trying to diff — a purchase is small and this cannot leave
                # a half-updated set of rows.
                purchase.items.all().delete()
                edited = form.save(commit=False)
                edited.edit_date = timezone.localdate()
                edited.save()
                for item in items:
                    MaterialPurchaseItem.objects.create(
                        purchase=edited,
                        material=item["material"],
                        ordered_qty=item["ordered_qty"],
                        unit_price=item["unit_price"],
                        line_total=item["line_total"],
                    )
                edited.refresh_status()
        except BillError as exc:
            messages.error(request, f"Purchase not saved: {exc}")
        else:
            messages.success(request, f"Purchase #{purchase.pk} updated.")
            return redirect("core:material_purchase_detail", pk=purchase.pk)
    else:
        form = MaterialPurchaseHeaderForm(instance=purchase, require_edit_reason=True)

    initial_items = [
        {
            "material_id": item.material_id,
            "material_name": item.material.name,
            "unit": item.material.get_unit_display(),
            "ordered_qty": f"{item.ordered_qty:.3f}",
            "unit_price": f"{item.unit_price:.2f}",
        }
        for item in purchase.items.select_related("material")
    ]

    return render(
        request,
        "core/material_purchase_create.html",
        _material_form_context(
            request, form, items=initial_items, is_edit=True, purchase=purchase
        ),
    )


@require_POST
@super_admin_required
def material_purchase_delete(request, pk):
    purchase = get_object_or_404(MaterialPurchase, pk=pk)
    label = f"Purchase #{purchase.pk} · {purchase.supplier.name}"
    with transaction.atomic():
        # CASCADE takes items and weigh entries with it.
        purchase.delete()
    messages.success(request, f"{label} deleted.")
    return redirect("core:material_purchase_list")


@login_required
def material_purchase_detail(request, pk):
    purchase = get_object_or_404(
        MaterialPurchase.objects.select_related("supplier", "created_by"),
        pk=pk,
    )
    items = (
        purchase.items.select_related("material")
        .prefetch_related("weigh_entries__submitted_by")
    )
    return render(
        request,
        "core/material_purchase_detail.html",
        {
            "purchase": purchase,
            "items": items,
            "weigh_form": MaterialWeighEntryForm(),
        },
    )


@require_POST
@login_required
def material_purchase_weigh_add(request, item_pk):
    item = get_object_or_404(
        MaterialPurchaseItem.objects.select_related("purchase"), pk=item_pk
    )
    form = MaterialWeighEntryForm(request.POST)
    if not form.is_valid():
        messages.error(request, f"Weigh entry not saved: {form.first_error()}")
        return redirect("core:material_purchase_detail", pk=item.purchase_id)

    with transaction.atomic():
        entry = form.save(commit=False)
        entry.purchase_item = item
        entry.submitted_by = request.user
        entry.save()
        item.recalculate_weighed()
        item.purchase.refresh_status()

    messages.success(
        request,
        f"Weighed {entry.weighed_qty} on {entry.weigh_date:%d %b %Y} · "
        f"{item.weighed_qty}/{item.ordered_qty} done.",
    )
    return redirect("core:material_purchase_detail", pk=item.purchase_id)


@require_POST
@login_required
def material_purchase_weigh_edit(request, pk):
    entry = get_object_or_404(
        MaterialWeighEntry.objects.select_related("purchase_item__purchase"),
        pk=pk,
    )
    form = MaterialWeighEntryForm(request.POST, instance=entry)
    if not form.is_valid():
        messages.error(request, f"Weigh entry not saved: {form.first_error()}")
        return redirect(
            "core:material_purchase_detail", pk=entry.purchase_item.purchase_id
        )

    with transaction.atomic():
        edited = form.save()
        edited.purchase_item.recalculate_weighed()
        edited.purchase_item.purchase.refresh_status()

    messages.success(request, "Weigh entry updated.")
    return redirect(
        "core:material_purchase_detail", pk=entry.purchase_item.purchase_id
    )


@require_POST
@login_required
def material_purchase_weigh_delete(request, pk):
    entry = get_object_or_404(
        MaterialWeighEntry.objects.select_related("purchase_item__purchase"),
        pk=pk,
    )
    with transaction.atomic():
        item = entry.purchase_item
        entry.delete()
        item.recalculate_weighed()
        item.purchase.refresh_status()

    messages.success(request, "Weigh entry removed.")
    return redirect("core:material_purchase_detail", pk=item.purchase_id)


# =========================================================== vehicle tracker
# Vehicles, riders, and a log of trips between them. No stock and no money —
# a trip is a leg with its own km reading (not an odometer), so month totals
# are a sum of legs. See models Vehicle / Rider / VehicleTrip.


def _month_km_for(qs, month_filter, group_field, name_field):
    """A small aggregation: total km and trip count per {vehicle,rider} in
    the given month. Feeds the two summary cards on the trip page."""
    scoped = month_filter.apply(qs, field="trip_date")
    return (
        scoped.values(group_field, name_field)
        .annotate(trips=Count("id"), total_km=Coalesce(Sum("km"), ZERO, output_field=MONEY))
        .order_by("-total_km", name_field)
    )


# ---- Vehicles (super-admin CRUD) ----


@super_admin_required
def vehicle_list(request):
    month_filter = get_month_filter(request)
    trips_this_month = month_filter.apply(
        VehicleTrip.objects.all(), field="trip_date"
    )
    km_by_vehicle = dict(
        trips_this_month.values("vehicle_id")
        .annotate(total=Coalesce(Sum("km"), ZERO, output_field=MONEY))
        .values_list("vehicle_id", "total")
    )
    vehicles = Vehicle.objects.order_by("name")
    for v in vehicles:
        v.km_this_month = km_by_vehicle.get(v.pk, ZERO)
    page_obj = _paginate(request, vehicles)
    return render(
        request,
        "core/vehicle_list.html",
        {
            "page_obj": page_obj,
            "vehicles": page_obj.object_list,
            "form": VehicleForm(),
            "month_filter": month_filter,
        },
    )


@require_POST
@super_admin_required
def vehicle_create(request):
    form = VehicleForm(request.POST)
    if not form.is_valid():
        messages.error(request, f"Vehicle not saved: {form.first_error()}")
    else:
        with transaction.atomic():
            vehicle = form.save()
        messages.success(request, f"Vehicle '{vehicle.name}' created.")
    return redirect("core:vehicle_list")


@require_POST
@super_admin_required
def vehicle_edit(request, pk):
    vehicle = get_object_or_404(Vehicle, pk=pk)
    form = VehicleForm(request.POST, instance=vehicle)
    if not form.is_valid():
        messages.error(request, f"Vehicle not saved: {form.first_error()}")
    else:
        with transaction.atomic():
            form.save()
        messages.success(request, "Vehicle updated.")
    return redirect("core:vehicle_list")


@require_POST
@super_admin_required
def vehicle_delete(request, pk):
    vehicle = get_object_or_404(
        Vehicle.objects.annotate(trip_count=Count("trips", distinct=True)), pk=pk
    )
    if vehicle.trip_count:
        messages.error(
            request,
            f"Cannot delete '{vehicle.name}' — {vehicle.trip_count} "
            f"trip{'' if vehicle.trip_count == 1 else 's'} still reference it. "
            f"Deactivate instead.",
        )
        return redirect("core:vehicle_list")
    try:
        vehicle.delete()
        messages.success(request, f"Vehicle '{vehicle.name}' deleted.")
    except ProtectedError:
        messages.error(request, f"Cannot delete '{vehicle.name}' — other records reference it.")
    return redirect("core:vehicle_list")


# ---- Riders (super-admin CRUD) ----


@super_admin_required
def rider_list(request):
    month_filter = get_month_filter(request)
    trips_this_month = month_filter.apply(VehicleTrip.objects.all(), field="trip_date")
    km_by_rider = dict(
        trips_this_month.values("rider_id")
        .annotate(total=Coalesce(Sum("km"), ZERO, output_field=MONEY))
        .values_list("rider_id", "total")
    )
    riders = Rider.objects.order_by("name")
    for r in riders:
        r.km_this_month = km_by_rider.get(r.pk, ZERO)
    page_obj = _paginate(request, riders)
    return render(
        request,
        "core/rider_list.html",
        {
            "page_obj": page_obj,
            "riders": page_obj.object_list,
            "form": RiderForm(),
            "month_filter": month_filter,
        },
    )


@require_POST
@super_admin_required
def rider_create(request):
    form = RiderForm(request.POST)
    if not form.is_valid():
        messages.error(request, f"Rider not saved: {form.first_error()}")
    else:
        with transaction.atomic():
            rider = form.save()
        messages.success(request, f"Rider '{rider.name}' created.")
    return redirect("core:rider_list")


@require_POST
@super_admin_required
def rider_edit(request, pk):
    rider = get_object_or_404(Rider, pk=pk)
    form = RiderForm(request.POST, instance=rider)
    if not form.is_valid():
        messages.error(request, f"Rider not saved: {form.first_error()}")
    else:
        with transaction.atomic():
            form.save()
        messages.success(request, "Rider updated.")
    return redirect("core:rider_list")


@require_POST
@super_admin_required
def rider_delete(request, pk):
    rider = get_object_or_404(
        Rider.objects.annotate(trip_count=Count("trips", distinct=True)), pk=pk
    )
    if rider.trip_count:
        messages.error(
            request,
            f"Cannot delete '{rider.name}' — {rider.trip_count} "
            f"trip{'' if rider.trip_count == 1 else 's'} still reference it. "
            f"Deactivate instead.",
        )
        return redirect("core:rider_list")
    try:
        rider.delete()
        messages.success(request, f"Rider '{rider.name}' deleted.")
    except ProtectedError:
        messages.error(request, f"Cannot delete '{rider.name}' — other records reference it.")
    return redirect("core:rider_list")


# ---- Trips ----


def _vehicle_trip_context(request):
    """Everything the trip page needs, whichever the caller. Reused by GET
    and (on failure) the POST fall-back."""
    month_filter = get_month_filter(request)

    vehicle_id = request.GET.get("vehicle", "").strip()
    rider_id = request.GET.get("rider", "").strip()

    trips = (
        VehicleTrip.objects.select_related("vehicle", "rider", "added_by")
        .order_by("-trip_date", "-id")
    )
    trips = month_filter.apply(trips, field="trip_date")
    if vehicle_id.isdigit():
        trips = trips.filter(vehicle_id=int(vehicle_id))
    if rider_id.isdigit():
        trips = trips.filter(rider_id=int(rider_id))

    page_obj = _paginate(request, trips)

    # Summary aggregates over the same month filter.
    scoped = month_filter.apply(VehicleTrip.objects.all(), field="trip_date")
    totals = scoped.aggregate(
        total_km=Coalesce(Sum("km"), ZERO, output_field=MONEY),
        total_trips=Count("id"),
    )

    by_vehicle = _month_km_for(
        VehicleTrip.objects.all(), month_filter, "vehicle_id", "vehicle__name"
    )
    by_rider = _month_km_for(
        VehicleTrip.objects.all(), month_filter, "rider_id", "rider__name"
    )
    most_vehicle = next(iter(by_vehicle), None)
    most_rider = next(iter(by_rider), None)

    return {
        "page_obj": page_obj,
        "trips": page_obj.object_list,
        "form": VehicleTripForm(),
        "month_filter": month_filter,
        "vehicles": Vehicle.objects.filter(is_active=True).order_by("name"),
        "riders": Rider.objects.filter(is_active=True).order_by("name"),
        "selected_vehicle": vehicle_id if vehicle_id.isdigit() else "",
        "selected_rider": rider_id if rider_id.isdigit() else "",
        "is_filtered": bool(vehicle_id or rider_id or not month_filter.is_all_time),
        "total_km": totals["total_km"],
        "total_trips": totals["total_trips"],
        "most_vehicle": most_vehicle,
        "most_rider": most_rider,
        "by_vehicle": by_vehicle,
        "by_rider": by_rider,
    }


@login_required
def vehicle_trip_list(request):
    return render(request, "core/vehicle_trip_list.html", _vehicle_trip_context(request))


def _vehicle_trip_redirect(request):
    """Preserve month + filter args after a write."""
    params = []
    for key in ("month", "vehicle", "rider"):
        value = request.GET.get(key) or request.POST.get(key)
        if value:
            params.append(f"{key}={value}")
    url = reverse("core:vehicle_trip_list")
    return redirect(f"{url}?{'&'.join(params)}" if params else url)


@require_POST
@login_required
def vehicle_trip_create(request):
    form = VehicleTripForm(request.POST)
    if not form.is_valid():
        messages.error(request, f"Trip not saved: {form.first_error()}")
        return _vehicle_trip_redirect(request)
    with transaction.atomic():
        trip = form.save(commit=False)
        trip.added_by = request.user
        trip.save()
    messages.success(request, f"Trip logged: {trip.km:g}km on {trip.trip_date:%d %b %Y}.")
    return _vehicle_trip_redirect(request)


@require_POST
@login_required
def vehicle_trip_edit(request, pk):
    trip = get_object_or_404(VehicleTrip, pk=pk)
    form = VehicleTripForm(request.POST, instance=trip)
    if not form.is_valid():
        messages.error(request, f"Trip not saved: {form.first_error()}")
        return _vehicle_trip_redirect(request)
    with transaction.atomic():
        form.save()
    messages.success(request, "Trip updated.")
    return _vehicle_trip_redirect(request)


@require_POST
@login_required
def vehicle_trip_delete(request, pk):
    trip = get_object_or_404(VehicleTrip, pk=pk)
    with transaction.atomic():
        trip.delete()
    messages.success(request, "Trip removed.")
    return _vehicle_trip_redirect(request)


@login_required
def vehicle_trip_pdf(request):
    context = _vehicle_trip_context(request)
    # The PDF is a full monthly report — replace the paged slice with the
    # complete filtered set so nothing is cut off at page 2.
    month_filter = context["month_filter"]
    trips = (
        month_filter.apply(
            VehicleTrip.objects.select_related("vehicle", "rider", "added_by"),
            field="trip_date",
        )
        .order_by("-trip_date", "-id")
    )
    if context["selected_vehicle"]:
        trips = trips.filter(vehicle_id=int(context["selected_vehicle"]))
    if context["selected_rider"]:
        trips = trips.filter(rider_id=int(context["selected_rider"]))
    context["trips"] = list(trips)
    context["generated_at"] = timezone.localtime()
    stamp = (
        month_filter.month.strftime("%Y-%m")
        if not month_filter.is_all_time
        else "all-time"
    )
    return _pdf_response(
        request,
        "core/vehicle_trip_pdf.html",
        context,
        f"senovka-vehicle-trips-{stamp}.pdf",
    )


# ================================================================ order book
# Quotations. Nothing here moves stock, balance or money — see Order for why
# these are their own model rather than a Bill with a status. Reference
# numbers come from ReferenceCounter, which survives deletion; see the model.


def _order_line_price(customer_id, product):
    """The price to quote for `product` when writing an OrderItem for
    `customer_id`. Mirrors the bill-create path: a CustomerPrice override
    wins over Product.default_price; a walk-in (no customer_id) gets the
    default."""
    if customer_id:
        override = CustomerPrice.objects.filter(
            customer_id=customer_id, product=product
        ).values_list("unit_price", flat=True).first()
        if override is not None:
            return override
    return product.default_price


def _parse_order_items(raw_json, customer_id):
    """Read the items JSON off an order POST and return validated dicts.

    Refuses malformed JSON, empty lists, unknown or inactive products,
    duplicated products, and negative numbers. Prices default to the
    customer's own quote if the JSON omitted one — the operator does not
    have to re-type what the AJAX endpoint pre-filled.
    """
    try:
        rows = json.loads(raw_json or "[]")
    except (ValueError, TypeError):
        raise BillError("Item list is malformed.")
    if not isinstance(rows, list) or not rows:
        raise BillError("Add at least one item.")

    ids = [r.get("product_id") for r in rows if isinstance(r, dict)]
    products = {p.pk: p for p in Product.objects.filter(pk__in=ids, is_active=True)}

    seen = set()
    items = []
    for index, row in enumerate(rows, start=1):
        if not isinstance(row, dict):
            raise BillError(f"Item {index}: malformed row.")
        product = products.get(row.get("product_id"))
        if product is None:
            raise BillError(f"Item {index}: product no longer available.")
        if product.pk in seen:
            raise BillError(f"{product} is on this order twice.")
        seen.add(product.pk)

        qty = _decimal(row.get("qty"), f"Item {index} qty", 3)
        if qty <= ZERO:
            raise BillError(f"Item {index}: quantity must be above 0.")

        raw_price = row.get("unit_price")
        if raw_price is None or str(raw_price).strip() == "":
            unit_price = _order_line_price(customer_id, product)
        else:
            unit_price = _decimal(raw_price, f"Item {index} price", 2)
        line_total = (qty * unit_price).quantize(
            Decimal("0.01"), rounding=ROUND_HALF_UP
        )
        items.append(
            {
                "product": product,
                "qty": qty,
                "unit_price": unit_price,
                "line_total": line_total,
            }
        )
    return items


@login_required
def order_list(request):
    status = request.GET.get("status", "").strip()
    valid_status = {v for v, _ in Order.Status.choices}
    if status not in valid_status:
        status = ""

    month_filter = get_month_filter(request)

    orders = (
        Order.objects.select_related("customer", "created_by")
        .annotate(item_count=Count("items", distinct=True))
        .order_by("-order_date", "-id")
    )
    if status:
        orders = orders.filter(status=status)
    orders = month_filter.apply(orders, field="order_date")

    page_obj = _paginate(request, orders)

    return render(
        request,
        "core/order_list.html",
        {
            "page_obj": page_obj,
            "orders": page_obj.object_list,
            "status": status,
            "statuses": Order.Status.choices,
            "month_filter": month_filter,
            "is_filtered": bool(status or not month_filter.is_all_time),
        },
    )


def _order_form_context(request, form, items=None, is_edit=False, order=None):
    """What both order form pages need. Includes annotated customers so the
    picker shows the running balance alongside each name."""
    return {
        "form": form,
        "customers": _billable_customers(),
        "categories": Category.objects.all(),
        "initial_items_json": json.dumps(items or []),
        "is_edit": is_edit,
        "order": order,
        # The endpoint the JS calls with the picked customer id — reused
        # verbatim from bill creation.
        "products_url_template": reverse(
            "core:bill_products", kwargs={"customer_id": 999999999}
        ),
        "walk_in_customer_id": _walk_in_customer().pk,
        "today": timezone.localdate(),
    }


@login_required
def order_create(request):
    form = OrderHeaderForm(request.POST or None)

    if request.method == "POST":
        try:
            with transaction.atomic():
                if not form.is_valid():
                    raise BillError(form.first_error())
                cust = form.cleaned_data.get("customer")
                items = _parse_order_items(
                    request.POST.get("items_json"),
                    cust.pk if cust else None,
                )

                order = form.save(commit=False)
                order.created_by = request.user
                order.save()  # save() assigns reference_no from ReferenceCounter.
                for item in items:
                    OrderItem.objects.create(
                        order=order,
                        product=item["product"],
                        qty=item["qty"],
                        unit_price=item["unit_price"],
                        line_total=item["line_total"],
                    )
                order.recalculate()
        except BillError as exc:
            messages.error(request, f"Quotation not saved: {exc}")
        else:
            messages.success(
                request,
                f"Quotation {order.reference_no} created ({len(items)} item"
                f"{'' if len(items) == 1 else 's'}, "
                f"{order.total_amount:,.2f}).",
            )
            return redirect("core:order_detail", pk=order.pk)

    return render(
        request,
        "core/order_create.html",
        _order_form_context(request, form),
    )


@login_required
def order_edit(request, pk):
    order = get_object_or_404(Order, pk=pk)

    if request.method == "POST":
        # A confirmed quotation being edited is a serious enough change to
        # want a reason on record — same as edit_reason on bills.
        was_confirmed = order.status == Order.Status.CONFIRMED
        form = OrderHeaderForm(
            request.POST, instance=order, require_edit_reason=was_confirmed
        )
        try:
            with transaction.atomic():
                if not form.is_valid():
                    raise BillError(form.first_error())
                cust = form.cleaned_data.get("customer")
                items = _parse_order_items(
                    request.POST.get("items_json"),
                    cust.pk if cust else None,
                )

                order.items.all().delete()
                edited = form.save(commit=False)
                if was_confirmed:
                    edited.edit_date = timezone.localdate()
                edited.save()
                for item in items:
                    OrderItem.objects.create(
                        order=edited,
                        product=item["product"],
                        qty=item["qty"],
                        unit_price=item["unit_price"],
                        line_total=item["line_total"],
                    )
                edited.recalculate()
        except BillError as exc:
            messages.error(request, f"Quotation not saved: {exc}")
        else:
            messages.success(request, f"Quotation {order.reference_no} updated.")
            return redirect("core:order_detail", pk=order.pk)
    else:
        was_confirmed = order.status == Order.Status.CONFIRMED
        form = OrderHeaderForm(
            instance=order, require_edit_reason=was_confirmed,
            initial={
                "customer_name": order.customer_name,
            },
        )

    initial_items = [
        {
            "product_id": item.product_id,
            "product_name": item.product.name,
            "size": item.product.size,
            "qty": f"{item.qty:.3f}",
            "unit_price": f"{item.unit_price:.2f}",
        }
        for item in order.items.select_related("product")
    ]

    return render(
        request,
        "core/order_create.html",
        _order_form_context(request, form, items=initial_items, is_edit=True, order=order),
    )


@require_POST
@super_admin_required
def order_delete(request, pk):
    order = get_object_or_404(Order, pk=pk)
    ref = order.reference_no
    with transaction.atomic():
        order.delete()  # cascades to OrderItem
    messages.success(request, f"Quotation {ref} deleted.")
    return redirect("core:order_list")


@login_required
def order_detail(request, pk):
    order = get_object_or_404(
        Order.objects.select_related("customer", "created_by"), pk=pk
    )
    items = order.items.select_related("product")
    return render(
        request,
        "core/order_detail.html",
        {
            "order": order,
            "items": items,
        },
    )


@require_POST
@login_required
def order_set_status(request, pk, status):
    order = get_object_or_404(Order, pk=pk)
    valid = {v for v, _ in Order.Status.choices}
    if status not in valid:
        messages.error(request, "Unknown status.")
        return redirect("core:order_detail", pk=pk)
    with transaction.atomic():
        order.status = status
        order.save(update_fields=["status"])
    messages.success(request, f"Quotation marked as {order.get_status_display()}.")
    return redirect("core:order_detail", pk=pk)


def _order_pdf_context(order):
    return {
        "order": order,
        "items": list(order.items.select_related("product")),
        "generated_at": timezone.localtime(),
    }


@login_required
def order_pdf(request, pk):
    order = get_object_or_404(
        Order.objects.select_related("customer"), pk=pk
    )
    return _pdf_response(
        request,
        "core/order_pdf.html",
        _order_pdf_context(order),
        f"quotation_{order.reference_no}.pdf",
    )


@login_required
def order_excel(request, pk):
    """Same content as the PDF, in .xlsx form. openpyxl is a pure-Python
    dependency, so unlike WeasyPrint this always works even on Windows
    without GTK."""
    from io import BytesIO

    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    order = get_object_or_404(
        Order.objects.select_related("customer"), pk=pk
    )
    items = list(order.items.select_related("product"))

    wb = Workbook()
    ws = wb.active
    ws.title = order.reference_no

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1F2937")
    right = Alignment(horizontal="right")

    # Masthead
    ws["A1"] = "Senovka Plastics — Quotation"
    ws["A1"].font = Font(bold=True, size=14)
    ws.merge_cells("A1:F1")

    ws["A2"] = "Ref No"; ws["B2"] = order.reference_no
    ws["C2"] = "Date"; ws["D2"] = order.order_date.strftime("%d %b %Y")
    ws["E2"] = "Status"; ws["F2"] = order.get_status_display()

    ws["A3"] = "Customer"; ws["B3"] = order.display_customer
    if order.valid_until:
        ws["C3"] = "Valid until"
        ws["D3"] = order.valid_until.strftime("%d %b %Y")

    for row in (2, 3):
        for col in ("A", "C", "E"):
            cell = ws[f"{col}{row}"]
            if cell.value:
                cell.font = Font(bold=True)

    # Items header
    HEADERS = ["No", "Product", "Size", "Qty", "Unit Price", "Line Total"]
    header_row = 5
    for idx, name in enumerate(HEADERS, start=1):
        cell = ws.cell(row=header_row, column=idx, value=name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # Items rows
    row_num = header_row + 1
    for i, item in enumerate(items, start=1):
        ws.cell(row=row_num, column=1, value=i)
        ws.cell(row=row_num, column=2, value=item.product.name)
        ws.cell(row=row_num, column=3, value=item.product.size or "—")
        c_qty = ws.cell(row=row_num, column=4, value=float(item.qty))
        c_price = ws.cell(row=row_num, column=5, value=float(item.unit_price))
        c_total = ws.cell(row=row_num, column=6, value=float(item.line_total))
        c_qty.alignment = right
        c_price.alignment = right
        c_total.alignment = right
        c_price.number_format = "#,##0.00"
        c_total.number_format = "#,##0.00"
        c_qty.number_format = "#,##0.000"
        row_num += 1

    # Totals
    row_num += 1

    def totals_row(label, value, bold=False):
        nonlocal row_num
        ws.cell(row=row_num, column=5, value=label).alignment = right
        c = ws.cell(row=row_num, column=6, value=float(value))
        c.alignment = right
        c.number_format = "#,##0.00"
        if bold:
            ws.cell(row=row_num, column=5).font = Font(bold=True)
            c.font = Font(bold=True)
        row_num += 1

    totals_row("Subtotal", order.subtotal)
    if order.delivery_charge:
        totals_row("Delivery", order.delivery_charge)
    if order.discount_amount:
        totals_row("Discount", -order.discount_amount)
    totals_row("Grand total", order.total_amount, bold=True)

    if order.notes:
        row_num += 1
        ws.cell(row=row_num, column=1, value="Notes").font = Font(bold=True)
        ws.cell(row=row_num, column=2, value=order.notes)
        ws.merge_cells(
            start_row=row_num, start_column=2,
            end_row=row_num, end_column=6,
        )

    # Column widths — a rough auto-size based on content length in each column.
    widths = {"A": 5, "B": 32, "C": 12, "D": 10, "E": 14, "F": 14}
    for letter, width in widths.items():
        ws.column_dimensions[letter].width = width

    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    response = HttpResponse(
        stream.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = (
        f'attachment; filename="quotation_{order.reference_no}.xlsx"'
    )
    return response


@login_required
def order_delivery_note_excel(request, pk):
    """Delivery Note for an order, laid out to the company's A4 template.

    The layout is not incidental — it mirrors `Delivery_Note_Template_A4.xlsx`
    cell for cell so a printed note from this system is indistinguishable from
    the one the office already uses. Column widths, row heights, merge ranges,
    fills and the 67% fit-to-page scale are all taken from that file; changing
    any of them changes what comes out of the printer, so they are written as
    explicit constants rather than "whatever looks right".

    Three columns are deliberately left empty for the delivery run to fill in
    by hand — vehicle/driver/helper, the "Checked at Customer's Place" column,
    and the payment block. Those are facts nobody knows until the lorry is
    loaded and the customer has signed, and pre-filling them with a guess is
    worse than leaving a ruled box.
    """
    import os
    import re
    from io import BytesIO

    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.worksheet.properties import PageSetupProperties

    order = get_object_or_404(Order.objects.select_related("customer"), pk=pk)
    items = list(order.items.select_related("product"))
    billing = BillingSettings.load()

    wb = Workbook()
    ws = wb.active
    ws.title = "Delivery Note"

    # ── Palette, straight off the template ───────────────────────────────
    NAVY = "FF1F4E79"        # headings, section bars, table header
    BAND = "FFD9E1F2"        # pale blue label fill
    GREY = "FFF2F2F2"        # grey label fill
    INK = "FF262626"         # label text
    MUTED = "FF404040"       # company address / small print
    FAINT = "FF595959"       # signature captions, disclaimer

    thin = Side(style="thin")
    medium = Side(style="medium")
    dashed = Side(style="dashed")
    box = Border(top=thin, bottom=thin, left=thin, right=thin)

    def F(size, bold=False, italic=False, color=None):
        """Arial at `size`. The template uses Arial throughout; Calibri only
        survives on cells that were never touched."""
        return Font(name="Arial", size=size, bold=bold, italic=italic, color=color)

    def fill(rgb):
        return PatternFill("solid", fgColor=rgb)

    def put(coord, value=None, font=None, align=None, bg=None, border=box,
            numfmt=None):
        cell = ws[coord]
        if value is not None:
            cell.value = value
        if font:
            cell.font = font
        if align:
            cell.alignment = align
        if bg:
            cell.fill = fill(bg)
        if border is not None:
            cell.border = border
        if numfmt:
            cell.number_format = numfmt
        return cell

    L = Alignment(horizontal="left", vertical="center")
    C = Alignment(horizontal="center", vertical="center")
    R = Alignment(horizontal="right", vertical="center")
    CW = Alignment(horizontal="center", vertical="center", wrap_text=True)
    LW = Alignment(horizontal="left", vertical="center", wrap_text=True)
    LT = Alignment(horizontal="left", vertical="top", wrap_text=True)

    QTY_FMT = "#,##0;\\(#,##0\\);\\-"

    # ── Geometry ─────────────────────────────────────────────────────────
    for col, width in (
        ("A", 6.0), ("B", 14.0), ("C", 30.0),
        ("D", 28.5546875), ("E", 23.0), ("F", 39.44140625),
    ):
        ws.column_dimensions[col].width = width

    # The template reserves 17 item rows (15–31). Honour that so a short note
    # prints with the same ruled box the office expects; grow only when an
    # order genuinely has more lines, because truncating an order is worse
    # than a note that runs a little longer.
    ITEM_ROWS = max(17, len(items))
    FIRST_ITEM = 15
    LAST_ITEM = FIRST_ITEM + ITEM_ROWS - 1
    R_TOTAL = LAST_ITEM + 1
    R_REMARK = R_TOTAL + 1
    R_GAP2 = R_REMARK + 1
    R_PAYHEAD = R_GAP2 + 1
    R_PAYROW1 = R_PAYHEAD + 1     # "Payment Method" / CHEQUE NUMBER / CASH / CHEQUE
    R_CASH = R_PAYROW1 + 1        # CASH row
    R_CHQ1 = R_CASH + 1           # first of five CHEQUE NO rows
    R_CHQ_LAST = R_CHQ1 + 4
    R_GAP3 = R_CHQ_LAST + 1
    R_SIGHEAD = R_GAP3 + 1
    R_SIGBOX = R_SIGHEAD + 1
    R_SIGCAP = R_SIGBOX + 1
    R_FOOTER = R_SIGCAP + 1

    heights = {1: 21.75, 2: 21.75, 3: 21.75, 4: 21.75, 5: 6.0, 6: 18.0}
    for r in range(7, 13):
        heights[r] = 18.75
    heights[13] = 6.0
    heights[14] = 30.0
    for r in range(FIRST_ITEM, LAST_ITEM + 1):
        heights[r] = 18.75
    heights[R_TOTAL] = 19.5
    heights[R_REMARK] = 24.0
    heights[R_GAP2] = 6.0
    heights[R_PAYHEAD] = 18.0
    heights[R_PAYROW1] = 18.0
    heights[R_CASH] = 18.0
    for r in range(R_CHQ1, R_CHQ_LAST + 1):
        heights[r] = 18.75
    heights[R_GAP3] = 6.0
    heights[R_SIGHEAD] = 15.75
    heights[R_SIGBOX] = 64.2
    heights[R_SIGCAP] = 13.5
    heights[R_FOOTER] = 24.0
    for row, h in heights.items():
        ws.row_dimensions[row].height = h

    # ── Rows 1–4 · masthead ──────────────────────────────────────────────
    # A1:B4 is the logo well. The dashed inner edges are the template's own
    # "drop the logo here" cue; we drop the real logo in and keep the frame.
    ws.merge_cells("A1:B4")
    put("A1", None, F(9, italic=True, color="FFA6A6A6"), CW, "FFFFFFFF",
        Border(top=thin, bottom=dashed, left=thin, right=dashed))
    put("B1", None, border=Border(top=thin, right=dashed))
    put("A2", None, border=Border(left=thin))
    put("B2", None, border=Border(right=dashed))
    put("A3", None, border=Border(left=thin))
    put("B3", None, border=Border(right=dashed))
    put("A4", None, border=Border(bottom=dashed, left=thin))
    put("B4", None, border=Border(bottom=dashed, right=dashed))

    ws.merge_cells("C1:D2")
    put("C1", billing.company_name.upper(), F(17, bold=True, color=NAVY), L,
        border=Border(top=thin))
    put("D1", None, border=Border(top=thin))

    ws.merge_cells("E1:F1")
    put("E1", "DELIVERY NOTE", F(18, bold=True, color="FFFFFFFF"), C, NAVY,
        Border(top=thin, right=medium))
    put("F1", None, border=Border(top=thin, right=medium))

    # Address / contact strip. Both lines fall back to a blank rather than a
    # placeholder — an unset field should print as nothing, not as "None".
    ws.merge_cells("C3:D3")
    address_line = " ".join((billing.address or "").split())
    put("C3", address_line, F(9, color=MUTED), L, border=None)
    ws.merge_cells("C4:D4")
    contact_bits = []
    if billing.phone:
        contact_bits.append(f"Tel: {billing.phone}")
    if billing.email:
        contact_bits.append(f"Email: {billing.email}")
    put("C4", "  |  ".join(contact_bits), F(9, color=MUTED), L, border=None)

    # Document reference block, right-hand side.
    # The delivery-note number rides on the order's own reference so the two
    # documents can always be matched up by eye: ORD-0008 -> DN-0008.
    dn_no = re.sub(r"^ORD-", "DN-", order.reference_no or "") or f"DN-{order.pk:04d}"
    today = timezone.localdate()
    for row, label, value in (
        (2, "D/N No", dn_no),
        (3, "Date", today.strftime("%d/%m/%Y")),
        (4, "Order No", order.reference_no or ""),
    ):
        put(f"E{row}", label, F(9, bold=True, color=INK), L, BAND)
        put(f"F{row}", value, F(10, bold=True, color=NAVY), L)

    put("A5", None, border=Border(left=thin))
    put("F5", None, border=Border(right=thin))

    # ── Row 6 · section bars ─────────────────────────────────────────────
    ws.merge_cells("A6:C6")
    put("A6", "CUSTOMER DETAILS", F(9, bold=True, color=NAVY), L, BAND)
    put("B6", None, border=Border(top=thin, bottom=thin))
    put("C6", None, border=Border(top=thin, bottom=thin, right=thin))
    ws.merge_cells("D6:F6")
    put("D6", "DELIVERY & TRANSPORT DETAILS", F(9, bold=True, color=NAVY), L,
        BAND, Border(top=thin, bottom=thin, left=thin, right=medium))
    put("E6", None, border=Border(top=thin, bottom=thin))
    put("F6", None, border=Border(top=thin, bottom=thin, right=medium))

    # ── Rows 7–12 · customer (left) and transport (right) ────────────────
    # Address is split across two rows the way the template does it, so a long
    # address wraps into the second line instead of overflowing the column.
    raw_address = (order.customer.address or "") if order.customer_id else ""
    addr_lines = [ln.strip() for ln in raw_address.splitlines() if ln.strip()]
    addr_1 = addr_lines[0] if addr_lines else ""
    addr_2 = ", ".join(addr_lines[1:]) if len(addr_lines) > 1 else ""
    phone = (order.customer.phone or "") if order.customer_id else ""

    ws.merge_cells("A7:B7")
    put("A7", "Customer Name", F(9, bold=True, color=INK), L, GREY)
    put("B7", None, border=Border(top=thin, bottom=thin, right=thin))
    put("C7", order.display_customer, F(10, color="FF000000"), L)

    ws.merge_cells("A8:B9")
    put("A8", "Customer Address", F(9, bold=True, color=INK), L, GREY)
    put("B8", None, border=Border(top=thin, right=thin))
    put("A9", None, border=Border(bottom=thin, left=thin))
    put("B9", None, border=Border(bottom=thin, right=thin))
    put("C8", addr_1, F(10, color="FF000000"), L)
    put("C9", addr_2, F(10, color="FF000000"), L)

    ws.merge_cells("A10:B10")
    put("A10", "Contact No", F(9, bold=True, color=INK), L, GREY)
    put("B10", None, border=Border(top=thin, bottom=thin, right=thin))
    put("C10", phone, F(10, color="FF000000"), L)

    # A11/A12 sit under the customer block with no label — the template keeps
    # them ruled and filled so the left and right columns end level.
    for row in (11, 12):
        put(f"A{row}", None, F(9, bold=True, color=INK),
            Alignment(vertical="center"), GREY,
            Border(top=thin, bottom=thin, left=thin))
        put(f"B{row}", None, F(9, bold=True, color=INK),
            Alignment(vertical="center"), GREY,
            Border(top=thin, bottom=thin, right=thin))
        put(f"C{row}", None, F(10, color="FF000000"), L)

    # Right column. Invoice/vehicle/driver/helper are blank by design — see
    # the docstring.
    transport = (
        ("Invoice No", "", medium),
        ("Order Date", order.order_date.strftime("%d/%m/%Y"), medium),
        ("Delivery Date", today.strftime("%d/%m/%Y"), medium),
        ("Vehicle No", "", medium),
        ("Driver Name", "", thin),
        ("Helper Name", "", thin),
    )
    for offset, (label, value, right_edge) in enumerate(transport):
        row = 7 + offset
        put(f"D{row}", label, F(9, bold=True, color=INK), L, GREY)
        ws.merge_cells(f"E{row}:F{row}")
        put(f"E{row}", value, F(10, color="FF000000"), L,
            border=Border(top=thin, bottom=thin, left=thin, right=right_edge))
        put(f"F{row}", None,
            border=Border(top=thin, bottom=thin, right=right_edge))

    put("A13", None, border=Border(left=thin))
    put("F13", None, border=Border(right=thin))

    # ── Row 14 · item table header ───────────────────────────────────────
    headers = [
        "No", "Size", "Item Name / Description",
        "Ordered Qty", "Issued Qty", "Checked at Customer's Place",
    ]
    for idx, title in enumerate(headers, start=1):
        cell = ws.cell(row=14, column=idx, value=title)
        cell.font = F(9, bold=True, color="FFFFFFFF")
        cell.fill = fill(NAVY)
        cell.alignment = CW
        cell.border = box

    # ── Item rows ────────────────────────────────────────────────────────
    # Product names in this system often lead with their own size ("63mm
    # elbow"), which would print the size twice once it has its own column.
    # Strip the prefix, and the trailing owner tag the catalogue carries.
    def split_name(product):
        size = (product.size or "").strip()
        name = product.name
        if size and name.upper().startswith(size.upper()):
            name = name[len(size):].strip(" -–")
        name = re.sub(
            r"\s*-\s*(SENOVKA|KRISHAN|SURESH)$", "", name, flags=re.IGNORECASE
        ).strip()
        return size, (name or product.name)

    for offset in range(ITEM_ROWS):
        row = FIRST_ITEM + offset
        item = items[offset] if offset < len(items) else None

        if item is not None:
            size, name = split_name(item.product)
            qty = item.qty
            # Quantities are whole units on a delivery note; a fractional
            # count would be a data-entry slip, so render what is stored but
            # drop a meaningless ".000".
            qty_out = int(qty) if qty == qty.to_integral_value() else float(qty)
            ws.cell(row=row, column=1, value=offset + 1)
            ws.cell(row=row, column=2, value=size)
            ws.cell(row=row, column=3, value=name)
            ws.cell(row=row, column=4, value=qty_out)
            # Issued defaults to ordered: the note goes out with the intent to
            # ship the full line, and the store amends by hand when short.
            ws.cell(row=row, column=5, value=qty_out)

        for col in range(1, 7):
            cell = ws.cell(row=row, column=col)
            cell.font = F(10)
            cell.alignment = L if col == 3 else C
            cell.border = box
            if col in (4, 5):
                cell.number_format = QTY_FMT

    # ── Total row ────────────────────────────────────────────────────────
    ws.merge_cells(f"A{R_TOTAL}:C{R_TOTAL}")
    put(f"A{R_TOTAL}", "TOTAL QUANTITY", F(10, bold=True, color=NAVY), R, BAND)
    put(f"B{R_TOTAL}", None, border=Border(top=thin, bottom=thin))
    put(f"C{R_TOTAL}", None, border=Border(top=thin, bottom=thin, right=thin))
    # Live formulas, not baked numbers: the store often crosses out an issued
    # quantity on the printed sheet and re-keys it, and the total should move
    # with it when the file is reopened.
    put(f"D{R_TOTAL}", f"=SUM(D{FIRST_ITEM}:D{LAST_ITEM})",
        F(10, bold=True, color=NAVY), C, BAND, numfmt=QTY_FMT)
    put(f"E{R_TOTAL}", f"=SUM(E{FIRST_ITEM}:E{LAST_ITEM})",
        F(10, bold=True, color=NAVY), C, BAND, numfmt=QTY_FMT)
    put(f"F{R_TOTAL}", None, bg=BAND)

    # ── Remarks ──────────────────────────────────────────────────────────
    ws.merge_cells(f"A{R_REMARK}:B{R_REMARK}")
    put(f"A{R_REMARK}", "Remarks / Special Instructions",
        F(9, bold=True, color=INK), LW, GREY)
    put(f"B{R_REMARK}", None, border=Border(top=thin, bottom=thin, right=thin))
    ws.merge_cells(f"C{R_REMARK}:F{R_REMARK}")
    put(f"C{R_REMARK}", (order.notes or "").strip(), F(10, color="FF000000"),
        LT, border=Border(top=thin, bottom=thin, left=thin, right=medium))
    for col in ("D", "E"):
        put(f"{col}{R_REMARK}", None, border=Border(top=thin, bottom=thin))
    put(f"F{R_REMARK}", None, border=Border(top=thin, bottom=thin, right=medium))

    put(f"A{R_GAP2}", None, border=Border(left=thin))
    put(f"F{R_GAP2}", None, border=Border(right=thin))

    # ── Payment / goods-received bars ────────────────────────────────────
    ws.merge_cells(f"A{R_PAYHEAD}:C{R_PAYHEAD}")
    put(f"A{R_PAYHEAD}", "PAYMENT DETAILS", F(9, bold=True, color=NAVY), L, BAND)
    put(f"B{R_PAYHEAD}", None, border=Border(top=thin, bottom=thin))
    put(f"C{R_PAYHEAD}", None, border=Border(top=thin, bottom=thin, right=thin))
    ws.merge_cells(f"D{R_PAYHEAD}:F{R_PAYHEAD}")
    put(f"D{R_PAYHEAD}", "GOODS RECEIVED CONFIRMATION",
        F(9, bold=True, color=NAVY), L, BAND,
        Border(top=thin, bottom=thin, left=thin, right=medium))
    put(f"E{R_PAYHEAD}", None, border=Border(top=thin, bottom=thin))
    put(f"F{R_PAYHEAD}", None, border=Border(top=thin, bottom=thin, right=medium))

    ws.merge_cells(f"A{R_PAYROW1}:B{R_PAYROW1}")
    put(f"A{R_PAYROW1}", "Payment Method", F(9, bold=True, color=INK), L, GREY)
    put(f"B{R_PAYROW1}", None, border=Border(top=thin, bottom=thin, right=thin))
    put(f"C{R_PAYROW1}", "CHEQUE NUMBER", F(9, bold=True, color=NAVY), L, BAND,
        Border(top=thin, bottom=thin, left=thin, right=medium))
    put(f"D{R_PAYROW1}", "CASH", F(9, bold=True, color=NAVY), L, BAND,
        Border(top=thin, bottom=thin, left=thin, right=medium))
    put(f"E{R_PAYROW1}", "CHEQUE", F(9, bold=True, color=NAVY), L, BAND,
        Border(top=thin, bottom=thin, left=thin))
    put(f"F{R_PAYROW1}", None, border=Border(top=thin, left=thin, right=thin))

    ws.merge_cells(f"A{R_CASH}:B{R_CASH}")
    put(f"A{R_CASH}", "CASH", F(9, bold=True, color=INK), C, GREY)
    put(f"B{R_CASH}", None, border=Border(top=thin, bottom=thin, right=thin))
    put(f"C{R_CASH}", None, F(9, bold=True, color=NAVY), L, BAND,
        Border(top=thin, bottom=thin, left=thin))
    put(f"D{R_CASH}", None, F(9, bold=True, color=NAVY), L, BAND,
        Border(top=thin, bottom=thin, left=thin, right=medium))
    put(f"E{R_CASH}", None, F(9, bold=True, color=NAVY), L, BAND,
        Border(top=thin, bottom=thin, left=thin))
    put(f"F{R_CASH}", None, border=Border(bottom=thin, left=thin, right=thin))

    for row in range(R_CHQ1, R_CHQ_LAST + 1):
        is_first = row == R_CHQ1
        is_last = row == R_CHQ_LAST
        ws.merge_cells(f"A{row}:B{row}")
        # Calibri here, not Arial: the template left these label cells on the
        # workbook default, and matching it keeps the column visually lighter
        # than the Arial labels above it.
        put(f"A{row}", "CHEQUE NO", Font(name="Calibri", size=11),
            Alignment(horizontal="center"))
        put(f"B{row}", None, border=Border(top=thin, bottom=thin, right=thin))
        # The last row's cheque-number cell steps up to 10pt in the template.
        put(f"C{row}", None, F(10 if is_last else 9), L)
        # F on the first row has no top edge — it butts against the CASH row
        # above, which already closed itself with a bottom border.
        f_top = None if is_first else thin
        put(f"D{row}", None, F(8, color=MUTED), LW,
            border=Border(top=thin, bottom=thin, left=thin, right=medium))
        put(f"E{row}", None, F(8, color=MUTED), LW,
            border=Border(top=thin, bottom=thin, left=thin, right=medium))
        put(f"F{row}", None, F(8, color=MUTED), LW,
            border=Border(top=f_top, bottom=thin, left=thin, right=thin))

    put(f"A{R_GAP3}", None, border=Border(left=thin))
    put(f"F{R_GAP3}", None, border=Border(right=thin))

    # ── Signature strip ──────────────────────────────────────────────────
    for start_col, label, right_edge in (
        ("A", "ISSUED BY (Stores)", thin),
        ("C", "DELIVERED BY (Driver)", thin),
        ("E", "RECEIVED BY (Customer)", medium),
    ):
        end_col = chr(ord(start_col) + 1)
        ws.merge_cells(f"{start_col}{R_SIGHEAD}:{end_col}{R_SIGHEAD}")
        put(f"{start_col}{R_SIGHEAD}", label, F(8, bold=True, color=NAVY), C,
            BAND, Border(top=thin, bottom=thin, left=thin, right=right_edge))
        put(f"{end_col}{R_SIGHEAD}", None,
            border=Border(top=thin, bottom=thin, right=right_edge))

        ws.merge_cells(f"{start_col}{R_SIGBOX}:{end_col}{R_SIGBOX}")
        put(f"{start_col}{R_SIGBOX}", None,
            border=Border(top=thin, bottom=thin, left=thin, right=right_edge))
        put(f"{end_col}{R_SIGBOX}", None,
            border=Border(top=thin, bottom=thin, right=right_edge))

    for start_col, caption, edge in (
        ("A", "Name  /  Date", Border(left=thin)),
        ("C", "Name  /  Date", None),
        ("E", "Signature with Rubber Stamp", Border(right=medium)),
    ):
        end_col = chr(ord(start_col) + 1)
        ws.merge_cells(f"{start_col}{R_SIGCAP}:{end_col}{R_SIGCAP}")
        put(f"{start_col}{R_SIGCAP}", caption, F(8, italic=True, color=FAINT), C,
            border=edge)
        put(f"{end_col}{R_SIGCAP}", None,
            border=Border(right=medium) if start_col == "E" else None)

    # ── Disclaimer ───────────────────────────────────────────────────────
    ws.merge_cells(f"A{R_FOOTER}:F{R_FOOTER}")
    put(
        f"A{R_FOOTER}",
        "Goods must be checked at the time of delivery. Claims for shortages "
        "or damages will not be entertained after the customer has signed "
        "this note. This is a delivery note only and is not a receipt for "
        "payment.",
        F(8, color=FAINT), CW,
        border=Border(bottom=thin, left=thin, right=medium),
    )
    for col in ("B", "C", "D", "E"):
        put(f"{col}{R_FOOTER}", None, border=Border(bottom=thin))
    put(f"F{R_FOOTER}", None, border=Border(bottom=thin, right=medium))

    # ── Logo into the well at A1:B4 ──────────────────────────────────────
    # Sized to the merged box (A+B ≈ 145px wide, four 21.75pt rows ≈ 116px)
    # so it sits inside the dashed frame instead of spilling over the company
    # name. A missing or unreadable file must never break the download — the
    # note is still valid without the mark.
    logo_path = os.path.join(
        os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
        "static", "images", "logo.jpeg",
    )
    if os.path.exists(logo_path):
        try:
            from openpyxl.drawing.image import Image as XLImage
            from PIL import Image as PILImage

            pil = PILImage.open(logo_path).convert("RGB")
            pil.thumbnail((138, 108), PILImage.LANCZOS)
            buf = BytesIO()
            pil.save(buf, format="PNG")
            buf.seek(0)

            img = XLImage(buf)
            img.width, img.height = pil.size
            img.anchor = "A1"
            ws.add_image(img)
        except Exception:
            pass

    # ── Page setup ───────────────────────────────────────────────────────
    # 67% on A4 portrait is what makes the six columns land inside the
    # printable width; fitToPage without a scale lets Excel pick its own and
    # the note comes out at a different size on every machine.
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    ws.page_setup.orientation = "portrait"
    ws.page_setup.paperSize = 9  # A4
    ws.page_setup.scale = 67
    ws.page_margins.left = 0.4
    ws.page_margins.right = 0.4
    ws.page_margins.top = 0.4
    ws.page_margins.bottom = 0.4
    ws.page_margins.header = 0.511811023622047
    ws.page_margins.footer = 0.511811023622047
    ws.print_area = f"A1:F{R_FOOTER}"
    ws.sheet_view.showGridLines = False

    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    response = HttpResponse(
        stream.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = (
        f'attachment; filename="delivery_note_{order.reference_no or order.pk}.xlsx"'
    )
    return response


@login_required
def order_production_check(request):
    """Aggregate ordered quantities per product across active orders and
    compare against current stock. The point is a single view where the
    operator can spot every product that needs production.

    Order status scope:
      sent      — default. Something has actually gone out to the customer.
      confirmed — always included.
      draft     — opt-in via ?scope=all. Drafts are still figures on paper.
      cancelled — never included; a cancelled order isn't owed anything.

    Filters:
      scope=sent|all|shortages   — as above; 'shortages' shows only rows
                                    that need production.
      q=<text>                    — product name filter.
    """
    scope = request.GET.get("scope", "sent").strip()
    if scope not in {"sent", "all", "shortages"}:
        scope = "sent"

    if scope == "all":
        active_statuses = [
            Order.Status.DRAFT, Order.Status.SENT, Order.Status.CONFIRMED,
        ]
    else:
        # 'sent' and 'shortages' both start from the same base — the
        # shortages filter is applied below on the aggregate.
        active_statuses = [Order.Status.SENT, Order.Status.CONFIRMED]

    query = request.GET.get("q", "").strip()

    # One aggregate query: total qty ordered per product across the picked
    # order statuses.
    ordered = (
        OrderItem.objects.filter(order__status__in=active_statuses)
        .values("product_id", "product__name", "product__size", "product__qty")
        .annotate(total_ordered=Coalesce(
            Sum("qty"),
            Decimal("0.000"),
            output_field=DecimalField(max_digits=12, decimal_places=3),
        ))
        .order_by("product__name", "product__size")
    )
    if query:
        ordered = ordered.filter(product__name__icontains=query)

    rows = []
    total_shortage = Decimal("0.000")
    total_ordered_all = Decimal("0.000")
    for row in ordered:
        stock = row["product__qty"] or Decimal("0.000")
        total_ordered = row["total_ordered"] or Decimal("0.000")
        shortage = max(total_ordered - stock, Decimal("0.000"))

        if stock <= 0:
            status = "out_of_stock"
        elif shortage > 0:
            status = "short"
        else:
            status = "sufficient"

        rows.append({
            "product_id": row["product_id"],
            "name": row["product__name"],
            "size": row["product__size"] or "",
            "stock": stock,
            "total_ordered": total_ordered,
            "shortage": shortage,
            "status": status,
        })
        total_ordered_all += total_ordered
        total_shortage += shortage

    if scope == "shortages":
        rows = [r for r in rows if r["shortage"] > 0 or r["status"] == "out_of_stock"]

    # For the per-row detail modal — which orders is this product in? Kept
    # to a small dict rather than fetched per click so the page has no
    # extra round-trips.
    orders_by_product = {}
    if rows:
        involved = OrderItem.objects.filter(
            order__status__in=active_statuses,
            product_id__in=[r["product_id"] for r in rows],
        ).select_related("order").order_by("-order__order_date")
        for oi in involved:
            orders_by_product.setdefault(oi.product_id, []).append({
                "ref": oi.order.reference_no,
                "pk": oi.order_id,
                "qty": oi.qty,
                "customer": oi.order.display_customer,
                "date": oi.order.order_date.strftime("%d %b %Y"),
                "status": oi.order.get_status_display(),
            })

    return render(
        request,
        "core/order_production_check.html",
        {
            "rows": rows,
            "scope": scope,
            "query": query,
            "orders_json": json.dumps(orders_by_product, default=str),
            "totals": {
                "row_count": len(rows),
                "total_ordered": total_ordered_all,
                "total_shortage": total_shortage,
                "shortage_row_count": sum(1 for r in rows if r["shortage"] > 0),
            },
        },
    )


# =========================================================== daily machine run
# The floor log: which machines ran today, who operated them, what they
# were making. Plus a single "other works" row per day for driver,
# material supply, material mixing, and anything else worth noting.


# ---- Machines master (super-admin CRUD) ----


@super_admin_required
def machine_list(request):
    machines = (
        Machine.objects.annotate(run_count=Count("daily_runs", distinct=True))
        .order_by("name")
    )
    page_obj = _paginate(request, machines)
    return render(
        request,
        "core/machine_list.html",
        {
            "page_obj": page_obj,
            "machines": page_obj.object_list,
            "form": MachineForm(),
        },
    )


@require_POST
@super_admin_required
def machine_create(request):
    form = MachineForm(request.POST)
    if not form.is_valid():
        messages.error(request, f"Machine not saved: {form.first_error()}")
    else:
        with transaction.atomic():
            machine = form.save()
        messages.success(request, f"Machine '{machine.name}' added.")
    return redirect("core:machine_list")


@require_POST
@super_admin_required
def machine_edit(request, pk):
    machine = get_object_or_404(Machine, pk=pk)
    form = MachineForm(request.POST, instance=machine)
    if not form.is_valid():
        messages.error(request, f"Machine not saved: {form.first_error()}")
    else:
        with transaction.atomic():
            form.save()
        messages.success(request, "Machine updated.")
    return redirect("core:machine_list")


@require_POST
@super_admin_required
def machine_delete(request, pk):
    machine = get_object_or_404(
        Machine.objects.annotate(run_count=Count("daily_runs", distinct=True)),
        pk=pk,
    )
    if machine.run_count:
        messages.error(
            request,
            f"Cannot delete '{machine.name}' — {machine.run_count} "
            f"daily entr{'y' if machine.run_count == 1 else 'ies'} still "
            f"reference it. Deactivate instead.",
        )
        return redirect("core:machine_list")
    try:
        machine.delete()
        messages.success(request, f"Machine '{machine.name}' deleted.")
    except ProtectedError:
        messages.error(request, f"Cannot delete '{machine.name}' — other records reference it.")
    return redirect("core:machine_list")


# ---- The daily run page ----


def _daily_run_date(request):
    """The date this daily-run page is looking at.

    Bookmarked and hand-edited URLs land here too, so garbage should
    default to today rather than 500 — same shape as _parse_date and
    get_month_filter.
    """
    raw = (request.GET.get("date") or request.POST.get("date") or "").strip()
    parsed = _parse_date(raw)
    if parsed is None:
        return timezone.localdate()
    return parsed


@login_required
def daily_run(request):
    """The floor log for one day.

    GET renders the form pre-populated with whatever has been logged for
    the picked date. POST processes every machine row + the other-works
    row in one transaction — if any row fails, none of them save, and the
    page re-renders with the fresh POST values so nothing is lost.
    """
    when = _daily_run_date(request)
    if when > timezone.localdate():
        messages.warning(request, "Future dates land on today.")
        when = timezone.localdate()

    machines = Machine.objects.filter(is_active=True).order_by("name")
    existing = {
        r.machine_id: r
        for r in DailyMachineRun.objects.filter(run_date=when).select_related("machine")
    }
    products = Product.objects.filter(is_active=True).order_by("name", "size")

    other = DailyOtherWork.objects.filter(run_date=when).first()

    if request.method == "POST":
        errors = []
        with transaction.atomic():
            # Machines: one form-set of hidden fields per machine, plus
            # status/operator/product/notes. Machine rows the operator did
            # not touch (all fields blank + status default) are ignored so
            # they do not clutter the audit log.
            for machine in machines:
                prefix = f"m-{machine.pk}"
                status = (request.POST.get(f"{prefix}-status") or "").strip()
                operator = (request.POST.get(f"{prefix}-operator") or "").strip()[:150]
                product_id = (request.POST.get(f"{prefix}-product") or "").strip()
                notes = (request.POST.get(f"{prefix}-notes") or "").strip()[:500]

                if status not in {
                    DailyMachineRun.Status.RUNNING,
                    DailyMachineRun.Status.NOT_WORKING,
                }:
                    # The picker only offers these two, so anything else is
                    # a bookmark from before the choices were fixed. Fall
                    # back to blank rather than 500 so the rest of the
                    # page's data can still save.
                    status = DailyMachineRun.Status.RUNNING

                # Skip rows the operator plainly did not fill in — they
                # would otherwise write a "running with no operator"
                # blank row every time the page is saved.
                already = existing.get(machine.pk)
                if (
                    already is None
                    and status == DailyMachineRun.Status.RUNNING
                    and not operator and not product_id and not notes
                ):
                    continue

                product = None
                if status == DailyMachineRun.Status.RUNNING and product_id.isdigit():
                    product = Product.objects.filter(
                        pk=int(product_id), is_active=True
                    ).first()

                if status == DailyMachineRun.Status.RUNNING and not operator:
                    errors.append(
                        f"{machine.name}: enter an operator or mark the machine as not working."
                    )
                    continue

                run, _ = DailyMachineRun.objects.update_or_create(
                    run_date=when,
                    machine=machine,
                    defaults={
                        "status": status,
                        "operator": operator if status == DailyMachineRun.Status.RUNNING else "",
                        "product": product if status == DailyMachineRun.Status.RUNNING else None,
                        "notes": notes,
                        "logged_by": request.user,
                    },
                )

            # Other works: single row per day. update_or_create is idempotent
            # under the unique(run_date), so re-saving the page is safe.
            other_form = DailyOtherWorkForm(
                request.POST, instance=other or DailyOtherWork(run_date=when)
            )
            if other_form.is_valid():
                other_row = other_form.save(commit=False)
                other_row.run_date = when
                other_row.logged_by = request.user
                if other:
                    other_row.edit_date = timezone.localdate()
                other_row.save()
            else:
                errors.append(f"Other works: {other_form.first_error()}")

            if errors:
                # Roll back everything — a half-saved daily log confuses the
                # accounting more than a re-typing.
                transaction.set_rollback(True)

        if errors:
            for msg in errors:
                messages.error(request, msg)
        else:
            messages.success(
                request,
                f"Daily log for {when:%d %b %Y} saved.",
            )
        return redirect(f"{reverse('core:daily_run')}?date={when.isoformat()}")

    # GET — hand the template a per-machine snapshot of what is on record.
    machine_rows = []
    for machine in machines:
        run = existing.get(machine.pk)
        machine_rows.append({
            "machine": machine,
            "run": run,
            "status": run.status if run else DailyMachineRun.Status.RUNNING,
            "operator": run.operator if run else "",
            "product_id": run.product_id if run and run.product_id else "",
            "notes": run.notes if run else "",
        })

    return render(
        request,
        "core/daily_run.html",
        {
            "when": when,
            "yesterday": when - timedelta(days=1),
            "tomorrow": when + timedelta(days=1),
            "is_today": when == timezone.localdate(),
            "machine_rows": machine_rows,
            "products": products,
            "other": other,
            "other_form": DailyOtherWorkForm(instance=other) if other else DailyOtherWorkForm(),
            "any_active_machines": bool(machines),
        },
    )


@login_required
def daily_run_history(request):
    """Every date that has anything logged — machine runs or other works —
    paginated. Read as a jump list into the per-day page."""
    machine_dates = set(
        DailyMachineRun.objects.values_list("run_date", flat=True).distinct()
    )
    other_dates = set(
        DailyOtherWork.objects.values_list("run_date", flat=True)
    )
    dates = sorted(machine_dates | other_dates, reverse=True)

    # Cheap counts per date — one query each, indexed on run_date.
    from collections import Counter
    machine_counts = Counter(
        DailyMachineRun.objects.filter(status=DailyMachineRun.Status.RUNNING)
        .values_list("run_date", flat=True)
    )
    not_working_counts = Counter(
        DailyMachineRun.objects.filter(status=DailyMachineRun.Status.NOT_WORKING)
        .values_list("run_date", flat=True)
    )
    has_other = {d for d in other_dates}

    rows = [
        {
            "date": d,
            "running": machine_counts.get(d, 0),
            "not_working": not_working_counts.get(d, 0),
            "has_other": d in has_other,
        }
        for d in dates
    ]
    page_obj = _paginate(request, rows)

    return render(
        request,
        "core/daily_run_history.html",
        {
            "page_obj": page_obj,
            "rows": page_obj.object_list,
        },
    )
